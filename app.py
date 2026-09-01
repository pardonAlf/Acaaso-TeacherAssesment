from flask import Flask, render_template, request, redirect, url_for, jsonify, session
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image
from datetime import datetime
 
from email.message import EmailMessage
import psycopg2
from flask import send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
 
from io import BytesIO
from flask import send_file
from datetime import datetime
from flask import session, redirect, url_for 
import threading
from dotenv import load_dotenv
import os,uuid
import json
import secrets
from flask import make_response

import os
import base64
import qrcode,time
from io import BytesIO
import base64

load_dotenv("llave.env")
try:
    import resend
    resend.api_key = os.getenv("RESEND_API_KEY")

    if not resend.api_key:
        print("❌ RESEND_API_KEY no definida")
        resend = None

except Exception as e:
    print("❌ Error importando resend:", str(e))
    resend = None

from openai import OpenAI
client = OpenAI()

app = Flask(__name__)
app.secret_key = "clave_secreta_super_segura"

from modules.sqlstudio import sqlstudio_bp
app.register_blueprint(sqlstudio_bp)

# Usuario de prueba
USUARIO_TEST = "admin"
PASSWORD_TEST = "1234"
EMAIL_REMITENTE = "pardoalf@gmail.com"
EMAIL_PASSWORD = "pxvr oyrf uhgb xugj"
 
from functools import wraps

def requiere_licencia(func):

    @wraps(func)
    def wrapper(*args, **kwargs):

        # No está logueado
        if 'user_id' not in session:
            return redirect(url_for('login'))

        # ROOT siempre puede entrar
        if session.get('rol') == 'root':
            return func(*args, **kwargs)

        cempre = session.get('cempre')

        if not cempre:
            return redirect(url_for('home'))

        # Consultar la BD directamente.
        # No dependemos de session['licencia_activa'],
        # porque una licencia puede activarse o vencer
        # mientras el usuario ya está logueado.
        conn = get_db_connection()
        cur = conn.cursor()

        try:

            cur.execute("""
                SELECT id
                FROM licencias
                WHERE cempre = %s
                  AND estado = 'ACTIVA'
                  AND fecha_expiracion >= CURRENT_TIMESTAMP
                LIMIT 1
            """, (cempre,))

            licencia = cur.fetchone()

        finally:
            cur.close()
            conn.close()

        # No tiene licencia válida
        if not licencia:
            return redirect(url_for('home'))

        return func(*args, **kwargs)

    return wrapper


@app.context_processor
def inject_version():
    return dict(version=get_version())

def get_version():
    try:
        with open("version.txt") as f:
            return f.read().strip()
    except:
        return "v1.0"
    
@app.route('/splash')
def splash():
    return render_template('splash.html')

@app.route('/')
def home():

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔹 planes individuales
    cur.execute("""
        SELECT
            id,
            nombre,
            precio,
            admins,
            profesores,
            alumnos,
            quizzes,
            periodicidad,
            orden
        FROM planes
        WHERE tipo = 'individual'
          AND activo = TRUE
        ORDER BY orden, id
    """)
    planes_individual = cur.fetchall()

    # 🔹 planes empresariales
    cur.execute("""
        SELECT
            id,
            nombre,
            precio,
            admins,
            profesores,
            alumnos,
            quizzes,
            periodicidad,
            orden
        FROM planes
        WHERE tipo = 'empresarial'
          AND activo = TRUE
        ORDER BY orden, id
    """)
    planes_empresarial = cur.fetchall()

    # ==========================
    # KPIs
    # ==========================

    cur.execute("SELECT COUNT(*) FROM quiz")
    total_quizzes = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM respuestas_alumno")
    total_respuestas = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM usuarios WHERE rol='profesor'")
    total_profesores = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM alumnos")
    total_alumnos = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM empresa")
    total_empresas = cur.fetchone()[0]

    cur.close()
    conn.close()

    return render_template(
        'index.html',
        planes_individual=planes_individual,
        planes_empresarial=planes_empresarial,
        total_quizzes=total_quizzes,
        total_respuestas=total_respuestas,
        total_profesores=total_profesores,
        total_alumnos=total_alumnos,
        total_empresas=total_empresas
    )

def es_admin():
    return session.get('rol') == 'admin'

def es_profesor():
    return session.get('rol') == 'profesor'

from werkzeug.security import check_password_hash

@app.route('/login', methods=['GET', 'POST'])
def login():

    conn = get_db_connection()
    cur = conn.cursor()

    # =========================================================
    # PLAN SELECCIONADO
    # =========================================================

    plan_id = request.args.get('plan_id')

    if plan_id:

        try:
            plan_id = int(plan_id)

            cur.execute("""
                SELECT id
                FROM planes
                WHERE id = %s
                  AND activo = TRUE
            """, (plan_id,))

            plan = cur.fetchone()

            if not plan:
                plan_id = None

        except (ValueError, TypeError):
            plan_id = None


    # Si llegó un plan válido, lo conservamos en sesión
    if plan_id:
        session['plan_id'] = plan_id


    # =========================================================
    # EMPRESAS
    # =========================================================

    cur.execute("""
        SELECT cempre, dempre
        FROM empresa
        WHERE estado = TRUE
        ORDER BY dempre
    """)

    empresas = cur.fetchall()


    # =========================================================
    # LOGIN
    # =========================================================

    if request.method == 'POST':

        usuario = request.form['usuario']
        password = request.form['password']

        cur.execute("""
            SELECT
                id,
                usuario,
                password,
                rol,
                cempre,
                usuario
            FROM usuarios
            WHERE usuario = %s
        """, (usuario,))

        user = cur.fetchone()

        if user:

            db_password = user[2]

            if db_password == password or check_password_hash(db_password, password):

                session['user_id'] = user[0]
                session['usuario'] = user[1]
                session['rol'] = user[3]
                session['cempre'] = user[4]
                session['usuario'] = user[5]
                
                
                # =========================================================
                # VALIDAR LICENCIA ACTIVA DE LA EMPRESA
                # =========================================================

                cur.execute("""
                    SELECT id
                    FROM licencias
                    WHERE cempre = %s
                    AND estado = 'ACTIVA'
                    AND fecha_expiracion >= CURRENT_TIMESTAMP
                    LIMIT 1
                """, (user[4],))

                licencia_activa = cur.fetchone()

                session['licencia_activa'] = bool(licencia_activa)

                cur.close()
                conn.close()

                if session['licencia_activa']:
                    return redirect(url_for('splash'))
                else:
                    return redirect(url_for('home'))


        cur.close()
        conn.close()

        return render_template(
            "login.html",
            empresas=empresas,
            error="Credenciales incorrectas"
        )


    cur.close()
    conn.close()

    return render_template(
        "login.html",
        empresas=empresas
    )

def require_admin():
    if 'user_id' not in session or session.get('rol') != 'admin':
        return False
    return True
 
def get_db_connection():
    database_url = os.getenv("DATABASE_URL")

    print("DATABASE_URL =", database_url)

    if database_url:
        print("👉 CONECTANDO A RENDER")
        return psycopg2.connect(database_url, sslmode='require')
    else:
        print("👉 CONECTANDO A LOCAL (SIN SSL)")
        return psycopg2.connect(
            dbname="BDTeacherAssessment",
            user="postgres",
            password="1234",
            host="127.0.0.1",
            port="5432",
            sslmode='disable'   # 👈 🔥 FORZAR SIN SSL
        )

@app.route('/init-db')
def init_db():

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔥 crear tabla planes
    cur.execute("""
    CREATE TABLE IF NOT EXISTS planes (
        id SERIAL PRIMARY KEY,
        nombre TEXT,
        precio INTEGER,
        admins INTEGER,
        profesores INTEGER,
        alumnos INTEGER,
        quizzes INTEGER,
        orden INTEGER
    );
    """)
    
    # 🔥 agregar columnas faltantes
    cur.execute("ALTER TABLE planes ADD COLUMN IF NOT EXISTS tipo TEXT;")
    cur.execute("ALTER TABLE planes ADD COLUMN IF NOT EXISTS activo BOOLEAN DEFAULT TRUE;")

    conn.commit()
    cur.close()
    conn.close()

    return "DB inicializada"

@app.route('/restore-db')
def restore_db():

    conn = get_db_connection()
    cur = conn.cursor()

    with open('backup.sql', 'r', encoding='utf-8') as f:
        sql = f.read()

    # 🔥 dividir en comandos simples
    statements = sql.split(';')

    for statement in statements:
        statement = statement.strip()
        if statement:
            try:
                cur.execute(statement)
            except Exception as e:
                print("Error en:", statement[:100], e)

    conn.commit()
    cur.close()
    conn.close()

    return "DB restaurada (parcial si hubo errores)"

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        usuario = request.form['usuario']
        password = request.form['password']
        rol = request.form['rol']

        conn = get_db_connection()
        cur = conn.cursor()

        try:
            cur.execute(
                "INSERT INTO usuarios (usuario, password, rol) VALUES (%s, %s, %s)",
                (usuario, password, rol)
            )
            conn.commit()
        except:
            return "El usuario ya existe ❌"

        conn.commit()
        cur.close()
        conn.close()

        return redirect('/login')

    return render_template('register.html')

@app.route('/dashboard_profesor')
def dashboard_profesor():
    
    # =========================================================
    # VALIDAR SESIÓN
    # =========================================================
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # =========================================================
    # ROOT SIEMPRE PUEDE ENTRAR
    # =========================================================
    if session.get('rol') != 'root':

        # =====================================================
        # VALIDAR LICENCIA ACTIVA
        # =====================================================
        if not session.get('licencia_activa', False):

            return redirect(url_for('home'))

    if session['rol'] == 'admin':
        # lógica admin
        pass
    
    conn = get_db_connection()
    cur = conn.cursor()

    if session['rol'] == 'admin':
        cur.execute("""
            SELECT 
                z.id,
                u.usuario,
                z.titulo, 
                COUNT(DISTINCT r.alumno_id) AS alumnos,
                (select count(distinct p.id)  
                    from preguntas p
                    where p.quiz_id=z.id) AS preguntas,
                z.codigo,
                z.publico,
                COALESCE(
                    (z.config_json::json->>'tiempo_minutos')::int,
                    0
                ) AS tiempo_minutos,
                z.multiple_intentos,
                z.enviar_solucionario,
                COALESCE(
                    (z.config_json::json->>'permitir_retroceder')::boolean,
                    true
                ) AS permitir_retroceder,
                COALESCE(
                    (z.config_json::json->>'permitir_sin_contestar')::boolean,
                    false
                ) AS permitir_sin_contestar
                FROM quiz z
                INNER JOIN usuarios u on u.id=z.usuario_id
                LEFT JOIN preguntas p ON p.quiz_id = z.id
                LEFT JOIN respuestas_alumno r ON r.pregunta_id = p.id
                WHERE z.estado = 'A'
                AND z.cempre= %s
                GROUP BY z.id,u.usuario, z.titulo, z.codigo, z.publico,z.multiple_intentos,z.enviar_solucionario
                ORDER BY z.id DESC
            """, (session['cempre'], ))
    else:
        cur.execute("""
            SELECT 
                z.id,
                z.titulo, 
                COUNT(DISTINCT r.alumno_id) AS alumnos,
                (select count(distinct p.id)  
                    from preguntas p
                    where p.quiz_id=z.id) AS preguntas,
                z.codigo,
                z.publico,
                COALESCE(
                    (z.config_json::json->>'tiempo_minutos')::int,
                    0
                ) AS tiempo_minutos,
                z.multiple_intentos,
                z.enviar_solucionario,
                COALESCE(
                    (z.config_json::json->>'permitir_retroceder')::boolean,
                    true
                ) AS permitir_retroceder,
                COALESCE(
                    (z.config_json::json->>'permitir_sin_contestar')::boolean,
                    false
                ) AS permitir_sin_contestar
                FROM quiz z
                LEFT JOIN preguntas p ON p.quiz_id = z.id
                LEFT JOIN respuestas_alumno r ON r.pregunta_id = p.id
                WHERE z.estado = 'A'
                AND z.cempre= %s
                AND z.usuario_id = %s
                GROUP BY z.id, z.titulo, z.codigo, z.publico,z.multiple_intentos,z.enviar_solucionario
                ORDER BY z.id DESC
            """, (session['cempre'], session['user_id']))
        
    quizzes = cur.fetchall()
    
    cur.execute("""
        SELECT id
        FROM cola_ia
        WHERE estado IN ('pendiente', 'procesando')
        ORDER BY id DESC
    """)

    cola = cur.fetchall()


    cur.close()
    conn.close()

    return render_template('dashboard_profesor.html', quizzes=quizzes,cola=cola)

@app.route('/profesores')
@requiere_licencia
def profesores():

    if 'rol' not in session or session['rol'] != 'admin':
        return "Acceso no autorizado", 403

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, dni, nombre, apellido, usuario,password,correo,fecha_creacion
        FROM usuarios
        WHERE rol = 'profesor' AND cempre = %s
        ORDER BY nombre
    """, (session['cempre'],))

    profesores = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('profesores.html', profesores=profesores)

@app.route('/reporte_alumno/<int:quiz_id>/<int:alumno_id>')
def reporte_alumno(quiz_id, alumno_id):

    intento = request.args.get('intento', 1)

    examen = obtener_examen_alumno(alumno_id, quiz_id, intento)
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    # 🔥 generar archivo usando tu función existente
    generar_y_enviar_reporte(
        examen["detalle"],
        examen["nota"],
        examen["correo"],
        examen["nombre"],
        alumno_id,
        examen["titulo"],
        examen["dni"],
        fecha,
        False  # ❌ no enviar correo
    )

    import os

    if os.name == "nt":
        ruta_pdf = f"reporte_{alumno_id}.pdf"
    else:
        ruta_pdf = f"/tmp/reporte_{alumno_id}.pdf"

    return send_file(
        ruta_pdf,
        as_attachment=True,
        download_name=f"{examen['nombre']}_reporte.pdf",
        mimetype='application/pdf'
    )
    
@app.route('/obtener_intentos/<int:alumno_id>/<int:quiz_id>')
def obtener_intentos(alumno_id, quiz_id):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT q.id, q.titulo, iq.intento_numero
        FROM intentos_quiz iq
        JOIN quiz q ON q.id = iq.quiz_id
        WHERE iq.alumno_id = %s
            AND iq.quiz_id = %s
            AND iq.activo = TRUE
            AND iq.activo=TRUE
        ORDER BY iq.intento_numero
    """, (alumno_id, quiz_id))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    data = [
        {
            "quiz_id": r[0],
            "quiz": r[1],
            "intento": r[2]
        }
        for r in rows
    ]

    return jsonify(data)

@app.route('/reporte_quiz/<int:quiz_id>')
def reporte_quiz(quiz_id):
    
    con_sol = request.args.get("solucion", "true").lower() == "true"

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔹 obtener preguntas y opciones
    cur.execute("""
        SELECT
            p.id,
            p.texto,
            p.tipo,
            p.explicacion,
            o.texto,
            o.es_correcta
        FROM preguntas p
        JOIN opciones o ON o.pregunta_id = p.id
        WHERE p.quiz_id = %s
        ORDER BY p.id, o.id
    """, (quiz_id,))
    

    data = cur.fetchall()
    
    cur.execute("""
        SELECT titulo
        FROM quiz
        WHERE id = %s
    """, (quiz_id,))

    titulo_quiz = cur.fetchone()[0]
    
    # 🔹 NOMBRE DE LA EMPRESA
    cur.execute("""
        SELECT dempre 
        FROM empresa
        WHERE cempre = %s
    """, (session['cempre'],))

    empresa = cur.fetchone()

    nombre_empresa = empresa[0] if empresa else "Empresa"
    
    cur.close()
    conn.close()

    # 🚀 GENERAR PDF
    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from datetime import datetime

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT logo_header
        FROM configuracion_reportes
        WHERE cempre = %s
    """, (session['cempre'],))

    config_logo = cur.fetchone()

    cur.close()
    conn.close()

    # Logo por defecto
    logo_path = "static/img/logo.png"

    # Logo configurado por la empresa
    if config_logo and config_logo[0]:
        logo_path = config_logo[0].lstrip("/")
    
    
    # 🔹 HEADER PROFESIONAL DEL REPORTE

    from reportlab.lib.utils import ImageReader

    # Obtener dimensiones reales del logo
    try:
        img_reader = ImageReader(logo_path)
        img_width, img_height = img_reader.getSize()

        # Área máxima disponible para el logo
        max_logo_width = 90
        max_logo_height = 55

        # Mantener siempre la proporción original
        escala = min(
            max_logo_width / img_width,
            max_logo_height / img_height,
            1
        )

        logo_width = img_width * escala
        logo_height = img_height * escala

    except Exception:
        logo_width = 80
        logo_height = 50


    logo = Image(
        logo_path,
        width=logo_width,
        height=logo_height
    )

    # Información del reporte
    fecha_actual = datetime.now()

    info_header = Paragraph(
        f"""
        <b>{nombre_empresa}</b><br/>
        <b>Solucionario de:</b> {titulo_quiz}<br/>
        <font size="8">
            Usuario: {session.get('usuario')} |
            Fecha: {fecha_actual.strftime("%d/%m/%Y")} |
            Hora: {fecha_actual.strftime("%H:%M")}
        </font>
        """,
        styles["Normal"]
    )

    # Separador vertical entre logo e información
    header_data = [[logo, info_header]]

    tabla_header = Table(
        header_data,
        colWidths=[125, 343],
        rowHeights=[70]
    )

    tabla_header.setStyle(TableStyle([

        # Borde exterior
        ('BOX', (0,0), (-1,-1), 0.8, colors.HexColor("#9AA0A6")),

        # Logo centrado verticalmente
        ('VALIGN', (0,0), (0,0), 'MIDDLE'),

        # Información centrada verticalmente
        ('VALIGN', (1,0), (1,0), 'MIDDLE'),

        # Separador vertical
        ('LINEAFTER', (0,0), (0,0), 0.8, colors.HexColor("#B8BEC5")),

        # Holgura del logo a la derecha
        ('RIGHTPADDING', (0,0), (0,0), 20),

        # Holgura izquierda del logo
        ('LEFTPADDING', (0,0), (0,0), 8),

        # Información
        ('LEFTPADDING', (1,0), (1,0), 12),
        ('RIGHTPADDING', (1,0), (1,0), 8),

        # Fondo
        ('BACKGROUND', (0,0), (-1,-1), colors.white),

        # Barra inferior profesional
        ('LINEBELOW', (0,0), (-1,-1), 4, colors.HexColor("#1554B8")),
    ]))
    elements.append(tabla_header)
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("<b>SOLUCIONARIO</b>", styles['Title']))
    elements.append(Spacer(1, 10))

   # 🔹 AGRUPAR POR PREGUNTA
    preguntas = {}

    for p_id, p_texto, p_tipo, p_explicacion, o_texto, correcta in data:

        if p_id not in preguntas:
            preguntas[p_id] = {
                "texto": p_texto,
                "tipo": p_tipo,
                "explicacion": p_explicacion,
                "opciones": []
            }

        preguntas[p_id]["opciones"].append((o_texto, correcta))

   # 🔹 CONTENIDO
    for i, pregunta in enumerate(preguntas.values(), start=1):

        elements.append(
            Paragraph(f"{i}. {pregunta['texto']}", styles['Heading3'])
        )
        elements.append(Spacer(1, 5))

        tabla_data = []
        
        letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

        for indice, (op_texto, correcta) in enumerate(pregunta["opciones"]):
            
            letra = letras[indice]
        
            if correcta and con_sol:
                texto = f"<b>{letra})</b>{op_texto} ✔ "
            else:
                texto = f"<b>{letra})</b> {op_texto}"

            tabla_data.append([Paragraph(texto, styles['Normal'])])

        tabla = Table(tabla_data)

        tabla.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BOX', (0,0), (-1,-1), 1, colors.black),
        ]))

        elements.append(tabla)

        # 🔹 EXPLICACIÓN
        if con_sol and pregunta["explicacion"]:
            elements.append(Spacer(1, 4))
            elements.append(
            Paragraph(
                f"<b>Explicación:</b> {pregunta['explicacion']}",
                styles["Normal"]
            )
        )

        elements.append(Spacer(1, 10))

    doc.build(elements)
    buffer.seek(0)

    #return send_file(buffer, as_attachment=False,
    #            download_name="reporte.pdf",
    #            mimetype='application/pdf')
        
    return send_file(buffer,
    as_attachment=False,
    download_name=f"{titulo_quiz}.pdf",
    mimetype='application/pdf')

@app.route('/alumno')
def alumno():
    if 'usuario' in session:
        return f"Bienvenido alumno {session['usuario']} 🎓"
    return redirect('/login')

@app.route('/registrar_admin', methods=['POST'])
def registrar_admin():

    data = request.get_json()

    tipo_registro = data.get("tipoRegistro")

    usuario = data.get("usuario")
    password = data.get("password")
    dni = data.get("dni")
    nombre = data.get("nombre")
    apellido = data.get("apellido")
    correo = data.get("correo")
    ruc = data.get("ruc")
    plan_id = data.get("plan_id")
    print("PLAN RECIBIDO EN REGISTRO:", plan_id)
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    # =========================================================
    # VALIDAR PLAN
    # =========================================================

    try:
        plan_id = int(plan_id)
    except (ValueError, TypeError):
        return jsonify({
            "mensaje": "Debe seleccionar un plan"
        }), 400


    cur.execute("""
        SELECT
            id,
            nombre,
            precio,
            periodicidad,
            admins,
            profesores,
            alumnos,
            quizzes
        FROM planes
        WHERE id = %s
        AND activo = TRUE
    """, (plan_id,))

    plan = cur.fetchone()

    if not plan:
        return jsonify({
            "mensaje": "El plan seleccionado no existe o está inactivo"
        }), 400


    # Datos reales del plan, obtenidos desde BD
    plan_id = plan[0]
    plan_nombre = plan[1]
    plan_precio = plan[2]
    plan_periodicidad = plan[3]
    plan_admins = plan[4]
    plan_profesores = plan[5]
    plan_alumnos = plan[6]
    plan_quizzes = plan[7]

    print("PLAN:", plan_id, plan_nombre, plan_precio, plan_periodicidad)

    try:

        # =========================================================
        # 1. VALIDAR QUE EL USUARIO NO EXISTA
        # =========================================================

        cur.execute("""
            SELECT id
            FROM usuarios
            WHERE usuario = %s
        """, (usuario,))

        if cur.fetchone():

            return jsonify({
                "mensaje": "El usuario ya existe"
            }), 400


        # =========================================================
        # 2. DETERMINAR LA EMPRESA
        # =========================================================

        if tipo_registro == "natural":

            # -----------------------------------------------------
            # PERSONA NATURAL
            # Se crea automáticamente una empresa.
            # -----------------------------------------------------

            empresa_nombre = f"{nombre} {apellido}".strip()

            cur.execute("""
                INSERT INTO empresa (
                    dempre,
                    licencia,
                    estado
                )
                VALUES (%s, FALSE, TRUE)
                RETURNING cempre
            """, (empresa_nombre,))

            cempre = cur.fetchone()[0]


        elif tipo_registro == "empresa":

            # -----------------------------------------------------
            # EMPRESA
            # El usuario seleccionó una empresa existente
            # o una empresa creada desde el formulario.
            # -----------------------------------------------------

            empresa_nombre = data.get("empresa")
            ruc = data.get("ruc")

            if not empresa_nombre:

                return jsonify({
                    "mensaje": "Debe seleccionar una empresa"
                }), 400


            # Buscar empresa

            cur.execute("""
                SELECT cempre
                FROM empresa
                WHERE dempre = %s
            """, (empresa_nombre,))

            row = cur.fetchone()


            if row:

                cempre = row[0]

            else:
    
                cur.execute("""
                    INSERT INTO empresa (
                        dempre,
                        ruc,
                        licencia,
                        estado
                    )
                    VALUES (%s, %s, FALSE, TRUE)
                    RETURNING cempre
                """, (empresa_nombre, ruc))

                cempre = cur.fetchone()[0]


        else:

            return jsonify({
                "mensaje": "Tipo de registro inválido"
            }), 400

        cur.execute("""
            SELECT id
            FROM usuarios
            WHERE dni = %s
            AND cempre = %s
        """, (dni, cempre))

        if cur.fetchone():

            conn.rollback()

            return jsonify({
                "campo": "dni",
                "mensaje": "El DNI ya está registrado en esta empresa. Registre otro DNI."
            }), 409
        # =========================================================
        # 3. CREAR USUARIO ADMINISTRADOR
        # =========================================================

        cur.execute("""
            INSERT INTO usuarios (
                usuario,
                password,
                rol,
                dni,
                nombre,
                apellido,
                correo,
                cempre
            )
            VALUES (
                %s,
                %s,
                'admin',
                %s,
                %s,
                %s,
                %s,
                %s
            )
            RETURNING id
        """, (
            usuario,
            password,
            dni,
            nombre,
            apellido,
            correo,
            cempre
        ))

        usuario_id = cur.fetchone()[0]
        
        session['user_id'] = usuario_id
        session['usuario'] = usuario
        session['rol'] = 'admin'
        session['cempre'] = cempre
        
        # =========================================================
        # 4. LICENCIA STARTER GRATUITA
        # =========================================================
        print("LICENCIA precio:", plan_precio)

        if plan_precio == 0:

            from datetime import datetime, timedelta

            fecha_inicio = datetime.now()
            fecha_expiracion = fecha_inicio + timedelta(days=7)

            # Generar código único
            codigo = "TA-" + secrets.token_hex(10).upper()

            cur.execute("""
                INSERT INTO licencias (
                    codigo,
                    cempre,
                    plan_id,
                    fecha_inicio,
                    fecha_expiracion,
                    estado,
                    periodicidad,
                    cantidad_periodos,
                    max_admins,
                    max_profesores,
                    max_alumnos,
                    max_quizzes,
                    origen,
                    creado_por
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    'GENERADA',
                    %s,
                    1,
                    %s,
                    %s,
                    %s,
                    %s,
                    'AUTOMATICA',
                    %s
                )
                RETURNING id, codigo
            """, (
                codigo,
                cempre,
                plan_id,
                fecha_inicio,
                fecha_expiracion,
                plan_periodicidad,
                plan_admins,
                plan_profesores,
                plan_alumnos,
                plan_quizzes,
                usuario_id
            ))

            licencia_id, codigo_licencia = cur.fetchone()

            print("LICENCIA STARTER:", licencia_id)
            print("CODIGO:", codigo_licencia)
            print("INICIO:", fecha_inicio)
            print("EXPIRA:", fecha_expiracion)


        # =========================================================
        # 4. CONFIRMAR TODO
        # =========================================================

        conn.commit()


        respuesta = {
            "mensaje": "Usuario administrador creado correctamente",
            "plan_id": plan_id,
            "precio": float(plan_precio)
        }

        if plan_precio == 0:
            respuesta["licencia_id"] = licencia_id
            respuesta["codigo_licencia"] = codigo_licencia
            respuesta["fecha_expiracion"] = fecha_expiracion.strftime("%d/%m/%Y")

        return jsonify(respuesta)


    except Exception as e:

        conn.rollback()

        print("========================================")
        print("ERROR REGISTRO:")
        print(str(e))
        print("========================================")

        return jsonify({
            "ok": False,
            "mensaje": "ERROR REAL: " + str(e)
        }), 500


    finally:

        cur.close()
        conn.close()


@app.route('/pago')
def pago():

    # =========================================================
    # 1. VALIDAR SESIÓN
    # =========================================================

    if 'user_id' not in session or 'cempre' not in session:
        return redirect(url_for('login'))

    # =========================================================
    # 2. OBTENER PLAN
    # =========================================================

    plan_id = request.args.get('plan_id', type=int)

    if not plan_id:
        return redirect(url_for('home'))

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                id,
                nombre,
                precio,
                periodicidad
            FROM planes
            WHERE id = %s
              AND activo = TRUE
        """, (plan_id,))

        plan = cur.fetchone()

        if not plan:
            return "Plan no válido", 400

        # Starter nunca debe entrar a pago
        if float(plan[2]) <= 0:
            return redirect(url_for('home'))

        return render_template(
            "pago.html",
            plan_id=plan[0],
            plan_nombre=plan[1],
            precio=float(plan[2]),
            periodicidad=plan[3]
        )

    finally:

        cur.close()
        conn.close()

@app.route('/admin')
def admin():
    if 'usuario' in session:
        return f"Panel admin {session['usuario']} ⚙️"
    return redirect('/login')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

# PASO 1: ingresar código
@app.route('/ingresar_codigo', methods=['GET', 'POST'])
def ingresar_codigo():
    if request.method == 'POST':
        codigo = request.form['codigo']

        # guardamos el código en sesión
        session['codigo_quiz'] = codigo
        session['modo'] = 'local'  # 🔥 AQUÍ

        return redirect('/ingresar_dni')

    return render_template('ingresar_codigo.html')


# PASO 2: ingresar DNI
@app.route('/ingresar_dni', methods=['GET', 'POST'])
def ingresar_dni():
    if request.method == 'POST':
        dni = request.form['dni']
         

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("SELECT id, nombre, apellido, correo FROM alumnos WHERE dni=%s  ", (dni, ))
        alumno = cur.fetchone()

        if alumno:
            alumno_id = alumno[0]
            correo_db = alumno[3]
            correo_form = request.form.get('correo')

            # 🔥 si no tiene correo → actualizar
            if not correo_db or not str(correo_db).strip():
    
                correo_form = request.form.get('correo')

                print("📩 correo recibido:", correo_form)

                if not correo_form or not correo_form.strip():
                    return render_template(
                        "ingresar_dni.html",
                        dni=dni,
                        mostrar_campos=True,
                        error="Debes ingresar tu correo"
                    )

                cur.execute(
                    "UPDATE alumnos SET correo=%s WHERE dni=%s",
                    (correo_form.strip(), dni)
                )
                conn.commit()
        else:
            cur.close()
            conn.close()
            return redirect(f"/registro_alumno?dni={dni}")

        # 🔥 obtener quiz
        codigo = session.get('codigo_quiz', '').strip().upper()

        # 🔍 primero buscar en salon_quiz
        cur.execute("""
            SELECT id, quiz_id
            FROM salon_quiz
            WHERE UPPER(TRIM(codigo)) = %s
        """, (codigo,))

        row = cur.fetchone()

        if row:
            salon_quiz_id = row[0]
            quiz_id = row[1]

            cur.close()
            conn.close()

            return redirect(url_for(
                'resolver_quiz_salon',
                salon_quiz_id=salon_quiz_id,
                alumno_id=alumno_id
            ))

        # 🔍 si no está en salon, usar lógica original
        cur.execute("SELECT id FROM quiz WHERE codigo=%s", (codigo,))
        quiz = cur.fetchone()

        if quiz:
            quiz_id = quiz[0]
            
            cur.execute("""
                SELECT multiple_intentos
                FROM quiz
                WHERE id = %s
            """, (quiz_id,))

            multiple_intentos = cur.fetchone()[0]
            
            if not multiple_intentos:
    
                cur.execute("""
                    SELECT COUNT(*)
                    FROM intentos_quiz
                    WHERE alumno_id = %s
                    AND quiz_id = %s
                    AND nota_final IS NOT NULL
                    AND activo = TRUE
                """, (alumno_id, quiz_id))

                ya_rindio = cur.fetchone()[0] > 0

                if ya_rindio:
    
                    cur.execute("""
                        SELECT
                            nota_final,
                            fecha_fin,
                            tiempo_total_segundos
                        FROM intentos_quiz
                        WHERE alumno_id = %s
                        AND quiz_id = %s
                        AND nota_final IS NOT NULL
                        AND activo=TRUE
                        ORDER BY intento_numero DESC
                        LIMIT 1
                    """, (alumno_id, quiz_id))

                    ultimo = cur.fetchone()

                    nota = ultimo[0] if ultimo else 0
                    fecha = ultimo[1].strftime("%d/%m/%Y %H:%M") if ultimo and ultimo[1] else "-"

                    if ultimo and ultimo[2]:
                        minutos = ultimo[2] // 60
                        segundos = ultimo[2] % 60
                        tiempo = f"{minutos:02}:{segundos:02}"
                    else:
                        tiempo = "--:--"

                    cur.execute("""
                        SELECT nombre, apellido
                        FROM alumnos
                        WHERE id = %s
                    """, (alumno_id,))

                    alumno = cur.fetchone()
                    nombre = f"{alumno[0]} {alumno[1]}" if alumno else "Alumno"

                    cur.close()
                    conn.close()

                    return render_template(
                        "quiz_bloqueado.html",
                        alumno=nombre,
                        nota=nota,
                        fecha=fecha,
                        tiempo=tiempo
                    )

            cur.close()
            conn.close()

            return redirect(url_for(
                'resolver_quiz',
                quiz_id=quiz_id,
                alumno_id=alumno_id
            ))
            
            

        # ❌ código inválido
        cur.close()
        conn.close()
        return "Código inválido", 404 

    return render_template('ingresar_dni.html')

@app.route('/resolver_quiz_salon/<int:salon_quiz_id>/<int:alumno_id>')
def resolver_quiz_salon(salon_quiz_id, alumno_id):
   

    conn = get_db_connection()
    cur = conn.cursor()
    
    # 🔍 obtener quiz_id (robusto)
    quiz_id = None

    if salon_quiz_id:
        cur.execute("""
            SELECT quiz_id
            FROM salon_quiz
            WHERE id = %s
        """, (salon_quiz_id,))
        
        row = cur.fetchone()
        
        if row:
            quiz_id = row[0]

    # 🔥 fallback (por si no viene de salón)
    if not quiz_id:
        quiz_id = request.args.get("quiz_id")

    # 🔥 última defensa
    if not quiz_id:
        return "No se pudo determinar el quiz", 400
    
    cur.execute("""
        SELECT config_json, multiple_intentos, titulo
        FROM quiz
        WHERE id = %s
    """, (quiz_id,))

    row = cur.fetchone()

    config_json = json.loads(row[0]) if row and row[0] else {}
    multiple_intentos = row[1] if row else False
    titulo = row[2] if row else "Titulo del quiz"

    cur.execute("""
        SELECT COALESCE(MAX(intento_numero), 0)
        FROM intentos_quiz
        WHERE alumno_id = %s AND quiz_id = %s
        AND activo=TRUE
    """, (alumno_id, quiz_id))

    ultimo = cur.fetchone()[0]
    nuevo_intento = ultimo + 1

    cur.execute("""
        INSERT INTO intentos_quiz (
            alumno_id,
            quiz_id,
            intento_numero,
            fecha_inicio
        )
        VALUES (
            %s,
            %s,
            %s,
            NOW()
        )
        RETURNING id
    """, (alumno_id, quiz_id, nuevo_intento))

    intento_id = cur.fetchone()[0]

    print("INTENTO CREADO:", intento_id)

    conn.commit()

    # 👇 AGREGA ESTO AQUÍ
    cur.execute("""
        SELECT id
        FROM intentos_quiz
        WHERE id = %s
         AND activo=TRUE
    """, (intento_id,))

    print("VERIFICACION:", cur.fetchone())

    # luego continúa con:
    # obtener preguntas
    cur.execute("""
        SELECT id, texto, tipo, explicacion
        FROM preguntas
        WHERE quiz_id = %s
    """, (quiz_id,))

    row = cur.fetchone()

    if not row:
        return "Error", 404

    # 🔍 obtener preguntas (igual que resolver_quiz)
    cur.execute("""
        SELECT id, texto, tipo, explicacion
        FROM preguntas
        WHERE quiz_id = %s
    """, (quiz_id,))

    preguntas = cur.fetchall()

    data = []

    for p in preguntas:
        cur.execute("""
            SELECT id, texto
            FROM opciones
            WHERE pregunta_id = %s
        """, (p[0],))

        opciones_db = cur.fetchall()

        opciones = []
        for op in opciones_db:
            opciones.append({
                "id": op[0],
                "texto": op[1]
            })

        data.append({
            'id': p[0],
            'texto': p[1],
            'tipo': p[2],
            'explicacion': p[3],
            'opciones': opciones
        })

    cur.close()
    conn.close()
    
    print("DATA DEBUG:", data)

    return render_template(
        'resolver_quiz.html',
        preguntas=data,
        quiz_id=quiz_id,
        alumno_id=alumno_id,
        salon_quiz_id=salon_quiz_id , # 🔥 clave
        intento_id=intento_id,
        config_json=config_json,
        multiple_intentos=multiple_intentos,
        titulo=titulo
    )
 

@app.route('/obtener_asignaciones/<int:salon_id>')
def obtener_asignaciones(salon_id):

    if 'user_id' not in session:
        return jsonify({"error": "No autenticado"}), 401

    conn = get_db_connection()
    cur = conn.cursor()

    if session["rol"] == "admin":

        cur.execute("""
            SELECT
                sq.id,
                q.titulo,
                sq.codigo,
                 s.usuario
            FROM salon_quiz sq
            JOIN quiz q ON q.id = sq.quiz_id
            JOIN salon s ON s.id=sq.salon_id
            WHERE sq.salon_id = %s
              AND q.cempre = %s
            ORDER BY sq.id DESC
        """, (
            salon_id,
            session["cempre"]
        ))

    else:   # profesor

        cur.execute("""
            SELECT
                sq.id,
                q.titulo,
                sq.codigo
            FROM salon_quiz sq
            JOIN quiz q ON q.id = sq.quiz_id
            WHERE sq.salon_id = %s
              AND q.usuario_id = %s
            ORDER BY sq.id DESC
        """, (
            salon_id,
            session["user_id"]
        ))

    data = cur.fetchall()

    cur.close()
    conn.close()

    if session["rol"] == "admin":
    
        resultado = [
            {
                "id": r[0],
                "titulo": r[1],
                "codigo": r[2],
                "usuario": r[3]
            }
            for r in data
        ]

    else:

        resultado = [
            {
                "id": r[0],
                "titulo": r[1],
                "codigo": r[2]
            }
            for r in data
        ]

    return jsonify(resultado)

 

@app.route('/completar_correo', methods=['GET', 'POST'])
def completar_correo():
    dni = request.args.get('dni')

    if request.method == 'POST':
        correo = request.form['correo']

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            "UPDATE alumnos SET correo=%s WHERE dni=%s",
            (correo, dni)
        )

        # 🔥 ahora buscamos alumno y redirigimos
        cur.execute("SELECT id FROM alumnos WHERE dni=%s", (dni,))
        alumno = cur.fetchone()

        codigo = session.get('codigo_quiz')

        cur.execute("SELECT id FROM quiz WHERE codigo=%s", (codigo,))
        quiz = cur.fetchone()

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for(
            'resolver_quiz',
            quiz_id=quiz[0],
            alumno_id=alumno[0]
        ))

    return render_template('completar_correo.html', dni=dni)

@app.route('/registro_alumno', methods=['GET', 'POST'])
def registro_alumno():
    dni = request.args.get('dni')  # viene desde la URL

    if request.method == 'POST':
        nombre = request.form['nombre']
        apellido = request.form['apellido']
        correo = request.form['correo']

        conn = get_db_connection()
        cur = conn.cursor()

        # 🔥 INSERT y obtener ID
        cur.execute("""
            INSERT INTO alumnos (dni, nombre, apellido, correo)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (dni, nombre, apellido, correo))

        alumno_id = cur.fetchone()[0]

        conn.commit()

        codigo = session.get('codigo_quiz')

        # 🔍 buscar en salon_quiz
        cur.execute("""
            SELECT id, quiz_id
            FROM salon_quiz
            WHERE codigo = %s
        """, (codigo,))

        row = cur.fetchone()

        if row:
            salon_quiz_id = row[0]

            cur.close()
            conn.close()

            return redirect(url_for(
                'resolver_quiz_salon',
                salon_quiz_id=salon_quiz_id,
                alumno_id=alumno_id
            ))

        # 🔍 fallback quiz normal
        cur.execute("SELECT id FROM quiz WHERE codigo=%s", (codigo,))
        quiz = cur.fetchone()

        if quiz:
            quiz_id = quiz[0]

            cur.close()
            conn.close()

            return redirect(url_for(
                'resolver_quiz',
                quiz_id=quiz_id,
                alumno_id=alumno_id
            ))

        cur.close()
        conn.close()

    return render_template('registro_alumno.html', dni=dni)

@app.route('/quiz')
def quiz():

    codigo = session.get('codigo_quiz')

    if not codigo:
        return "No hay código en sesión"

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔥 obtener quiz_id desde código
    cur.execute("SELECT id FROM quiz WHERE codigo=%s", (codigo,))
    quiz = cur.fetchone()

    if not quiz:
        return "Código inválido"

    quiz_id = quiz[0]

    # 🔥 ahora sí traer preguntas correctas
    cur.execute("""
        SELECT id, texto, tipo, explicacion
        FROM preguntas
        WHERE quiz_id = %s
    """, (quiz_id,))
    preguntas = cur.fetchall()

    data = []

    for p in preguntas:
        cur.execute("SELECT id, texto FROM opciones WHERE pregunta_id=%s", (p[0],))
        opciones_db = cur.fetchall()

        opciones = []
        for op in opciones_db:
            opciones.append({
                "id": op[0],
                "texto": op[1]
            })

        data.append({
            'id': p[0],
            'texto': p[1],
            'tipo': p[2],
            'explicacion': p[3],
            'opciones': opciones
        })

    cur.close()
    conn.close()

    return render_template('quiz.html', preguntas=data)

    
@app.route('/crear_quiz', methods=['GET', 'POST'])
@requiere_licencia
def crear_quiz():
    
    if request.method == 'POST':    
        titulo = request.form['titulo']
        config_json = request.form.get("config_json")
        usuario = session.get('usuario')
        
        multiple_intentos = request.form.get('multiple_intentos') in ['true', 'on', '1']
        enviar_solucionario = request.form.get("enviar_solucionario") in ['true', 'on', '1']
        publico = not (request.form.get("privado") in ['true', 'on', '1'])
        
        config = json.loads(config_json)

        config["multiple_intentos"] = multiple_intentos
        config["enviar_solucionario"] = enviar_solucionario
        config["privado"] = not publico

        config_json = json.dumps(config)
        
        if not config_json or config_json=='':
            config_json = json.dumps(obtener_config_default())
        
        if not usuario:
            return redirect('/login')

        conn = get_db_connection()
        cur = conn.cursor()

        # crear quiz
        cur.execute(
            """
            INSERT INTO quiz
            (
                titulo,
                cempre,
                usuario_id,
                usuario,
                estado,
                multiple_intentos,
                enviar_solucionario,
                publico,
                config_json
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
            """,
            (
                titulo,
                session['cempre'],
                session['user_id'],
                usuario,
                'A',
                multiple_intentos,
                enviar_solucionario,
                publico,
                config_json
            )
        )
        quiz_id = cur.fetchone()[0]
        
        # recorrer preguntas
        orden = 1
        for key, texto in request.form.items():
            if key.startswith("pregunta_"):

                if not texto.strip():
                    continue

                num = key.split("_")[1]
                tipo = request.form.get(f"tipo_{num}")

                cur.execute(
                "INSERT INTO preguntas (quiz_id, texto, tipo, norden) VALUES (%s, %s, %s, %s) RETURNING id",
                (quiz_id, texto, tipo, orden)
                )
                orden += 1
                pregunta_id = cur.fetchone()[0]

                if tipo == "vf":
                    correcta = request.form.get(f"correcta_{num}")
                    explicacion = request.form.get(f"explicacion_{num}")

                    cur.execute(
                        "INSERT INTO opciones (pregunta_id, texto, es_correcta) VALUES (%s, %s, %s)",
                        (pregunta_id, "Verdadero", correcta == "Verdadero")
                    )

                    cur.execute(
                        "INSERT INTO opciones (pregunta_id, texto, es_correcta) VALUES (%s, %s, %s)",
                        (pregunta_id, "Falso", correcta == "Falso")
                    )

                    cur.execute(
                        "UPDATE preguntas SET explicacion=%s WHERE id=%s",
                        (explicacion, pregunta_id)
                    )

                else:
                    correcta = request.form.get(f"correcta_{num}")
                    print("CORRECTA RECIBIDA:", correcta)
                    explicacion = request.form.get(f"explicacion_{num}")
                    print("EXPLICACION RECIBIDA:", explicacion)
                    for i in range(1, 6):
                        opcion = request.form.get(f"opcion_{num}_{i}")
                        if opcion:
                            es_correcta = str(i) == correcta

                            cur.execute(
                                "INSERT INTO opciones (pregunta_id, texto, es_correcta) VALUES (%s, %s, %s)",
                                (pregunta_id, opcion, es_correcta)
                            )
                    # 🔥 FALTABA ESTO
                    cur.execute(
                        "UPDATE preguntas SET explicacion=%s WHERE id=%s",
                        (explicacion, pregunta_id)
                    )
                            

        conn.commit()
        cur.close()
        conn.close()

        return redirect('/dashboard_profesor')

    return render_template( 
        'crear_quiz.html',
        config={
            "config_json": obtener_config_default()
        }
    )


@app.route('/generar_quiz_ia', methods=['POST'])
def generar_quiz_ia():
    data = request.get_json()
    print("JSON RECIBIDO:", data)
    prompt_usuario = data.get('prompt')
    cantidad = data.get('cantidad')
    tipo = data.get('tipo')
    titulo = data.get("titulo")
    config_json = data.get("config_json", "{}")
    origen = data.get('origen', 'prompt')
    contenido_extra = data.get('contenido_extra')  # aquí irá texto del archivo o quiz
    
    # 🔥 NUEVO — leer configuración m
    multiple_intentos = str(data.get("multiple_intentos")).lower() in ["true", "1", "on"] 
    enviar_solucionario = str(data.get("enviar_solucionario")).lower() in ["true", "1", "on"] 
    publico = str(data.get("publico")).lower() in ["true", "1", "on"]

    if origen == "prompt":
        base = prompt_usuario

    elif origen == "archivo":
        base = f"Basado EXCLUSIVAMENTE en el siguiente texto:\n{contenido_extra}"

    elif origen == "quiz_previo":
        base = f"Convierte el siguiente contenido en un quiz estructurado con respuestas:\n{contenido_extra}"

    else:
        base = prompt_usuario

    prompt = f"""
    {base} 
    
    Genera un quiz con estas condiciones:
    - Cantidad total de preguntas: {cantidad}
    - Tipo de preguntas: {tipo}

    Devuelve SOLO JSON válido en este formato:
    [
      {{
        "tipo": "vf",
        "texto": "...",
        "correcta": "Verdadero" ,
        "explicacion": "..."
      }},
      {{
        "tipo": "multiple",
        "texto": "...",
        "opciones": [
            "opcion 1",
            "opcion 2",
            "opcion 3",
            "opcion 4",
            "opcion 5"
        ],
        "correcta": "A",
        "explicacion": "..."
      }}
    ]
    
    - Siempre generar exactamente 5 opciones.
    - El contenido de las opciones debe ser texto real, NO letras.
    - El campo "correcta" debe contener ÚNICAMENTE una letra: A, B, C, D o E.
    - Nunca devolver el texto de la respuesta en el campo "correcta".
    - Nunca devolver dos respuestas correctas.

    Genera EXACTAMENTE lo que el usuario pide .
    la explicacion es un breve refuerzo de por que la opcion elegida en (opcion) o en (VF) es la correcta
    No expliques nada mas. Solo JSON puro.
     
    """

    conn = get_db_connection()
    cur = conn.cursor()

    usuario_id = session.get("user_id")
    usuario = session.get("usuario")
    cempre = session.get("cempre")

    cur.execute(""" 
        INSERT INTO cola_ia ( prompt, cantidad, tipo, estado, usuario_id, 
                            usuario, cempre, multiple_intentos, enviar_solucionario, publico,
                            titulo, origen, contenido_extra,config_json ) 
        VALUES (%s, %s, %s, 'pendiente', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) 
        RETURNING id 
    """, ( 
        prompt, cantidad, tipo, usuario_id, usuario, 
        cempre, multiple_intentos, enviar_solucionario, publico,
        titulo, origen, contenido_extra,config_json
    ))
    cola_id = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "status": "encolado",
        "cola_id": cola_id
    })

@app.route('/procesar_examen', methods=['POST'])
def procesar_examen():

    archivo = request.files.get('archivo')  
    
    titulo = request.form.get("titulo")

    multiple_intentos = str(request.form.get("multiple_intentos")).lower() in ["true", "1", "on"]
    enviar_solucionario = str(request.form.get("enviar_solucionario")).lower() in ["true", "1", "on"]
    publico = str(request.form.get("publico")).lower() in ["true", "1", "on"]

    if not archivo:
        return jsonify({"error": "No se envió archivo"}), 400

    # 🔥 LEER COMO TEXTO (IMPORTANTE)
    try:
        contenido = archivo.read().decode("utf-8")
    except:
        contenido = archivo.read().decode("latin-1")

    prompt = f"""
    Convierte este examen en preguntas tipo quiz:

    {contenido}

    Devuelve SOLO JSON en este formato:
    [
      {{
        "tipo": "vf",
        "texto": "...",
        "correcta": "Verdadero",
        "explicacion": ""
      }},
      {{
        "tipo": "multiple",
        "texto": "...",
        "opciones": ["A","B","C","D","E"],
        "correcta": "A"
      }}
    ]

    No expliques nada. Solo JSON válido.
    No olvidar cada respuesta debe estar precedida con las letras A - E (opcion multiple)
    """
    conn = get_db_connection()
    cur = conn.cursor()

    usuario_id = session.get("user_id")
    usuario = session.get("usuario")
    cempre = session.get("cempre")

    cur.execute("""
        INSERT INTO cola_ia (
            prompt,
            contenido_extra,
            origen,
            estado,
            usuario_id,
            usuario,
            cempre,
            multiple_intentos,
            enviar_solucionario,
            publico,
            titulo
        )
        VALUES (
            %s, %s, %s, 'pendiente',
            %s, %s, %s, %s,
            %s, %s, %s
        )
        RETURNING id
    """, (
        prompt,
        contenido,
        "importar",
        usuario_id,
        usuario,
        cempre,
        multiple_intentos,
        enviar_solucionario,
        publico,
        titulo
    ))

    cola_id = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "status": "encolado",
        "cola_id": cola_id
    })
    

     

@app.route('/generar_quiz_desde_archivo', methods=['POST'])
def generar_quiz_desde_archivo():

    archivo = request.files.get('archivo')
    cantidad = request.form.get('cantidad')
    tipo = request.form.get('tipo')

    if not archivo:
        return jsonify({"error": "No se envió archivo"}), 400

    try:
        contenido = archivo.read().decode("utf-8")
    except:
        contenido = archivo.read().decode("latin-1")

    prompt = f"""
    Basado en el siguiente contenido:

    {contenido}

    Genera un quiz con estas condiciones:
    - Cantidad total: {cantidad}
    - Tipo: {tipo}

    REGLAS OBLIGATORIAS:

    1. Si tipo = "vf":
    - TODAS las preguntas deben ser Verdadero/Falso
    - Alternar respuestas: Verdadero, Falso, Verdadero, Falso...

    2. Si tipo = "multiple":
    - TODAS deben ser opción múltiple (5 opciones, A-E)
    - La respuesta correctas deben ROTAR así:
        A, B, C, D, E, A, B, C...

    3. Si tipo = "mixto":
    - Mezclar VF y múltiple
    - Mantener reglas anteriores en cada tipo

    4. Las respuestas NO deben ser todas iguales
    5. No repetir patrones obvios
    6. Generar preguntas coherentes con el contenido
    7. Las opciones deben ser textos reales (ej: "Egipto", "Roma", etc.)
   NO usar letras como contenido de opciones
   Las letras A–E son solo referenciales para la respuesta correcta

    FORMATO (SOLO JSON):
    [
    {{
        "tipo": "vf",
        "texto": "...",
        "correcta": "Verdadero",
        "explicacion": "..."
    }},
    {{
        "tipo": "multiple",
        "texto": "...",
        "opciones": ["opcion 1", "opcion 2", "opcion 3", "opcion 4", "opcion 5"]
        "correcta": "A",
        "explicacion": "..."
    }}
    ]
    La explicacion es una reseña breve de por que esa opcion es verdader o correcta
    No expliques nada. Solo JSON válido.
    """
    conn = get_db_connection()
    cur = conn.cursor()

    usuario_id = session.get("user_id")
    usuario = session.get("usuario")
    cempre = session.get("cempre")

    # puedes ajustar defaults si quieres
    multiple_intentos = str(request.form.get("multiple_intentos")).lower() in ["true", "1", "on"]
    enviar_solucionario = str(request.form.get("enviar_solucionario")).lower() in ["true", "1", "on"]
    publico = str(request.form.get("publico")).lower() in ["true", "1", "on"]
    titulo = request.form.get("titulo") or "Quiz desde archivo"

    cur.execute(""" 
        INSERT INTO cola_ia (
            prompt, cantidad, tipo, estado, usuario_id, 
            usuario, cempre, multiple_intentos, enviar_solucionario, publico,
            titulo, origen, contenido_extra
        ) 
        VALUES (%s, %s, %s, 'pendiente', %s, %s, %s, %s, %s, %s, %s, %s, %s) 
        RETURNING id 
    """, ( 
        prompt, cantidad, tipo, usuario_id, usuario, 
        cempre, multiple_intentos, enviar_solucionario,publico, 
        titulo, "archivo", contenido
    ))

    cola_id = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "status": "encolado",
        "cola_id": cola_id
    })

@app.route('/editar_quiz/<int:quiz_id>', methods=['GET', 'POST'])
def editar_quiz(quiz_id):

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':

        titulo = request.form['titulo']
        multiple_intentos = request.form.get("multiple_intentos") == "on"
        enviar_solucionario = bool(request.form.get("enviar_solucionario"))
        publico = not (request.form.get("privado") == "on")
        config_json = request.form.get("config_json")

        print("CONFIG_JSON RECIBIDO:", repr(config_json))

        if config_json:
            config = json.loads(config_json)
        else:
            config = obtener_config_default()

        config["multiple_intentos"] = multiple_intentos
        config["enviar_solucionario"] = enviar_solucionario
        config["privado"] = not publico

        config_json = json.dumps(config)
        
        if not config_json or config_json=='':
            config_json = json.dumps(obtener_config_default())
        
        
        # actualizar título
        cur.execute(
            """
            UPDATE quiz
            SET titulo=%s,
                multiple_intentos=%s,
                enviar_solucionario=%s,
                publico=%s,
                config_json=%s
            WHERE id=%s
            """,
            (
                titulo,
                multiple_intentos,
                enviar_solucionario,
                publico,
                config_json,
                quiz_id
            )
        )

        # 🔥 borrar todo
#       cur.execute("DELETE FROM opciones WHERE pregunta_id IN (SELECT id FROM preguntas WHERE quiz_id=%s)", (quiz_id,))
#        cur.execute("DELETE FROM preguntas WHERE quiz_id=%s", (quiz_id,))

        # 🔥 volver a insertar (igual que crear_quiz)
        
        for key, texto in request.form.items():
            if key.startswith("pregunta_") and not key.startswith("pregunta_id_"):

                if not texto.strip():
                    continue

                num = key.split("_")[1]
                tipo = request.form.get(f"tipo_{num}")
                
                pregunta_id = request.form.get(f"pregunta_id_{num}")

                if pregunta_id:

                    cur.execute("""
                        UPDATE preguntas
                        SET texto=%s,
                            tipo=%s,
                            explicacion=%s
                        WHERE id=%s
                    """, (
                        texto,
                        tipo,
                        request.form.get(f"explicacion_{num}"),
                        pregunta_id
                    ))

                else:
                    
                    cur.execute("""
                            SELECT COALESCE(MAX(norden), 0) + 1
                            FROM preguntas
                            WHERE quiz_id = %s
                        """, (quiz_id,))

                    nuevo_orden = cur.fetchone()[0]

                    cur.execute("""
                        INSERT INTO preguntas (
                            quiz_id,
                            texto,
                            tipo,
                            explicacion,
                            norden
                        )
                        VALUES (%s, %s, %s, %s, %s)
                        RETURNING id
                    """, (
                        quiz_id,
                        texto,
                        tipo,
                        request.form.get(f"explicacion_{num}"),
                        nuevo_orden
                    ))
                    pregunta_id = cur.fetchone()[0]

                if tipo == "vf":
                    correcta = request.form.get(f"correcta_{num}")
                    explicacion = request.form.get(f"explicacion_{num}")
                    
                    opcion_id_v = request.form.get(f"opcion_id_{num}_verdadero")
                    opcion_id_f = request.form.get(f"opcion_id_{num}_falso")

                    if opcion_id_v:
    
                        cur.execute("""
                            UPDATE opciones
                            SET texto=%s,
                                es_correcta=%s
                            WHERE id=%s
                        """, (
                            "Verdadero",
                            correcta == "Verdadero",
                            opcion_id_v
                        ))

                    else:

                        cur.execute("""
                            INSERT INTO opciones (pregunta_id, texto, es_correcta)
                            VALUES (%s, %s, %s)
                        """, (
                            pregunta_id,
                            "Verdadero",
                            correcta == "Verdadero"
                        ))

                    if opcion_id_f:
    
                        cur.execute("""
                            UPDATE opciones
                            SET texto=%s,
                                es_correcta=%s
                            WHERE id=%s
                        """, (
                            "Falso",
                            correcta == "Falso",
                            opcion_id_f
                        ))

                    else:

                        cur.execute("""
                            INSERT INTO opciones (pregunta_id, texto, es_correcta)
                            VALUES (%s, %s, %s)
                        """, (
                            pregunta_id,
                            "Falso",
                            correcta == "Falso"
                        ))

                    cur.execute(
                        "UPDATE preguntas SET explicacion=%s WHERE id=%s",
                        (explicacion, pregunta_id)
                    )

                else:
                    correcta = request.form.get(f"correcta_{num}")

                    for i in range(1, 6):
                        
                        opcion_id = request.form.get(f"opcion_id_{num}_{i}")
                        
                        opcion = request.form.get(f"opcion_{num}_{i}")
                        if opcion:
                            es_correcta = (opcion == correcta) or (str(i) == correcta)

                            if opcion_id:
    
                                cur.execute("""
                                    UPDATE opciones
                                    SET texto=%s,
                                        es_correcta=%s
                                    WHERE id=%s
                                """, (
                                    opcion,
                                    es_correcta,
                                    opcion_id
                                ))

                            else:

                                cur.execute("""
                                    INSERT INTO opciones (pregunta_id, texto, es_correcta)
                                    VALUES (%s, %s, %s)
                                """, (
                                    pregunta_id,
                                    opcion,
                                    es_correcta
                                ))

        conn.commit()
        cur.close()
        conn.close()

        return redirect(url_for('dashboard_profesor'))

    # 🔹 GET (lo que ya tenías)
    cur.execute("""
        SELECT titulo,
        multiple_intentos,
        enviar_solucionario,
        publico,
        config_json
    FROM quiz
    WHERE id=%s
    """, (quiz_id,))

    row = cur.fetchone()
    
    config_json = json.loads(row[4]) if row[4] else obtener_config_default()
    
    print("ROW =", row)
    print("CONFIG_JSON =", config_json)
    print("multiple BD =", row[1])
    print("solucionario BD =", row[2])
    print("publico BD =", row[3])

   

    quiz = {
        "titulo": row[0],

        "multiple_intentos": config_json.get("multiple_intentos", row[1]),
        "enviar_solucionario": config_json.get("enviar_solucionario", row[2]),
        "publico": not config_json.get("privado", not row[3]),
        "privado": config_json.get("privado", not row[3]),

        "config_json": config_json
    }   
    
    cur.execute("""
        SELECT id, texto, tipo, explicacion
        FROM preguntas
        WHERE quiz_id=%s
        ORDER BY norden
    """, (quiz_id,))
    preguntas = cur.fetchall()

    data = []

    for p in preguntas:
        cur.execute("""
            SELECT id, texto, es_correcta
            FROM opciones
            WHERE pregunta_id=%s
            ORDER BY id
        """, (p[0],))

        opciones = cur.fetchall()

        data.append({
            'id': p[0],
            'texto': p[1],
            'tipo': p[2],
            'explicacion': p[3],
            'opciones': opciones
        })

    cur.close()
    conn.close()
    
    config = quiz
    
    print(quiz["config_json"])

    return render_template(
        "editar_quiz.html",
        quiz=quiz,
        preguntas=data
    )
    
import random
import string


@app.route("/generar_quiz_desde_imagen", methods=["POST"])
def generar_quiz_desde_imagen():

    if "imagen" not in request.files:
        return jsonify({"error": "No se recibió imagen"}), 400

    imagen = request.files["imagen"]
    prompt = request.form["prompt"]
    titulo = request.form["titulo"]
    config_json = request.form.get("config_json") or json.dumps(obtener_config_default())
    multiple_intentos = str(request.form.get("multiple_intentos")).lower() in ["true", "1", "on"]
    enviar_solucionario = str(request.form.get("enviar_solucionario")).lower() in ["true", "1", "on"]
    publico = str(request.form.get("publico")).lower() in ["true", "1", "on"]
     

    tmp_dir = os.path.abspath(os.path.join("tmp", "ia"))
    os.makedirs(tmp_dir, exist_ok=True)

    extension = os.path.splitext(imagen.filename)[1].lower()
    nombre = f"{uuid.uuid4()}{extension}"

    ruta = os.path.join(tmp_dir, nombre)

    imagen.save(ruta)
    
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO cola_ia (
            prompt,
            estado,
            usuario_id,
            usuario,
            cempre,
            titulo,
            origen,
            contenido_extra,
            config_json,
            multiple_intentos, 
            enviar_solucionario, 
            publico
            
        )
        VALUES (
            %s,
            'pendiente',
            %s,
            %s,
            %s,
            %s,
            'imagen',
            %s,
            %s,
            %s,
            %s,
            %s
        )
       RETURNING id
    """, (
        prompt,
        session["user_id"],
        session["usuario"],
        session["cempre"],
        titulo,
        ruta,
        config_json,
        multiple_intentos, 
        enviar_solucionario, 
        publico
    ))
    
    cola_id = cur.fetchone()[0]

    conn.commit()
    cur.close()
    conn.close()
    

    return jsonify({
        "status": "encolado",
        "cola_id": cola_id
    })

   

def generar_codigo_unico(cur):
    
    while True:

        codigo = ''.join(
            random.choices(
                string.ascii_uppercase + string.digits,
                k=6
            )
        )

        # 🔍 ¿Existe en quiz?
        cur.execute(
            "SELECT 1 FROM quiz WHERE codigo=%s",
            (codigo,)
        )

        if cur.fetchone():
            continue

        # 🔍 ¿Existe en salon_quiz?
        cur.execute(
            "SELECT 1 FROM salon_quiz WHERE codigo=%s",
            (codigo,)
        )

        if cur.fetchone():
            continue

        return codigo


@app.route("/usuarios")
def usuarios():

    if "usuario" not in session:
        return redirect("/login")

    if session["rol"] != "root":
        return redirect("/")

    return render_template("usuarios.html")   

@app.route("/api/empresas")
def api_empresas():

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            cempre,
            dempre
        FROM empresa
        WHERE estado = true
        ORDER BY dempre
    """)

    filas = cur.fetchall()

    cur.close()
    conn.close()

    empresas = []

    for fila in filas:
        empresas.append({
            "id": fila[0],
            "nombre": fila[1]
        })

    return jsonify(empresas)

@app.route("/api/usuarios/guardar", methods=["POST"])
def api_guardar_usuario():

    datos = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        if not datos["id"]:

            cur.execute("""
                INSERT INTO usuarios
                (
                    usuario,
                    nombre,
                    apellido,
                    correo,
                    password,
                    rol,
                    cempre
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (
                datos["usuario"],
                datos["nombre"],
                datos["apellido"],
                datos["correo"],
                datos["password"],
                datos["rol"],
                datos["cempre"]
            ))

        conn.commit()

        return jsonify({
            "ok": True
        })

    except Exception as e:

        conn.rollback()

        mensaje = "Ocurrió un error al guardar el usuario."

        if "usuarios_usuario_key" in str(e):
            mensaje = "El nombre de usuario ya existe."

        return jsonify({
            "ok": False,
            "mensaje": mensaje
        }), 400

    finally:

        cur.close()
        conn.close()
@app.route("/api/usuarios")
def api_usuarios():

    # if "usuario" not in session:
    #     return jsonify([])

    # if session["rol"] != "root":
    #     return jsonify([])
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            id,
            usuario,
            nombre,
            apellido,
            correo,
            rol,
            cempre
        FROM usuarios
        ORDER BY usuario
    """)

    filas = cur.fetchall()

    cur.close()
    conn.close()

    usuarios = []

    for fila in filas:
        usuarios.append({
            "id": fila[0],
            "usuario": fila[1],
            "nombre": fila[2],
            "apellido": fila[3],
            "correo": fila[4],
            "rol": fila[5],
            "cempre": fila[6]
        })

    return jsonify(usuarios)
        
@app.route('/generar_codigo_unico/<int:id>')
def generar_codigo(id):

    conn = get_db_connection()
    cur = conn.cursor()

    codigo = generar_codigo_unico(cur)

    cur.execute("UPDATE quiz SET codigo=%s WHERE id=%s", (codigo, id))
    conn.commit()

    cur.close()
    conn.close()

    # 🔥 devolver JSON (NO redirect)
    return jsonify({"codigo": codigo})

@app.route('/eliminar_asignacion/<int:id>', methods=['DELETE'])
def eliminar_asignacion(id):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM salon_quiz WHERE id = %s", (id,))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"status": "ok"})


@app.route('/buscar_alumnos')
def buscar_alumnos():

    texto = request.args.get('q', '').strip()

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id,
               dni,
               nombre,
               apellido,
               correo
        FROM alumnos
        WHERE
            dni ILIKE %s
            OR nombre ILIKE %s
            OR apellido ILIKE %s
        ORDER BY apellido, nombre
        LIMIT 20
    """, (
        f"%{texto}%",
        f"%{texto}%",
        f"%{texto}%"
    ))

    return jsonify(cursor.fetchall())

@app.route('/buscar_alumnos_por_salon')
def buscar_alumnos_por_salon():

    texto = request.args.get('q')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DISTINCT a.id, a.dni, a.nombre, a.apellido, a.correo
        FROM salon s
        JOIN salon_quiz sq ON s.id = sq.salon_id
        JOIN respuestas_alumno ra ON sq.quiz_id = ra.quiz_id
        JOIN alumnos a ON a.id = ra.alumno_id
        WHERE s.descripcion ILIKE %s
    """, (f"%{texto}%",))

    return jsonify(cursor.fetchall())

@app.route('/buscar_salones')
def buscar_salones():

    q = request.args.get('q')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, descripcion
        FROM salon
        WHERE descripcion ILIKE %s
        LIMIT 10
    """, (f"%{q}%",))

    return jsonify(cursor.fetchall())



@app.route('/enviar_quiz/<int:quiz_id>')
def vista_enviar_quiz(quiz_id):

    salon_id = request.args.get('salon_id')

    conn = get_db_connection()
    cursor = conn.cursor()

    # 🔹 quiz
    cursor.execute("""
        SELECT titulo, codigo
        FROM quiz
        WHERE id = %s
    """, (quiz_id,))
    quiz = cursor.fetchone()

    titulo_quiz, codigo_quiz = quiz

    # 🔥 lógica con o sin filtro
    if salon_id:

        cursor.execute("""
            SELECT DISTINCT a.id, a.nombre, a.apellido, a.dni, a.correo
            FROM alumnos a
            JOIN respuestas_alumno r ON r.alumno_id = a.id
            JOIN salon_quiz sq ON sq.quiz_id = r.quiz_id
            WHERE sq.salon_id = %s
              AND r.quiz_id = %s
            ORDER BY a.apellido, a.nombre
        """, (salon_id, quiz_id))

    else:
        # sin filtro → como ya lo tienes
        cursor.execute("""
            SELECT DISTINCT a.id, a.nombre, a.apellido, a.dni, a.correo
            FROM alumnos a
            JOIN respuestas_alumno r ON r.alumno_id = a.id
            WHERE r.quiz_id = %s
            ORDER BY a.apellido, a.nombre
        """, (quiz_id,))

    alumnos = cursor.fetchall()

    # 🔹 lista de salones (para combo)
    cursor.execute("""
        SELECT id, descripcion
        FROM salon
        ORDER BY descripcion
    """)
    salones = cursor.fetchall()

    return render_template(
        'enviar_quiz.html',
        alumnos=alumnos,
        quiz_id=quiz_id,
        titulo_quiz=titulo_quiz,
        codigo_quiz=codigo_quiz,
        salones=salones,
        salon_id_actual=salon_id
    )
    
@app.route('/buscar_alumnos_por_salon_id')
def alumnos_por_salon_id():

    salon_id = request.args.get('id')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DISTINCT a.id, a.dni, a.nombre, a.apellido, a.correo
        FROM salon_quiz sq
        INNER JOIN respuestas_alumno ra ON sq.quiz_id = ra.quiz_id
        INNER JOIN alumnos a ON a.id = ra.alumno_id
        WHERE sq.salon_id = %s
    """, (salon_id,))

    return jsonify(cursor.fetchall())
    
@app.route('/buscar_alumno/<dni>')
def buscar_alumno(dni):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT nombre, apellido,correo FROM alumnos WHERE dni=%s", (dni,))
    alumno = cur.fetchone()

    cur.close()
    conn.close()

    if alumno:
        return {"existe": True, "nombre": alumno[0], "apellido": alumno[1], "correo": alumno[2]}
    else:
        return {"existe": False}

@app.route('/ingresar_quiz', methods=['GET', 'POST'])
def ingresar_quiz():
    if request.method == 'POST':
        codigo = request.form['codigo']
        dni = request.form['dni']

        conn = get_db_connection()
        cur = conn.cursor()

        # buscar quiz por código
        cur.execute("SELECT id FROM quiz WHERE codigo=%s", (codigo,))
        quiz = cur.fetchone()

        if not quiz:
            return "❌ Código inválido"

        quiz_id = quiz[0]

        # buscar alumno por dni
        cur.execute("SELECT id, nombre, apellido,correo FROM alumnos WHERE dni=%s", (dni,))
        alumno = cur.fetchone()

        if alumno:
            alumno_id = alumno[0]
            return redirect(url_for('resolver_quiz', quiz_id=quiz_id, alumno_id=alumno_id))
        else:
            return redirect(url_for('registro_alumno', codigo=codigo, dni=dni))

    return render_template('ingresar_quiz.html')

@app.route('/quiz/<codigo>', methods=['GET', 'POST'])
def acceso_quiz(codigo):

    conn = get_db_connection()
    cur = conn.cursor()
    session['modo'] = 'web'

    # buscar quiz
    cur.execute("SELECT id, titulo, multiple_intentos FROM quiz WHERE codigo=%s", (codigo,))
    quiz = cur.fetchone()    
    
    if not quiz:
        return "❌ Código inválido"

    quiz_id = quiz[0]
    titulo_quiz = quiz[1]
    multiple_intentos = quiz[2]

    if request.method == 'POST':
        dni = request.form['dni']
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        correo = request.form.get('correo')

        # buscar alumno
        cur.execute("SELECT id, nombre, apellido,correo FROM alumnos WHERE dni=%s", (dni,))
        alumno = cur.fetchone()

        if alumno:
            alumno_id = alumno[0]
            correo_db = alumno[3]
            
            cur.execute("""
                SELECT COUNT(*) 
                FROM respuestas_alumno r
                INNER JOIN preguntas p ON p.id = r.pregunta_id
                WHERE r.alumno_id = %s AND p.quiz_id = %s
            """, (alumno_id, quiz_id))

            ya_respondio = cur.fetchone()[0] > 0
    
            if not multiple_intentos and ya_respondio:
                response = make_response(render_template("bloqueado.html", quiz_titulo=titulo_quiz))
                response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
                return response

            # 🔥 si no tiene correo → actualizar
            if not correo_db or not str(correo_db).strip():

                if not correo or not correo.strip():
                    return "❌ Debes ingresar tu correo"

                cur.execute(
                    "UPDATE alumnos SET correo=%s WHERE dni=%s",
                    (correo.strip(), dni)
                )
                conn.commit()
        else:
            if not nombre or not apellido or not correo:
                return "❌ Debes completar nombre, apellido y correo"

            cur.execute(
                "INSERT INTO alumnos (dni, nombre, apellido, correo) VALUES (%s, %s, %s, %s) RETURNING id",
                (dni, nombre, apellido, correo,)
            )
            alumno_id = cur.fetchone()[0]
            conn.commit()

        cur.close()
        conn.close()
         
        return redirect(url_for('resolver_quiz', quiz_id=quiz_id, alumno_id=alumno_id))

    return render_template('login_quiz.html', codigo=codigo)

@app.route('/iniciar_quiz', methods=['POST'])
def iniciar_quiz():

    data = request.get_json()
    salon_quiz_id = data.get("salon_quiz_id")
    quiz_id = data["quiz_id"]
    alumno_id = data["alumno_id"]
    
    conn = get_db_connection()
    cur = conn.cursor()

    # obtener siguiente intento
    cur.execute("""
        SELECT COALESCE(MAX(intento_numero),0)
        FROM intentos_quiz
        WHERE alumno_id=%s
            AND quiz_id=%s
            AND activo=TRUE
    """, (alumno_id, quiz_id))

    ultimo = cur.fetchone()[0]
    nuevo_intento = ultimo + 1

    # crear intento
    cur.execute("""
        INSERT INTO intentos_quiz
            (alumno_id, quiz_id, intento_numero)
        VALUES
            (%s,%s,%s)
        RETURNING id
    """, (alumno_id, quiz_id, nuevo_intento))

    intento_id = cur.fetchone()[0]

    conn.commit()

    cur.close()
    conn.close()

    return jsonify({
        "status": "ok",
        "intento_id": intento_id
    })

@app.route('/resolver_quiz/<int:quiz_id>/<int:alumno_id>')
def resolver_quiz(quiz_id, alumno_id):
    
    salon_quiz_id = None
 

    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        select nombre||' '||apellido 
           from alumnos
           WHERE id=%s
        """, (alumno_id,))
    
    row = cur.fetchone()
    
    alumno_nombre=row[0] if row else "Alumno Prueba"
    
    cur.execute("""
        SELECT config_json, multiple_intentos,titulo
        FROM quiz
        WHERE id = %s
    """, (quiz_id,))

    row = cur.fetchone()

    config_json = json.loads(row[0]) if row and row[0] else {}
    multiple_intentos = row[1] if row else False
    titulo=row[2] if row else "Titulo del quiz"

    # obtener preguntas
    cur.execute("""
        SELECT id, texto, tipo,explicacion
        FROM preguntas
        WHERE quiz_id = %s
    """, (quiz_id,))
    preguntas = cur.fetchall()

    data = []

    for p in preguntas:
        cur.execute("""
            SELECT id, texto
            FROM opciones
            WHERE pregunta_id = %s
        """, (p[0],))

        opciones_raw = cur.fetchall()

        opciones = [
            {
                "id": op[0],       # 🔥 IMPORTANTE
                "texto": op[1]
            }
            for op in opciones_raw
        ]
        
        data.append({
            'id': p[0],
            'texto': p[1],
            'tipo': p[2],
            'explicacion': p[3],
            'opciones': opciones
        })

    cur.close()
    conn.close()

    return render_template(
        'resolver_quiz.html',
        preguntas=data,
        quiz_id=quiz_id,
        alumno_id=alumno_id,
        salon_quiz_id=salon_quiz_id,
        config_json=config_json,
        multiple_intentos=multiple_intentos,
        titulo=titulo,
        alumno_nombre=alumno_nombre
    )
    
@app.route('/guardar_respuestas', methods=['POST'])
def guardar_respuestas():

    data = request.get_json()

    alumno_id = data['alumno_id']
    respuestas = data['respuestas']
    preguntas = data.get("preguntas", [])
    salon_quiz_id = data.get('salon_quiz_id')
    
    tiempo_por_pregunta = data.get("tiempo_por_pregunta", {})

    conn = get_db_connection()
    cur = conn.cursor()

    correctas = 0
    total = len(preguntas)
    detalle = []
     
    valor_por_pregunta = 20 / total if total > 0 else 0

    # 🔍 obtener quiz_id
    if salon_quiz_id:
        cur.execute("""
            SELECT quiz_id
            FROM salon_quiz
            WHERE id = %s
        """, (salon_quiz_id,))
        
        row = cur.fetchone()
        quiz_id = row[0] if row else None
    else:
        quiz_id = data.get("quiz_id")
        
    if not quiz_id:
        print("⚠️ reconstruyendo quiz_id desde preguntas")

        if preguntas:
    
            primera_pregunta = preguntas[0]

            cur.execute("""
                SELECT quiz_id
                FROM preguntas
                WHERE id = %s
            """, (int(primera_pregunta),))

            row = cur.fetchone()
            quiz_id = row[0] if row else None
        
    intento_id = data["intento_id"] 

    # 🔥 fallback de seguridad (NO rompe nada)
    if preguntas:
    
        primera_pregunta = preguntas[0]

        cur.execute("""
            SELECT quiz_id
            FROM preguntas
            WHERE id = %s
        """, (int(primera_pregunta),))

        row = cur.fetchone()
        quiz_id = row[0] if row else None

    # 🔥 LOOP PRINCIPAL
    for pregunta_id, opcion_id in respuestas.items():

        cur.execute("SELECT texto FROM preguntas WHERE id = %s", (pregunta_id,))
        pregunta_texto = cur.fetchone()[0]

        cur.execute("""
            SELECT id, texto, es_correcta
            FROM opciones
            WHERE pregunta_id = %s
        """, (pregunta_id,))
        opciones = cur.fetchall()

        opciones_detalle = []
        puntaje = 0

        for op in opciones:
            marcada = (op[0] == int(opcion_id))

            if op[2] and marcada:
                puntaje = valor_por_pregunta

            opciones_detalle.append({
                "texto": op[1],
                "correcta": op[2],
                "marcada": marcada
            })

        detalle.append({
            "pregunta": pregunta_texto,
            "opciones": opciones_detalle,
            "puntaje": puntaje
        })

        cur.execute("""
            INSERT INTO respuestas_alumno
            (
                alumno_id,
                pregunta_id,
                opcion_id,
                salon_quiz_id,
                quiz_id,
                intento_id,
                tiempo_segundos
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            alumno_id,
            pregunta_id,
            opcion_id,
            salon_quiz_id,
            quiz_id,
            intento_id,
            tiempo_por_pregunta.get(str(pregunta_id), 0)
        ))
        if puntaje > 0:
            correctas += 1

    conn.commit()

    # 🔹 datos alumno
    cur.execute("""
        SELECT nombre, apellido, dni, correo
        FROM alumnos
        WHERE id = %s
    """, (alumno_id,))
    alumno = cur.fetchone()

    nombre_completo = f"{alumno[0]} {alumno[1]}"
    dni = alumno[2]
    correo = alumno[3]

    # 🔹 título quiz
    cur.execute("""
        SELECT titulo
        FROM quiz q
        JOIN salon_quiz sq ON sq.quiz_id = q.id
        WHERE sq.id = %s
    """, (salon_quiz_id,))
    quiz = cur.fetchone()

    titulo_quiz = quiz[0] if quiz else "Quiz"

    nota = round((correctas / total) * 20, 2)
    fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    cur.execute("""
    UPDATE intentos_quiz
    SET
        nota_final = %s,
        fecha_fin = NOW(),
        tiempo_total_segundos = EXTRACT(EPOCH FROM (NOW() - fecha_inicio))::INTEGER
    WHERE id = %s
    """, (nota, intento_id))
    
    conn.commit()   # ← AGREGAR ESTA LÍNEA
    
    cur.execute("""
        SELECT enviar_solucionario
        FROM quiz
        WHERE id = %s
    """, (quiz_id,))
    
    print("DEBUG quiz_id:", quiz_id)
    

    row = cur.fetchone()
    valor = row[0] if row else False

    enviar_solucionario = True if valor in [True, 't', 'true', '1', 1] else False
    
    print("DEBUG enviar_solucionario:", enviar_solucionario)

    cur.close()
    conn.close()
    
    print("DEBUG RAW DB:", valor)
    print("DEBUG NORMALIZADO:", enviar_solucionario)

    # 🚀 BACKGROUND TASK
    threading.Thread(
        target=generar_y_enviar_reporte,
        args=(detalle, nota, correo, nombre_completo, alumno_id, titulo_quiz, dni, fecha_hora, enviar_solucionario)
    ).start()

    # ⚡ RESPUESTA RÁPIDA
    return jsonify({
        "status": "ok",
        "correctas": correctas,
        "total": total,
        "nota": nota
    })
    
    
@app.route('/enviar_reporte_manual', methods=['POST'])
def enviar_reporte_manual():

    data = request.get_json()

    alumno_id = data['alumno_id']
    quiz_id = data['quiz_id']
    intento_id = data['intento_id']

    try:
        examen = obtener_examen_alumno(alumno_id, quiz_id, intento_id)

        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

        generar_y_enviar_reporte(
            examen["detalle"],
            examen["nota"],
            examen["correo"],
            examen["nombre"],
            alumno_id,
            examen["titulo"],
            examen["dni"],
            fecha,
            True
        )

        return jsonify({"status": "ok"})

    except Exception as e:
        print("❌ ERROR INTERNO:", str(e))

        # 👇 ESTO ES CLAVE
        return jsonify({"status": "error", "mensaje": str(e)})
    
@app.route('/exportar_quiz_excel/<int:quiz_id>')
def exportar_quiz_excel(quiz_id):

    con_sol = request.args.get("solucion", "true").lower() == "true"
    titulo=""

    conn = get_db_connection()
    cur = conn.cursor()
    
    # 🔹 título
    cur.execute("SELECT titulo FROM quiz WHERE id = %s", (quiz_id,))
    titulo_quiz = cur.fetchone()[0]
    
    if con_sol:
        titulo = titulo_quiz + " - Con Solución"
    else:
        titulo = titulo_quiz + " - Sin Solución"

    # 🔹 preguntas + opciones
    cur.execute("""
        SELECT p.id, p.texto, o.texto, o.es_correcta
        FROM preguntas p
        JOIN opciones o ON p.id = o.pregunta_id
        WHERE p.quiz_id = %s
        ORDER BY p.id, o.id
    """, (quiz_id,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    # 🔹 agrupar preguntas
    preguntas = defaultdict(list)
    for r in rows:
        preguntas[(r[0], r[1])].append({
            "opcion": r[2],
            "correcta": r[3]
        })

    # 📊 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz"

    # 🔹 columnas
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 60

    # =========================
    # 🔥 HEADER
    # =========================

    # 🔹 LOGO en A1:B2
    try:
        logo = XLImage("static/img/logo.png")

        base_width = 70
        ratio = base_width / logo.width
        logo.width = base_width
        logo.height = int(logo.height * ratio)

        ws.add_image(logo, "A1")
    except:
        pass

    # 🔹 TÍTULO en A3:B3
    ws.merge_cells('A3:B3')
    celda_titulo = ws.cell(row=3, column=1, value=titulo_quiz)
    celda_titulo.font = Font(size=14, bold=True)
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")

    # 🔹 línea separadora (fila 4)
    thin = Side(style='thin')
    for col in range(1, 3):
        ws.cell(row=4, column=col).border = Border(bottom=thin)

    # =========================
    # 🔹 CONTENIDO
    # =========================

    fila = 5
    letras = ['a', 'b', 'c', 'd', 'e', 'f']
    contador = 1

    for (pid, texto_pregunta), opciones in preguntas.items():

        # 🔹 PREGUNTA
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)

        texto = f"{contador}. {texto_pregunta}"
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(wrap_text=True)

        fila += 1

        # 🔹 OPCIONES
        for i, op in enumerate(opciones):
            letra = letras[i] if i < len(letras) else f"{i})"
            texto_op = f"   {letra}) {op['opcion']}"

            if con_sol and op["correcta"]:
                texto_op += "  ✔"

            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)

            celda_op = ws.cell(row=fila, column=1, value=texto_op)
            celda_op.alignment = Alignment(wrap_text=True)

            fila += 1

        fila += 1
        contador += 1

    # 🔹 guardar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{titulo}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )   
   
@app.route('/salones')
def ver_salones():

    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # =========================================================
    # CONTROL DE LICENCIA
    # =========================================================

    if session.get('rol') != 'root':

        if not session.get('licencia_activa', False):
            return redirect(url_for('home'))
        
        

    user_id = session['user_id']
    rol = session['rol']
    cempre = session['cempre']

    conn = get_db_connection()
    cur = conn.cursor()

    # 🧠 lógica según rol
    if rol == 'admin':
        filtro = "WHERE s.estado='A' AND s.cempre = %s"
        params = (cempre,)
    else:
        # profesor
        filtro = "WHERE s.estado='A' AND s.cempre = %s AND s.usuario_id = %s"
        params = (cempre, user_id)

    query = f"""
        SELECT 
            s.id, 
            s.codigo, 
            s.descripcion, 
            s.fecha_creacion, 
            s.estado,
            COUNT(sq.id) as total_quizzes
        FROM salon s
        LEFT JOIN salon_quiz sq ON sq.salon_id = s.id
        {filtro}
        GROUP BY s.id
        ORDER BY s.id DESC
    """

    cur.execute(query, params)
    salones = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("salones.html", salones=salones)

@app.route('/editar_salon/<int:id>', methods=['GET', 'POST'])
def editar_salon(id):

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        codigo = request.form['codigo']
        descripcion = request.form['descripcion']
        estado = request.form['estado']

        cur.execute("""
            UPDATE salon
            SET codigo = %s,
                descripcion = %s,
                estado = %s
            WHERE id = %s
        """, (codigo, descripcion, estado, id))

        conn.commit()
        cur.close()
        conn.close()

        return redirect('/salones')

    # 🔍 cargar datos actuales
    cur.execute("""
        SELECT id, codigo, descripcion, estado
        FROM salon
        WHERE id = %s
    """, (id,))

    salon = cur.fetchone()

    cur.close()
    conn.close()

    return render_template("editar_salon.html", salon=salon)

@app.route('/eliminar_salon/<int:id>')
def eliminar_salon(id):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE salon
        SET estado = 'I'
        WHERE id = %s
    """, (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect('/salones') 

@app.route('/guardar_salon', methods=['POST'])
def guardar_salon():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    codigo = request.form['codigo']
    descripcion = request.form['descripcion']
    usuario = session['usuario']
    user_id = session['user_id']
    cempre = session['cempre']
    
    print ("Empresa:",cempre)

    if not codigo or not descripcion:
        return "Campos obligatorios", 400

    conn = get_db_connection()
    cur = conn.cursor()
    
    # 🔍 Verificar código duplicado
    cur.execute("""
        SELECT 1
        FROM salon
        WHERE UPPER(codigo)=UPPER(%s)
    """, (codigo,))

    if cur.fetchone():

        cur.close()
        conn.close()

        return jsonify({
            "status": "error",
            "campo": "codigo",
            "mensaje": "El código del salón ya existe. Registre otro código."
        })

    cur.execute("""
        INSERT INTO salon (codigo, descripcion,usuario, usuario_id, cempre, estado)
        VALUES (%s, %s, %s, %s, %s,'A')
    """, (codigo, descripcion,usuario, user_id, cempre))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "status": "ok"
    })

@app.route('/asignar_quiz')
@requiere_licencia
def asignar_quiz():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    rol = session['rol']
    cempre = session['cempre']

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔹 SALONES
    if rol == 'admin':
        cur.execute("""
            SELECT id, codigo, descripcion
            FROM salon
            WHERE cempre = %s
            ORDER BY id DESC
        """, (cempre,))
    else:
        cur.execute("""
            SELECT id, codigo, descripcion
            FROM salon
            WHERE cempre = %s AND usuario_id = %s
            ORDER BY id DESC
        """, (cempre, user_id))

    salones = cur.fetchall()

    # 🔹 QUIZZES
    if rol == 'admin':
        cur.execute("""
            SELECT id, titulo
            FROM quiz
            WHERE cempre = %s
            ORDER BY id DESC
        """, (cempre,))
    else:
        cur.execute("""
            SELECT id, titulo
            FROM quiz
            WHERE cempre = %s AND usuario_id = %s
            ORDER BY id DESC
        """, (cempre, user_id))

    quizzes = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("asignar_quiz.html", salones=salones, quizzes=quizzes,rol=rol)

from psycopg2 import errors

@app.route('/guardar_asignacion', methods=['POST'])
def guardar_asignacion():

    data = request.get_json()

    salon_id = data['salon_id']
    quiz_id = data['quiz_id']
    
    cempre=session['cempre']

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        codigo = generar_codigo_unico_salon(cur)

        cur.execute("""
            INSERT INTO salon_quiz (
                salon_id,
                quiz_id,
                codigo,
                cempre,
                estado
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                'A'
            )
        """, (
            salon_id,
            quiz_id,
            codigo,
            cempre
        ))

        conn.commit()

    except Exception as e:
        conn.rollback()

        # 🔥 detectar duplicado
        if "uq_salon_quiz" in str(e):
            return jsonify({
                "status": "error",
                "message": "Este quiz ya fue asignado a este salón"
            })

        return jsonify({
            "status": "error",
            "message": "Error inesperado"
        })

    finally:
        cur.close()
        conn.close()

    return jsonify({"status": "ok"})

@app.route('/resultados_salon', methods=['GET', 'POST'])
@requiere_licencia
def resultados_salon():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    rol = session['rol']
    cempre = session['cempre']


    conn = get_db_connection()
    cur = conn.cursor()

    salon_id = request.form.get('salon_id') or request.args.get('salon_id')

    # 🔹 SALONES FILTRADOS
    if rol == 'admin':
        cur.execute("""
            SELECT id, codigo, descripcion
            FROM salon
            WHERE cempre = %s
            ORDER BY id DESC
        """, (cempre,))
    else:
        cur.execute("""
            SELECT id, codigo, descripcion
            FROM salon
            WHERE cempre = %s AND usuario_id = %s
            ORDER BY id DESC
        """, (cempre, user_id))

    salones = cur.fetchall()

    quizzes = []
    tabla = []
    notas = []
    nombres = []
    preguntas_top = []
    promedio_quiz = []
    aprobados = 0
    desaprobados = 0
    cantidad_alumnos = 0
    cantidad_evaluaciones = 0
    intentos = {}
    intentos_detalle = {}
    detalle_intentos = {}
    codigo=""
    descripcion=""
    fecha_creacion=""
    
    tiempo_promedio = "00:00"

    distribucion = [0, 0, 0, 0]

    distribucion_detalle = {
        "0-10": [],
        "11-13": [],
        "14-16": [],
        "17-20": []
    }

    verdes = 0
    amarillos = 0
    rojos = 0

    detalle_aprobados = {
        "Aprobados": [],
        "Desaprobados": []
    }
    distribucion= []
    distribucion_detalle = []

    # 🔐 VALIDAR SALON_ID (CLAVE)
    if salon_id:

        if rol == 'admin':
            cur.execute("""
                SELECT id,codigo,descripcion,fecha_creacion FROM salon
                WHERE id = %s AND cempre = %s
            """, (salon_id, cempre))
        else:
            cur.execute("""
                SELECT id,codigo,descripcion,fecha_creacion FROM salon
                WHERE id = %s AND cempre = %s AND usuario_id = %s
            """, (salon_id, cempre, user_id))
            
        salon = cur.fetchone()

        if not salon:
            cur.close()
            conn.close()
            return "No autorizado", 403

        codigo = salon[1]
        descripcion = salon[2]
        fecha_creacion = salon[3].strftime("%d/%m/%Y")
        
        cur.execute("""
            select count(distinct(alumno_id) ) cantidad
                from respuestas_alumno
                inner join salon_quiz sq on sq.id=salon_quiz_id
                where salon_id=%s      AND
                      cempre=%s   
            """, (salon_id, cempre ))
        
        cantidad_alumnos=cur.fetchone()[0]

    # 🔹 PROCESO SOLO SI ES POST
    if request.method == 'POST' and salon_id:

# ==========================================================
# 1. OBTENER RESULTADOS DE LOS ALUMNOS
# ==========================================================
        cur.execute("""
                WITH ultimo_intento AS (

                SELECT
                    ra.alumno_id,
                    ra.quiz_id,
                    ra.salon_quiz_id,
                    MAX(ra.intento_id) AS intento_id

                FROM respuestas_alumno ra

                JOIN intentos_quiz iq
                    ON iq.id = ra.intento_id
                AND iq.activo = TRUE
                where   ra.salon_quiz_id IS NOT NULL

                GROUP BY
                    ra.alumno_id,
                    ra.quiz_id,
                    ra.salon_quiz_id
            ),cantidad_intentos AS (

            SELECT
                ra.alumno_id,
                ra.quiz_id,
                ra.salon_quiz_id,
                count(distinct ra.intento_id) AS cantidad

            FROM respuestas_alumno ra

            JOIN intentos_quiz iq
                ON iq.id = ra.intento_id
            AND iq.activo = TRUE
            where   ra.salon_quiz_id IS NOT NULL

            GROUP BY
                ra.alumno_id,
                ra.quiz_id,
                ra.salon_quiz_id
        )  

                SELECT 
                    a.dni,
                    a.id AS alumno_id,
                    a.nombre,
                    a.apellido,
                    a.correo,
                    ra.salon_quiz_id,
                    q.id AS quiz_id,
                    q.titulo,
                    iq.intento_numero,
                    ci.cantidad AS cantidad_intentos,
                    ROUND(
                        (
                            COUNT(CASE WHEN o.es_correcta THEN 1 END)::decimal
                            / NULLIF(COUNT(*), 0)
                        ) * 20,
                    2) AS nota,
                    MAX(sq.fecha_asignacion) as fecha

                FROM respuestas_alumno ra

                JOIN ultimo_intento ui
                    ON ui.alumno_id = ra.alumno_id
                AND ui.quiz_id = ra.quiz_id
                AND ui.salon_quiz_id = ra.salon_quiz_id
                AND ui.intento_id = ra.intento_id

                JOIN alumnos a ON a.id = ra.alumno_id
                JOIN opciones o ON o.id = ra.opcion_id
            JOIN quiz q
                    ON q.id = ra.quiz_id

                JOIN salon_quiz sq
                    ON sq.id = ra.salon_quiz_id

                JOIN salon s
                    ON s.id = sq.salon_id
                LEFT JOIN intentos_quiz iq ON iq.id = ra.intento_id  AND iq.activo=TRUE
                LEFT JOIN cantidad_intentos ci
                        ON ci.alumno_id = ra.alumno_id
                    AND ci.quiz_id = ra.quiz_id
                    AND ci.salon_quiz_id = ra.salon_quiz_id

                WHERE sq.salon_id = %s
                AND s.cempre = %s

                GROUP BY 
                a.id, 
                a.dni, 
                a.nombre, 
                a.apellido, 
                q.id, 
                q.titulo, 
                iq.intento_numero,
				 cantidad,
                ra.salon_quiz_id

                ORDER BY a.nombre
            """, (salon_id, cempre ))

        data = cur.fetchall()
    
        
        cur.execute("""
            SELECT
                ra.alumno_id,
                ra.salon_quiz_id,
                ra.intento_id,
                iq.intento_numero,
                ROUND(
                    (
                        COUNT(CASE WHEN o.es_correcta THEN 1 END)::decimal
                        / NULLIF(COUNT(*),0)
                    ) * 20,
                2) AS nota,
                iq.fecha_inicio,
                iq.fecha_fin        

            FROM respuestas_alumno ra

            JOIN intentos_quiz iq
                ON iq.id = ra.intento_id  AND iq.activo=TRUE

            JOIN opciones o
                ON o.id = ra.opcion_id

            JOIN salon_quiz sq
                ON sq.id = ra.salon_quiz_id

            WHERE sq.salon_id = %s 
            AND ra.salon_quiz_id IS NOT NULL

            GROUP BY
                ra.alumno_id,
                ra.salon_quiz_id,
                ra.intento_id,
                iq.intento_numero,
                iq.fecha_inicio,
                iq.fecha_fin

            ORDER BY
                ra.alumno_id,
                iq.intento_numero,
                ra.salon_quiz_id
        """, (salon_id,))
        
       

        data_intentos = cur.fetchall()
        
        detalle_intentos = defaultdict(dict)

# ==========================================================
# 2. OBTENER QUIZZES DEL SALÓN
# ==========================================================
        cur.execute("""
            SELECT
                sq.id,
                sq.codigo,
                q.id,
                q.titulo
            FROM salon_quiz sq
            JOIN quiz q
                ON q.id = sq.quiz_id
            JOIN salon s
                ON s.id = sq.salon_id
            WHERE sq.salon_id = %s
            AND s.cempre = %s
            ORDER BY sq.id
        """, (salon_id, cempre))

        todos_quizzes = cur.fetchall()
        
        cur.execute("""
            select count(distinct codigo) cantidad  from salon_quiz
                where salon_id=%s and
                      cempre=%s
        """,(salon_id,cempre))
        
        cantidad_evaluaciones=cur.fetchone()[0]

# ==========================================================
# 3. CONSTRUIR PIVOT DE RESULTADOS
# ==========================================================
        resultado = {}
        
        for row in data_intentos:
    
            alumno_id, salon_quiz_id,intento_id, intento, nota, fecha_inicio, fecha_fin = row
            
            if fecha_inicio and fecha_fin:
                segundos = int((fecha_fin - fecha_inicio).total_seconds())
                minutos = segundos // 60
                segundos = segundos % 60
                tiempo = f"{minutos:02d}:{segundos:02d}"
            else:
                tiempo = "--:--"

            detalle_intentos[alumno_id].setdefault(intento, {})

            detalle_intentos[alumno_id][intento][salon_quiz_id] = {
                "nota": nota,
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
                "tiempo": tiempo
            }
            
             

        for row in data:
             
    
            dni, alumno_id, nombre, apellido, correo, salon_quiz_id, quiz_id, quiz, intento, cantidad_intentos, nota, fecha = row = row

            if alumno_id not in resultado:
                resultado[alumno_id] = {
                    "alumno_id": alumno_id,
                    "dni": dni,
                    "alumno": f"{nombre} {apellido}",
                    "correo": correo
                }

            
            resultado[alumno_id][salon_quiz_id] = {
                "nota": nota,
                "intento": intento,
                "cantidad_intentos": cantidad_intentos,  # activos
                "fecha": fecha,
                "quiz_id": quiz_id
            }

        quizzes = [
            {
                "id": q[0],          # 👈 ahora es salon_quiz_id
                "codigo": q[1],      # 👈 código de salon_quiz
                "quiz_id": q[2],     # 👈 lo guardamos porque lo necesitaremos
                "titulo": q[3]
            }
            for q in todos_quizzes
        ]
        
# ==========================================================
# para los intentos por linea
# ==========================================================
        intentos = {}
        intentos_detalle = {}
# ==========================================================
# 4. CONSTRUIR TABLA PARA LA VISTA
# ==========================================================

        tabla = []

        for alumno_id, fila in resultado.items():

            suma = 0
            evaluaciones = 0

            for quiz in quizzes:

                celda = fila.get(quiz["id"])
                notas = []

                if alumno_id in detalle_intentos:

                    for intento in detalle_intentos[alumno_id]:

                        if quiz["id"] in detalle_intentos[alumno_id][intento]:
                            notas.append(
                                float(detalle_intentos[alumno_id][intento][quiz["id"]]["nota"])
                            )

                if notas:
                    celda["nota"] = round(sum(notas) / len(notas), 2)

                if isinstance(celda, dict):

                    fila[quiz["id"]] = celda

                    if celda["nota"] > 0:
                        suma += celda["nota"]
                        evaluaciones += 1

                else:

                    fila[quiz["id"]] = {
                        "nota": 0,
                        "intento": 0
                    }

            fila["promedio"] = round(
                suma / evaluaciones, 2
            ) if evaluaciones else 0

            tabla.append(fila)
            
            
 
            
# ==========================================================
# 5. CALCULAR KPIs Y GRÁFICOS
# ==========================================================

        cur.execute("""
            SELECT
                ROUND(AVG(iq.tiempo_total_segundos))
            FROM intentos_quiz iq
            JOIN salon_quiz sq
                ON sq.quiz_id = iq.quiz_id
            WHERE sq.salon_id = %s
            AND iq.tiempo_total_segundos IS NOT NULL
            AND iq.nota_final IS NOT NULL
            AND iq.activo=TRUE
        """, (salon_id,))
        
        tiempo_promedio = cur.fetchone()[0] or 0
        minutos = int(tiempo_promedio // 60)
        segundos = int(tiempo_promedio % 60)

        tiempo_promedio = f"{minutos:02d}:{segundos:02d}"

        notas = [fila["promedio"] for fila in tabla]
        distribucion = [0, 0, 0, 0]

        distribucion_detalle = {
            "0-10": [],
            "11-13": [],
            "14-16": [],
            "17-20": []
        }

        for fila in tabla:

            promedio = fila["promedio"]
            alumno = f'{fila["alumno"]} ({promedio:.2f})'

            if promedio <= 10:
                distribucion[0] += 1
                distribucion_detalle["0-10"].append(alumno)

            elif promedio <= 13:
                distribucion[1] += 1
                distribucion_detalle["11-13"].append(alumno)

            elif promedio <= 16:
                distribucion[2] += 1
                distribucion_detalle["14-16"].append(alumno)

            else:
                distribucion[3] += 1
                distribucion_detalle["17-20"].append(alumno)
                
        aprobados = sum(1 for nota in notas if nota >= 11)
        desaprobados = sum(1 for nota in notas if nota < 11)
        
        verdes = sum(1 for nota in notas if nota >= 17)
        amarillos = sum(1 for nota in notas if 11 <= nota < 17)
        rojos = sum(1 for nota in notas if nota < 11)
        
        detalle_aprobados = {
            "Aprobados": [],
            "Desaprobados": []
        }

        for fila in tabla:

            alumno = f'{fila["alumno"]} ({fila["promedio"]:.2f})'

            if fila["promedio"] >= 11:
                detalle_aprobados["Aprobados"].append(alumno)
            else:
                detalle_aprobados["Desaprobados"].append(alumno)
                
        nombres = [fila["alumno"] for fila in tabla]
        
# ==========================================================
# 6. CALCULAR ESTADÍSTICAS POR QUIZ
# ==========================================================
         
        cur.execute("""
            WITH ultimo_intento AS (
                SELECT
                    alumno_id,
                    quiz_id,
                    salon_quiz_id,
                    MAX(intento_id) AS intento_id
                FROM respuestas_alumno
                WHERE  salon_quiz_id IS NOT NULL
                GROUP BY
                    alumno_id,
                    quiz_id,
                    salon_quiz_id
            ),

            notas_quiz AS (

                SELECT
                    ra.alumno_id,
                    ra.quiz_id,

                    ROUND(
                        (
                            COUNT(CASE WHEN o.es_correcta THEN 1 END)::decimal
                            / NULLIF(COUNT(*),0)
                        ) * 20,
                    2) AS nota

                FROM respuestas_alumno ra

                JOIN ultimo_intento ui
                    ON ui.alumno_id = ra.alumno_id
                AND ui.quiz_id = ra.quiz_id
                AND ui.salon_quiz_id = ra.salon_quiz_id
                AND ui.intento_id = ra.intento_id
                AND ra.salon_quiz_id IS NOT NULL
                JOIN opciones o
                    ON o.id = ra.opcion_id

                GROUP BY
                    ra.alumno_id,
                    ra.quiz_id,
                    ra.salon_quiz_id
            )

            SELECT

                sq.id,
                sq.codigo,
                q.titulo,
                MAX(sq.fecha_asignacion),

                ROUND(AVG(nq.nota),2) AS promedio,

                COUNT(CASE WHEN nq.nota >= 11 THEN 1 END) AS aprobados,

                COUNT(CASE WHEN nq.nota < 11 THEN 1 END) AS desaprobados,

                MAX(nq.nota) AS maxima,

                MIN(nq.nota) AS minima

            FROM notas_quiz nq

            JOIN quiz q
                ON q.id = nq.quiz_id

            JOIN salon_quiz sq
                ON sq.quiz_id = q.id

            WHERE sq.salon_id = %s

            GROUP BY
                sq.id,
                sq.codigo,
                q.titulo

            ORDER BY promedio DESC
        """, (salon_id,))
        
        rows = cur.fetchall()

        promedio_quiz = []
   
 

        for row in rows:

            quiz_id = row[0]

            notas = []

            for fila in tabla:

                celda = fila.get(quiz_id)

                if isinstance(celda, dict):
                    notas.append(float(celda["nota"]))

            if notas:

                promedio = round(sum(notas) / len(notas), 2)
                maxima = max(notas)
                minima = min(notas)
                aprobados = len([n for n in notas if n >= 11])
                desaprobados = len([n for n in notas if n < 11])

            else:

                promedio = 0
                maxima = 0
                minima = 0
                aprobados = 0
                desaprobados = 0

            promedio_quiz.append({
                "id": row[0],
                "codigo": row[1],
                "titulo": row[2],
                "fecha": row[3],
                "promedio": promedio,
                "aprobados": aprobados,
                "desaprobados": desaprobados,
                "maxima": maxima,
                "minima": minima
            })


       
    cur.close()
    conn.close()
    

    return render_template(
        "resultados_salon.html",
        salones=salones,
        quizzes=quizzes,
        resultado=tabla,
        codigo =codigo,
        descripcion=descripcion,

        cantidad_alumnos=cantidad_alumnos,
        cantidad_evaluaciones=cantidad_evaluaciones,
        fecha_creacion=fecha_creacion,
        salon_seleccionado=salon_id or "",
        notas=notas,
        aprobados=aprobados,
        desaprobados=desaprobados,
        detalle_aprobados=detalle_aprobados,
        detalle_intentos=detalle_intentos,
        tiempo_promedio=tiempo_promedio,
        nombres=nombres,
        distribucion=distribucion,
        promedio_quiz=promedio_quiz,
        distribucion_detalle=distribucion_detalle,
        verdes=verdes,
        amarillos=amarillos,
        rojos=rojos,
        intentos=intentos,
        preguntas_top=preguntas_top
    )
@app.route('/obtener_alumnos_salon', methods=['POST'])
def obtener_alumnos_salon():

    data = request.json

    salon_id = data.get('salon_id')

    if not salon_id:
        return jsonify({
            "error": "Salón no válido"
        }), 400

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT DISTINCT
            a.id,
            a.nombre,
            a.apellido,
            a.correo
        FROM alumnos a
        JOIN respuestas_alumno ra
            ON ra.alumno_id = a.id
        JOIN salon_quiz sq
            ON sq.id = ra.salon_quiz_id
        WHERE sq.salon_id = %s
        ORDER BY a.apellido, a.nombre
    """, (salon_id,))

    alumnos = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify({
        "ok": True,
        "alumnos": [
            {
                "id": alumno[0],
                "nombre": f"{alumno[1]} {alumno[2]}",
                "correo": alumno[3]
            }
            for alumno in alumnos
        ]
    })
    
def construir_tabla(resultado, quizzes):

    tabla = []

    for  fila in resultado.values():

        suma = 0
        evaluaciones = 0

        for quiz in quizzes:

            celda = fila.get(quiz["id"])

            if isinstance(celda, dict):

                fila[quiz["id"]] = celda

                if celda["nota"] > 0:
                    suma += celda["nota"]
                    evaluaciones += 1

            else:

                fila[quiz["id"]] = {
                    "nota": 0,
                    "intento": 0
                }

        fila["promedio"] = round(
            suma / evaluaciones, 2
        ) if evaluaciones else 0

        tabla.append(fila)

    return tabla
    
@app.route('/eliminar_respuestas', methods=['POST'])
def eliminar_respuestas():
    alumno_id = request.form.get('alumno_id')
    quiz_id = request.form.get('quiz_id')

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔥 borrar SOLO ese quiz del alumno
    cur.execute("""
        DELETE FROM respuestas_alumno
        WHERE alumno_id = %s
        AND (quiz_id = %s OR quiz_id IS NULL)
    """, (alumno_id, quiz_id))

    conn.commit()
    cur.close()
    conn.close()

    salon_id = request.form.get('salon_id')

    return '', 204


@app.route('/obtener_quizzes_alumno/<int:alumno_id>')
def obtener_quizzes_alumno(alumno_id):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT DISTINCT ra.quiz_id, q.titulo
        FROM respuestas_alumno ra
        JOIN quiz q ON q.id = ra.quiz_id
        WHERE ra.alumno_id = %s
    """, (alumno_id,))

    data = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify([
        {"quiz_id": r[0], "titulo": r[1]} for r in data
    ])
    

@app.route('/resultados/<int:quiz_id>')
def ver_resultados(quiz_id):

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔹 RESULTADOS GENERALES
    cur.execute("""
        WITH total_preguntas AS (

            SELECT
                quiz_id,
                COUNT(*) AS total
            FROM preguntas
            GROUP BY quiz_id

        )

        SELECT
            a.id,
            a.dni,
            a.nombre,
            a.apellido,

            COUNT(CASE WHEN o.es_correcta THEN 1 END) AS correctas,

            tp.total,

            ROUND(
                (
                    COUNT(CASE WHEN o.es_correcta THEN 1 END)::decimal
                    / tp.total
                ) * 20,
            2) AS nota

        FROM respuestas_alumno r

        JOIN alumnos a
            ON a.id = r.alumno_id

        JOIN opciones o
            ON o.id = r.opcion_id

        JOIN preguntas p
            ON p.id = r.pregunta_id

        JOIN total_preguntas tp
            ON tp.quiz_id = p.quiz_id

        WHERE p.quiz_id = %s

        GROUP BY
            a.id,
            a.dni,
            a.nombre,
            a.apellido,
            tp.total

        ORDER BY nota DESC
    """, (quiz_id,))

    resultados = cur.fetchall()

    # 🔥 1. TOP 10
    top = resultados[:10]
    #top_nombres = [f"{r[1]} {r[2]}" for r in top]
    top_nombres = [f"{(r[1] + ' ' + r[2])[:10]}" for r in top]
    top_puntajes = [r[3] for r in top]

    # 🔥 2. APROBADOS VS DESAPROBADOS
    Excelente = sum(1 for r in resultados if r[6] >=18)
    MuyBien = sum(1 for r in resultados if r[6] <18 and r[6]>=16)
    Bien = sum(1 for r in resultados if r[6] <16 and r[6]>=13)
    Regular =sum(1 for r in resultados if r[6] <13 and r[6]>=11)
    Mal =sum(1 for r in resultados if r[6] <11)

    aprobados_data = [Excelente, MuyBien,Bien,Regular,Mal]

    # 🔥 3. DISTRIBUCIÓN DE NOTAS
    rangos = {
    "0-2": 0,
    "3-5": 0,
    "6-8": 0,
    "9-10":0,
    "11-12": 0,
    "13-14": 0,
    "15-17": 0,
    "18-20": 0
}

    for r in resultados:
        nota = r[6]
        if nota <= 2:
            rangos["0-2"] += 1
        elif nota <= 5:
            rangos["3-5"] += 1
        elif nota <= 8:
            rangos["6-8"] += 1
        elif nota <= 10:
            rangos["9-10"] += 1
        elif nota <= 12:
            rangos["11-12"] += 1
        elif nota <= 14:
            rangos["13-14"] += 1
        elif nota <= 17:
            rangos["15-17"] += 1
        else:
            rangos["18-20"] += 1

    notas_labels = list(rangos.keys())
    notas_data = list(rangos.values())

    # 🔥 4. PREGUNTAS MÁS FALLADAS (simple demo)
    cur.execute("""
        SELECT 
            p.id,
            p.texto,

            COUNT(DISTINCT CASE 
                WHEN o.es_correcta = false THEN r.alumno_id 
            END) AS alumnos_fallaron,

            COUNT(DISTINCT r.alumno_id) AS total_alumnos,

            ROUND(
                (
                    COUNT(DISTINCT CASE 
                        WHEN o.es_correcta = false THEN r.alumno_id 
                    END)::decimal
                    /
                    NULLIF(COUNT(DISTINCT r.alumno_id), 0)
                ) * 100
            ,1) AS porcentaje_error

        FROM respuestas_alumno r
        JOIN preguntas p ON r.pregunta_id = p.id
        JOIN opciones o ON r.opcion_id = o.id

        WHERE p.quiz_id = %s

        GROUP BY p.id, p.texto

        ORDER BY porcentaje_error DESC
        LIMIT 10;
    """, (quiz_id,))

    preguntas = cur.fetchall()
    preguntas = [p for p in preguntas if p[4] > 0] 
    preguntas_labels = [(p[1] or "Pregunta")[:30] for p in preguntas]
    preguntas_tooltips = [p[1] for p in preguntas]
    preguntas_fallos = [float(p[4]) for p in preguntas]  # 👈 ahora % no conteo
    
    # lista de alumnos
    promedio = round(
        sum(r[6] for r in resultados) / len(resultados),
        2
    ) if resultados else 0
    
    # mayor nota
    mayor_nota = max(r[6] for r in resultados) if resultados else 0
    
    cur.execute("SELECT titulo,codigo,multiple_intentos FROM quiz WHERE id = %s", (quiz_id,))
    resultado = cur.fetchone()

    if resultado is not None:
        titulo_quiz = resultado[0]
        codigo_quiz = resultado[1]
        multiple_intentos=resultado[2]
    else:
        titulo_quiz = "Quiz no encontrado"
        
    cur.execute("""
        SELECT
            alumno_id,
            COUNT(*)
        FROM intentos_quiz
        WHERE quiz_id = %s
        AND nota_final IS NOT NULL
        AND activo=TRUE
        GROUP BY alumno_id
    """, (quiz_id,))

    intentos_por_alumno = cur.fetchall()

    intentos_dict = {row[0]: row[1] for row in intentos_por_alumno}
    
    cur.execute("""
        SELECT COUNT(*) 
        FROM preguntas 
        WHERE quiz_id = %s
    """, (quiz_id,))

    total_preguntas = cur.fetchone()[0]
    
    cur.execute("""
        SELECT
            alumno_id,
            ROUND(AVG(tiempo_total_segundos))
        FROM intentos_quiz
        WHERE quiz_id = %s
        AND tiempo_total_segundos IS NOT NULL
        AND nota_final IS NOT NULL
        AND activo=TRUE
        GROUP BY alumno_id
    """, (quiz_id,))

    tiempos_promedio = {
        alumno_id: tiempo
        for alumno_id, tiempo in cur.fetchall()
    }
    
    cur.execute("""
        SELECT
            iq.alumno_id,
            iq.intento_numero,
            iq.nota_final,
            iq.tiempo_total_segundos,
            iq.fecha_fin,

            COUNT(CASE WHEN o.es_correcta THEN 1 END) AS correctas,
            COUNT(ra.id) AS total

        FROM intentos_quiz iq

        JOIN respuestas_alumno ra
            ON ra.intento_id = iq.id

        JOIN opciones o
            ON o.id = ra.opcion_id

        WHERE iq.quiz_id = %s  AND iq.activo=TRUE

        GROUP BY
            iq.alumno_id,
            iq.intento_numero,
            iq.nota_final,
            iq.tiempo_total_segundos,
            iq.fecha_fin

        ORDER BY
            iq.alumno_id,
            iq.intento_numero
    """, (quiz_id,))

    intentos_detalle = {}

    for alumno_id, intento, nota, tiempo, fecha, correctas, total in cur.fetchall():
    
        if alumno_id not in intentos_detalle:
            intentos_detalle[alumno_id] = []

        intentos_detalle[alumno_id].append({
            "intento": intento,
            "nota": nota,
            "tiempo": tiempo,
            "fecha": fecha,
            "correctas": correctas,
            "total": total
        })
        
     # 🔥 4. KPI's de tiempo
    cur.execute("""
        SELECT
            ROUND(AVG(tiempo_total_segundos)),
            MIN(tiempo_total_segundos),
            MAX(tiempo_total_segundos)
        FROM intentos_quiz
        WHERE quiz_id = %s
            AND tiempo_total_segundos IS NOT NULL
            AND nota_final IS NOT NULL
            AND activo=TRUE
    """, (quiz_id,))

    tiempo_promedio_examen, examen_mas_rapido, examen_mas_lento = cur.fetchone()

    tiempo_promedio_examen = tiempo_promedio_examen or 0
    examen_mas_rapido = examen_mas_rapido or 0
    examen_mas_lento = examen_mas_lento or 0 
    
    cur.execute("""
        SELECT
            COALESCE(SUM(tiempo_total_segundos),0)
        FROM intentos_quiz
        WHERE quiz_id = %s
            AND tiempo_total_segundos IS NOT NULL
            AND nota_final IS NOT NULL
            AND activo=TRUE
    """, (quiz_id,))

    tiempo_total_examenes = cur.fetchone()[0] or 0
    
    cur.execute("""
        SELECT
            p.id,
            p.norden,
            p.texto,
            ROUND(AVG(ra.tiempo_segundos),1) AS promedio
        FROM preguntas p
        JOIN respuestas_alumno ra
            ON ra.pregunta_id = p.id
        WHERE p.quiz_id = %s
        AND ra.tiempo_segundos IS NOT NULL
        GROUP BY
            p.id,
            p.norden,
            p.texto
        ORDER BY promedio DESC;
    """, (quiz_id,))

    tiempo_preguntas = cur.fetchall()
    tiempo_labels = [f"P{p[1]}" for p in tiempo_preguntas]
    tiempo_promedios = [float(p[3]) for p in tiempo_preguntas]
    tiempo_tooltips = [p[2] for p in tiempo_preguntas]
    

    cur.execute("""
        SELECT
            p.norden,
            p.texto,
            COUNT(*) AS respuestas,
            SUM(CASE WHEN o.es_correcta THEN 1 ELSE 0 END) AS correctas,
            ROUND(
                100.0 * SUM(CASE WHEN o.es_correcta THEN 1 ELSE 0 END) / COUNT(*),
                1
            ) AS porcentaje
        FROM preguntas p
        JOIN respuestas_alumno ra
            ON ra.pregunta_id = p.id
        JOIN opciones o
            ON ra.opcion_id = o.id
        WHERE p.quiz_id = %s
        GROUP BY
            p.norden,
            p.texto
        ORDER BY porcentaje ASC;
    """, (quiz_id,))
    
    resultados_aciertos = cur.fetchall()

    aciertos_labels = [f"P{r[0]}" for r in resultados_aciertos]
    aciertos_tooltips = [r[1] for r in resultados_aciertos]
    aciertos_porcentaje = [float(r[4]) for r in resultados_aciertos]
    aciertos_respuestas = [int(r[2]) for r in resultados_aciertos]
    
    cur.execute("""
                select count(distinct(alumno_id) ) cantidad
                    from respuestas_alumno
                    where quiz_id=%s       
                """, (quiz_id, ))
            
    cantidad_alumnos=cur.fetchone()[0]

    cur.close()
    conn.close()
    
     

    return render_template(
        "resultados.html",
        resultados=resultados,
        quiz_id=quiz_id,
        codigo_quiz=codigo_quiz,
        multiple_intentos=multiple_intentos,
        top_nombres=top_nombres,
        top_puntajes=top_puntajes,
        cantidad_alumnos=cantidad_alumnos,
        aprobados_data=aprobados_data,
        notas_labels=notas_labels,
        notas_data=notas_data,
        total_preguntas=total_preguntas,
        preguntas_labels=preguntas_labels,
        preguntas_fallos=preguntas_fallos,
        preguntas_tooltips=preguntas_tooltips,
        intentos=intentos_dict,
        titulo_quiz=titulo_quiz,
        tiempos_promedio=tiempos_promedio,
        promedio=promedio ,
        mayor_nota=mayor_nota,
        intentos_detalle=intentos_detalle,
        tiempo_promedio_examen=tiempo_promedio_examen,
        examen_mas_rapido=examen_mas_rapido,
        examen_mas_lento=examen_mas_lento,
        tiempo_total_examenes=tiempo_total_examenes ,
        tiempo_labels=tiempo_labels,
        tiempo_tooltips=tiempo_tooltips,
        tiempo_promedios=tiempo_promedios ,
        aciertos_labels=aciertos_labels,
        aciertos_porcentaje=aciertos_porcentaje,
        aciertos_tooltips=aciertos_tooltips,
        aciertos_respuestas=aciertos_respuestas 
    )
    
    
@app.route('/ver_quiz_alumno/<int:quiz_id>/<int:alumno_id>')
def ver_quiz_alumno(quiz_id, alumno_id):

    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    cur = conn.cursor()
    
    intento = request.args.get('intento', 1)
    
    cur.execute("""
    SELECT id
        FROM intentos_quiz
        WHERE alumno_id = %s AND quiz_id = %s AND intento_numero = %s  AND activo=TRUE
    """, (alumno_id, quiz_id, intento))

    row = cur.fetchone()
    intento_id = row[0] if row else None
    
    # 🔹 Obtener preguntas + opciones + respuesta del alumno
    cur.execute("""
    SELECT 
        p.texto AS pregunta,
        o.texto AS opcion,
        o.es_correcta,
        o.id AS opcion_id,

        (
            SELECT ra.opcion_id
            FROM respuestas_alumno ra
            WHERE ra.pregunta_id = p.id
            AND ra.alumno_id = %s
            AND (
                ra.intento_id = %s
                OR ra.intento_id IS NULL
            )
            ORDER BY ra.id DESC
            LIMIT 1
        ) AS respuesta_alumno

    FROM preguntas p
    JOIN opciones o ON o.pregunta_id = p.id
    WHERE p.quiz_id = %s
    ORDER BY p.id, o.id
    """, (alumno_id, intento_id, quiz_id))

    data = cur.fetchall()

    # 🔥 AGRUPAR POR PREGUNTA
    preguntas = {}

    for row in data:
        pregunta = row[0]

        if pregunta not in preguntas:
            preguntas[pregunta] = []

        preguntas[pregunta].append({
            "opcion": row[1],
            "es_correcta": row[2],
            "opcion_id": row[3],
            "respuesta_alumno": row[4]
        })
        
    # 🔥 calcular nota
    cur.execute("""
        SELECT 
            ROUND(
                (
                    COUNT(CASE WHEN r.opcion_id = oc.id THEN 1 END)::decimal
                    / COUNT(p.id)
                ) * 20, 2
            )
        FROM preguntas p

        LEFT JOIN LATERAL (
            SELECT ra.opcion_id
            FROM respuestas_alumno ra
            WHERE ra.pregunta_id = p.id
            AND ra.alumno_id = %s
            AND (
                ra.intento_id = %s OR ra.intento_id IS NULL
            )
            ORDER BY ra.id DESC
            LIMIT 1
        ) r ON true

        LEFT JOIN opciones oc 
            ON oc.pregunta_id = p.id AND oc.es_correcta = true

        WHERE p.quiz_id = %s
    """, (alumno_id, intento_id, quiz_id))
    resultado_nota = cur.fetchone()
    nota = resultado_nota[0] if resultado_nota and resultado_nota[0] else 0

    cur.execute("""
        SELECT nombre, apellido
        FROM alumnos
        WHERE id = %s
    """, (alumno_id,))

    alumno = cur.fetchone()

    alumno_nombre = f"{alumno[0]} {alumno[1]}" if alumno else "Alumno"
    
    cur.close()
    conn.close()
    
    

    return render_template(
        "ver_quiz_alumno.html",
        preguntas=preguntas,
        alumno_nombre=alumno_nombre,
        fecha=datetime.now().strftime("%d/%m/%Y"),
        nota=nota
    )

    

from flask import session, redirect

@app.route('/salon/<codigo>', methods=['GET', 'POST'])
def acceso_salon(codigo):

    conn = get_db_connection()
    cur = conn.cursor()
    session['modo'] = 'web'

    # 🔍 buscar salon_quiz
    cur.execute("""
        SELECT sq.id, q.id, q.titulo, q.multiple_intentos
        FROM salon_quiz sq
        JOIN quiz q ON q.id = sq.quiz_id
        WHERE sq.codigo = %s
    """, (codigo,))

    row = cur.fetchone()
    

    if not row:
        return "❌ Código inválido"

    salon_quiz_id = row[0]
    quiz_id = row[1]
    titulo_quiz = row[2]
    multiple_intentos = row[3]

    if request.method == 'POST':
        dni = request.form['dni']
        nombre = request.form.get('nombre')
        apellido = request.form.get('apellido')
        correo = request.form.get('correo')

        # 🔍 buscar alumno
        cur.execute("SELECT id, nombre, apellido, correo FROM alumnos WHERE dni=%s", (dni,))
        alumno = cur.fetchone()

        if alumno:
            alumno_id = alumno[0]
            nombre_db = alumno[1]
            apellido_db = alumno[2]
            correo_db = alumno[3]

            incompleto = (
                not nombre_db or not str(nombre_db).strip() or
                not apellido_db or not str(apellido_db).strip() or
                not correo_db or not str(correo_db).strip()
            )

            # 🔴 SI ESTÁ INCOMPLETO → MOSTRAR FORMULARIO
            if incompleto:
                return render_template(
                    "login_quiz.html",
                    codigo=codigo,
                    dni=dni,
                    nombre=nombre_db or "",
                    apellido=apellido_db or "",
                    correo=correo_db or "",
                    mostrar_form=True
                )
            # 🔥 VALIDACIÓN AQUÍ (ANTES)
            # 🔥 VALIDACIÓN INTENTOS (como ya tienes)
            cur.execute("""
                SELECT COUNT(DISTINCT pregunta_id)
                FROM respuestas_alumno
                WHERE alumno_id = %s AND salon_quiz_id = %s
            """, (alumno_id, salon_quiz_id))

            ya_respondio = cur.fetchone()[0] > 0

            if not multiple_intentos and ya_respondio:
                response = make_response(render_template("bloqueado.html", quiz_titulo=titulo_quiz))
                response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
                return response

            # correo
            if not correo_db or not str(correo_db).strip():

                if not correo or not correo.strip():
                    return "❌ Debes ingresar tu correo"

                cur.execute(
                    "UPDATE alumnos SET correo=%s WHERE dni=%s",
                    (correo.strip(), dni)
                )
                conn.commit()

        else:
            if not nombre or not apellido or not correo:
                return render_template(
                    "login_quiz.html",
                    codigo=codigo,
                    dni=dni,
                    nombre=nombre or "",
                    apellido=apellido or "",
                    correo=correo or "",
                    mostrar_form=True
                )

        # 🔍 verificar si ya existe
        # 🔥 SI NO EXISTE → CREAR
        if not alumno:
            cur.execute("""
                INSERT INTO alumnos (dni, nombre, apellido, correo)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (dni, nombre, apellido, correo))

            alumno_id = cur.fetchone()[0]
            conn.commit()

        cur.close()
        conn.close()

        return redirect(url_for(
            'resolver_quiz_salon',
            salon_quiz_id=salon_quiz_id,
            alumno_id=alumno_id
        ))

    return render_template('login_quiz.html', codigo=codigo)

    
@app.route('/eliminar_quiz/<int:quiz_id>')
def eliminar_quiz(quiz_id):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE quiz 
        SET estado='I', fmodificacion=NOW()
        WHERE id=%s
    """, (quiz_id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for('dashboard_profesor'))

def generar_y_enviar_reporte(detalle, nota, correo, nombre_completo, alumno_id, titulo_quiz, dni, fecha_hora, enviar_solucionario):
    
    limpiar_reportes_temporales()
    
    # 📁 Ruta compatible local + Render
    if os.name == "nt":
        carpeta_tmp = os.path.join(os.getcwd(), "tmp", "reportes")
        os.makedirs(carpeta_tmp, exist_ok=True)
        ruta_pdf = os.path.join(carpeta_tmp, f"reporte_{alumno_id}.pdf")
    else:
        ruta_pdf = f"/tmp/reporte_{alumno_id}.pdf"

    doc = SimpleDocTemplate(ruta_pdf, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # 🔹 Logo (por si falla en Render)
    try:
        logo = Image("static/img/logo.png", width=80, height=50)
    except:
        logo = Paragraph("ACAASO", styles['Normal'])

    info = [[logo, Paragraph(f"""
    <b>ACAASO</b><br/>
    <b>Quiz:</b> {titulo_quiz}<br/>
    <b>Alumno:</b> {nombre_completo}<br/>
    <b>DNI:</b> {dni}<br/>
    <b>Correo:</b> {correo}<br/>
    <b>Fecha:</b> {fecha_hora}
    """, styles['Normal'])]]

    tabla_header = Table(info, colWidths=[100, 350])
    tabla_header.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1.5, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,-1), colors.whitesmoke),
    ]))

    elements.append(tabla_header)
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("<b>RESULTADO DEL QUIZ</b>", styles['Title']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>Nota final: {nota}</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))

    letras = ['a', 'b', 'c', 'd', 'e']

    for i, item in enumerate(detalle, start=1):
        elements.append(Paragraph(f"{i}. {item['pregunta']}", styles['Heading3']))
        elements.append(Spacer(1, 5))

        data_tabla = []
        fila = []

        for j, op in enumerate(item["opciones"]):
            letra = letras[j]
            texto_base = op["texto"]

            if op["correcta"] and op["marcada"]:
                texto = f'<font color="green">({letra}) ✅ {texto_base}</font>'
            elif not op["correcta"] and op["marcada"]:
                texto = f'<font color="red">({letra}) ❌ {texto_base}</font>'
            elif op["correcta"]:
                texto = f'<font color="green">({letra}) ✔ {texto_base}</font>'
            else:
                texto = f'({letra}) {texto_base}'

            fila.append(Paragraph(texto, styles['Normal']))

            if len(fila) == 3:
                data_tabla.append(fila)
                fila = []

        if fila:
            data_tabla.append(fila)

        tabla = Table(data_tabla)
        tabla.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BOX', (0,0), (-1,-1), 1, colors.black),
        ]))

        elements.append(tabla)
        elements.append(Spacer(1, 10))

    # 🔥 Generar PDF
    try:
        doc.build(elements)        
    except Exception as e:       
        return


    if resend is None:
        return
    
    if enviar_solucionario:
        try:
            # 📎 Leer PDF y convertir a base64
            with open(ruta_pdf, "rb") as f:
                pdf_bytes = f.read()

            pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")

            # 📧 Enviar con Resend
            import time

            t0 = time.time()
            resend.Emails.send({
                
                "from": "ACAASO <pardoalf@acaaso.com>",
                "to": correo,
                "subject": "Resultado de tu Quiz",
                "html": f"""
                <div style="font-family: Arial, sans-serif; background-color:#f4f6f8; padding:20px;">
                    
                    <div style="max-width:600px; margin:auto; background:white; border-radius:8px; overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,0.1);">
                        
                        <!-- HEADER -->
                        <div style="text-align:center; padding:20px; background:#0d47a1;">
                            <span style="
                                font-size:42px;
                                font-weight:900;
                                color:#29b6f6;
                                letter-spacing:4px;
                                font-family: Arial Black, Arial, sans-serif;
                            ">
                                ACAASO
                            </span>
                        </div>

                        <!-- BODY -->
                        <div style="padding:25px; color:#333;">
                            
                            <p style="font-size:16px;">
                                Estimado(a) <b>{nombre_completo}</b>,
                            </p>

                            <p>
                                ACAASO Assessment le informa que ya se encuentra disponible el resultado de su evaluación.
                            </p>

                            <div style="background:#f1f5fb; padding:15px; border-radius:6px; margin:20px 0;">
                                <p style="margin:5px 0;"><b>Quiz:</b> {titulo_quiz}</p>
                                <p style="margin:5px 0;"><b>Fecha:</b> {fecha_hora}</p>
                            </div>

                            <div style="text-align:center; margin:25px 0;">
                                <p style="margin-bottom:5px;">Su nota final es:</p>
                                <span style="font-size:28px; font-weight:bold; color:#0d47a1;">
                                    {nota}
                                </span>
                            </div>

                            <p>
                                A continuación, se adjunta el reporte detallado de su evaluación para su revisión.
                            </p>

                            <p style="margin-top:30px;">
                                Cordial saludo,<br>
                                <b>ACAASO Assessment</b>
                            </p>

                        </div>

                        <!-- FOOTER -->
                        <div style="background:#f4f6f8; padding:15px; text-align:center; font-size:12px; color:#777;">
                            © {fecha_hora[:4]} ACAASO - Todos los derechos reservados
                        </div>

                    </div>

                </div>
                """,
                "attachments": [
                    {
                        "filename": f"reporte_{alumno_id}.pdf",
                        "content": pdf_base64
                    }
                ]
            })

            # 🧹 limpiar archivo
            try:
                os.remove(ruta_pdf)
            except:
                pass

        except Exception as e:
            print("❌ ERROR RESEND:", str(e))
    else:
        print("ℹ️ solucionario no enviado")

def limpiar_reportes_temporales():
    if os.name != "nt":
        return

    carpeta_tmp = os.path.join(os.getcwd(), "tmp", "reportes")

    if not os.path.exists(carpeta_tmp):
        return

    ahora = time.time()
    limite = 60 * 60  # 1 hora

    for archivo in os.listdir(carpeta_tmp):

        ruta = os.path.join(carpeta_tmp, archivo)

        if os.path.isfile(ruta) and archivo.lower().endswith(".pdf"):

            try:
                if ahora - os.path.getmtime(ruta) > limite:
                    os.remove(ruta)
                    print(f"🧹 PDF temporal eliminado: {archivo}")

            except Exception as e:
                print(f"⚠️ No se pudo eliminar {archivo}: {e}")
        
@app.route("/contactos", methods=["GET", "POST"])
def contactos():

    if request.method == "POST":

        nombre = request.form["nombre"].strip()
        institucion = request.form["institucion"].strip()
        cargo = request.form["cargo"].strip()
        correo = request.form["correo"].strip()
        whatsapp = request.form["whatsapp"].strip()
        pais = request.form["pais"].strip()
        mensaje = request.form["mensaje"].strip()

        desea_demo = "demo" in request.form

        fecha_demo = request.form.get("fecha_demo") or None
        horario = request.form.get("horario") or None

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO contactos_demo
            (
                nombre,
                institucion,
                cargo,
                correo,
                whatsapp,
                pais,
                mensaje,
                desea_demo,
                fecha_demo,
                horario
            )
            VALUES
            (
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s
            )
        """, (
            nombre,
            institucion,
            cargo,
            correo,
            whatsapp,
            pais,
            mensaje,
            desea_demo,
            fecha_demo,
            horario
        ))

        conn.commit()

        cur.close()
        conn.close()
        
        try:

            resend.Emails.send({

                "from": "ACAASO <pardoalf@acaaso.com>",

               "to": [
                    "pardoalf@acaaso.com",
                    "pardoalf@gmail.com"
                ],

                "reply_to": correo,

                "subject": f"🎯 Nueva solicitud de demostración - {institucion}",

                "html": f"""

                <div style="font-family:Arial;background:#f4f6f8;padding:25px;">

                    <div style="
                        max-width:700px;
                        margin:auto;
                        background:white;
                        border-radius:12px;
                        overflow:hidden;
                        box-shadow:0 4px 15px rgba(0,0,0,.15);
                    ">

                        <div style="
                            background:#0d47a1;
                            color:white;
                            padding:20px;
                            text-align:center;
                        ">

                            <h2 style="margin:0;">
                                Nueva solicitud de demostración
                            </h2>

                        </div>

                        <div style="padding:30px;">

                            <table style="width:100%;font-size:15px;">

                                <tr>
                                    <td style="font-weight:bold;width:180px;">Nombre</td>
                                    <td>{nombre}</td>
                                </tr>

                                <tr>
                                    <td style="font-weight:bold;">Institución</td>
                                    <td>{institucion}</td>
                                </tr>

                                <tr>
                                    <td style="font-weight:bold;">Cargo</td>
                                    <td>{cargo}</td>
                                </tr>

                                <tr>
                                    <td style="font-weight:bold;">Correo</td>
                                    <td>{correo}</td>
                                </tr>

                                <tr>
                                    <td style="font-weight:bold;">WhatsApp</td>
                                    <td>{whatsapp}</td>
                                </tr>

                                <tr>
                                    <td style="font-weight:bold;">País</td>
                                    <td>{pais}</td>
                                </tr>

                            </table>

                            <hr>

                            <h4>Mensaje</h4>

                            <div style="
                                background:#f7f7f7;
                                padding:15px;
                                border-radius:8px;
                            ">

                                {mensaje}

                            </div>

                            <br>

                            <table style="width:100%;">

                                <tr>
                                    <td><b>Solicita demostración</b></td>
                                    <td>{"Sí" if desea_demo else "No"}</td>
                                </tr>

                                <tr>
                                    <td><b>Fecha sugerida</b></td>
                                    <td>{fecha_demo or "-"}</td>
                                </tr>

                                <tr>
                                    <td><b>Horario</b></td>
                                    <td>{horario or "-"}</td>
                                </tr>

                            </table>

                        </div>

                    </div>

                </div>

                """

            })
            
            try:

                resend.Emails.send({

                    "from": "ACAASO <pardoalf@acaaso.com>",

                    "to": correo,

                    "subject": "Hemos recibido su solicitud - ACAASO Teacher Assessment",

                    "html": f"""

                    <div style="font-family:Arial,sans-serif;background:#f4f6f8;padding:20px;">

                        <div style="
                            max-width:650px;
                            margin:auto;
                            background:white;
                            border-radius:10px;
                            overflow:hidden;
                            box-shadow:0 2px 8px rgba(0,0,0,.12);
                        ">

                            <div style="
                                background:#0d47a1;
                                text-align:center;
                                padding:25px;
                            ">

                                <span style="
                                    font-size:40px;
                                    font-weight:900;
                                    color:#29b6f6;
                                    letter-spacing:4px;
                                    font-family:Arial Black,Arial,sans-serif;
                                ">
                                    ACAASO
                                </span>

                                <div style="
                                    color:white;
                                    font-size:15px;
                                    margin-top:8px;
                                ">
                                    Teacher Assessment
                                </div>

                            </div>

                            <div style="padding:35px;">

                                <p style="font-size:16px;">
                                    Estimado(a)
                                    <b>{nombre}</b>,
                                </p>

                                <p>
                                    Muchas gracias por comunicarse con
                                    <b>ACAASO Teacher Assessment</b>.
                                </p>

                                <p>
                                    Hemos recibido correctamente su solicitud de información.
                                </p>

                                <p>
                                    Uno de nuestros especialistas revisará la información enviada y
                                    se comunicará con usted a la brevedad para coordinar una
                                    demostración personalizada de nuestra plataforma.
                                </p>

                                <div style="
                                    background:#eef6ff;
                                    border-left:5px solid #0d47a1;
                                    padding:15px;
                                    margin:25px 0;
                                ">

                                    <b>Institución:</b> {institucion}<br>
                                    <b>Correo:</b> {correo}

                                </div>

                                <p>
                                    Mientras tanto puede visitar nuestro sitio web y conocer
                                    las principales funcionalidades de ACAASO Teacher Assessment.
                                </p>

                                <br>

                                <p>

                                    Cordialmente,

                                    <br><br>

                                    <b>Equipo ACAASO</b>

                                </p>

                            </div>

                            <div style="
                                background:#f1f5f9;
                                text-align:center;
                                padding:15px;
                                font-size:12px;
                                color:#666;
                            ">

                                Este es un correo automático.
                                No es necesario responder este mensaje.

                            </div>

                        </div>

                    </div>

                    """

                })

            except Exception as e:

                print("❌ Error enviando correo al cliente:", e)

        except Exception as e:

            print("❌ Error enviando correo:", e)

        return "", 200

    return render_template(
        "contactos.html",
        ok=request.args.get("ok")
    )

@app.route('/eliminar_intento', methods=['POST'])
def eliminar_intento():

    data = request.get_json()

    alumno_id = data['alumno_id']
    quiz_id = data['quiz_id']
    intento = int(data['intento'])  # 🔥 importante

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # 🔍 obtener intento_id real
        cur.execute("""
            SELECT id
            FROM intentos_quiz
            WHERE alumno_id = %s
                AND quiz_id = %s
                AND intento_numero = %s
                AND activo=TRUE 
        """, (alumno_id, quiz_id, intento))

        row = cur.fetchone()

        if not row:
            return jsonify({"ok": False, "error": "No encontrado"})

        intento_id = row[0]

        # 🔥 borrar respuestas
        cur.execute("""
            UPDATE intentos_quiz
            SET
                activo = FALSE,
                fecha_eliminacion = NOW(),
                eliminado_por = %s
            WHERE id = %s
        """, (session["user_id"], intento_id))

        conn.commit()

        return jsonify({"ok": True})

    except Exception as e:
        print("❌ ERROR eliminar:", str(e))
        return jsonify({"ok": False, "error": str(e)})

    finally:
        cur.close()
        conn.close()


@app.route('/enviar_codigo_quiz', methods=['POST'])
def endpoint_enviar_codigo_quiz():

    data = request.json

    quiz_id = data.get('quiz_id')
    alumnos = data.get('alumnos')  # lista de IDs

    if not quiz_id or not alumnos:
        return jsonify({"error": "Datos incompletos"}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔹 obtener info del quiz
    cur.execute("""
        SELECT titulo, codigo
        FROM quiz
        WHERE id = %s
    """, (quiz_id,))
    quiz = cur.fetchone()

    if not quiz:
        return jsonify({"error": "Quiz no encontrado"}), 404

    titulo_quiz, codigo_quiz = quiz

    # 🔹 obtener alumnos
    cur.execute(f"""
        SELECT id, nombre, apellido, correo
        FROM alumnos
        WHERE id = ANY(%s)
    """, (alumnos,))

    alumnos_data = cur.fetchall()

    enviados = 0

    for a in alumnos_data:
        alumno_id, nombre, apellido, correo = a
        nombre_completo = f"{nombre} {apellido}"

        cur.execute("""
            INSERT INTO cola_email
            (
                destinatario,
                nombre,
                titulo_quiz,
                codigo_quiz
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s
            )
        """, (
            correo,
            nombre_completo,
            titulo_quiz,
            codigo_quiz
        ))

        enviados += 1
        
    conn.commit()

    cur.close()
    conn.close()    

    return jsonify({
        "ok": True,
        "enviados": enviados,
        "mensaje": f"{enviados} correos agregados a la cola de envío."
    })

def enviar_correo_licencia(
        correo,
        nombre_completo,
        plan_nombre,
        codigo_licencia,
        fecha_inicio,
        fecha_expiracion
    ):

    try:

        with open("static/img/banner_enviar3.png", "rb") as archivo:
            banner_base64 = base64.b64encode(
                archivo.read()
            ).decode("utf-8")

        resend.Emails.send({

            "from": "ACAASO <pardoalf@acaaso.com>",

            "to": correo,

            "subject": "Su licencia ha sido generada – ACAASO Teacher Assessment",

            "html": f"""
            <div style="
                font-family: Arial, sans-serif;
                max-width: 650px;
                margin: auto;
                border: 1px solid #e5e7eb;
                border-radius: 12px;
                overflow: hidden;
                color: #334155;
            ">

                <img
                    src="cid:banner_licencia"
                    style="
                        width: 100%;
                        display: block;
                    "
                >

                <div style="padding: 30px;">

                    <h2 style="
                        color: #1e293b;
                        margin-top: 0;
                    ">
                        ¡Su licencia ha sido generada!
                    </h2>

                    <p>
                        Estimado(a) <b>{nombre_completo}</b>,
                    </p>

                    <p>
                        Le informamos que su licencia de
                        <b>ACAASO Teacher Assessment</b>
                        ha sido generada correctamente.
                    </p>

                    <div style="
                        background: #f8fafc;
                        border: 1px solid #e2e8f0;
                        border-radius: 10px;
                        padding: 20px;
                        margin: 25px 0;
                    ">

                        <p>
                            <b>Plan:</b> {plan_nombre}
                        </p>

                        <p>
                            <b>Código de licencia:</b><br>

                            <span style="
                                font-size: 20px;
                                font-weight: bold;
                                color: #2563eb;
                            ">
                                {codigo_licencia}
                            </span>
                        </p>

                        <p>
                            <b>Fecha de inicio:</b>
                            {fecha_inicio}
                        </p>

                        <p>
                            <b>Fecha de vencimiento:</b>
                            {fecha_expiracion}
                        </p>

                    </div>

                    <p><b>Instrucciones para activar su licencia:</b></p>

                    <ol>
                        <li>Ingrese a ACAASO Teacher Assessment.</li>
                        <li>Inicie sesión con su usuario y contraseña.</li>
                        <li>Ingrese al módulo de Licencia.</li>
                        <li>Registre el código de licencia enviado en este correo.</li>
                    </ol>

                    <p>
                        Conserve este correo como respaldo de su licencia.
                    </p>

                    <p>
                        Atentamente,<br>
                        <b>ACAASO Teacher Assessment</b>
                    </p>

                </div>

                <div style="
                    background: #f8fafc;
                    padding: 15px;
                    text-align: center;
                    font-size: 12px;
                    color: #64748b;
                ">
                    © ACAASO Assessment
                </div>

            </div>
            """,

            "attachments": [
                {
                    "filename": "banner_enviar_quiz.png",
                    "content": banner_base64,
                    "content_type": "image/png",
                    "content_id": "banner_licencia"
                }
            ]
        })

        print(
            f"✅ CORREO DE LICENCIA ENVIADO A: {correo}"
        )

        return True


    except Exception as e:

        print(
            "❌ ERROR ENVIANDO CORREO DE LICENCIA:",
            str(e)
        )

        return False
    
def enviar_codigo_quiz(
        correo,
        nombre_completo,
        titulo_quiz,
        codigo_quiz,
        tipo_acceso="quiz"
    ):
    
    #import resend
    
    if tipo_acceso == "salon":
        link_quiz = f"https://teacherassessment.onrender.com/salon/{codigo_quiz}"
    else:
        link_quiz = f"https://teacherassessment.onrender.com/quiz/{codigo_quiz}"
        
    qr_png = generar_qr(link_quiz)
    #banner_url = "https://acaaso-teacherassessment.onrender.com/static/img/banner_enviar_quiz.png"
    #banner_url = "http://localhost:5000/static/img/banner_enviar_quiz.png"
    with open("static/img/banner_enviar3.png", "rb") as f:
        banner_png = f.read() 
    banner_base64 = base64.b64encode(banner_png).decode("utf-8")

    try:
        resend.Emails.send({
            "from": "ACAASO <pardoalf@acaaso.com>",
            "to": correo,
            "subject": "Acceso a evaluación – ACAASO Assessment",
            "html": f"""
                <div style="font-family: Arial, sans-serif; background-color:#f4f6f8; padding:20px;">

                <div style="
                    width:100%;
                    max-width:600px;
                    margin:0 auto;
                    background:white;
                    border-radius:8px;
                    overflow:hidden;
                    box-shadow:0 2px 8px rgba(0,0,0,0.1);
                ">

                    <!-- HEADER -->
                    <div style="
                        position:relative;
                        height:120px;
                        overflow:hidden;
                    ">

                        <img src="cid:banner_quiz"
                            alt="ACAASO"
                            style="
                                width:100%;
                                height:120px;
                                object-fit:cover;
                                display:block;
                            ">
                    </div>

                    <div style="
                        text-align:center;
                        padding:18px 20px 12px 20px;
                        background:white;
                    ">

                        <div style="
                            font-size:42px;
                            font-weight:900;
                            color:#0d47a1;
                            letter-spacing:4px;
                            font-family:Arial Black, Arial, sans-serif;
                            line-height:1;
                        ">
                            ACAASO
                        </div>

                        <div style="
                            margin-top:6px;
                            font-size:16px;
                            letter-spacing:3px;
                            color:#666;
                            font-weight:600;
                        ">
                            Teacher Assessment
                        </div>

                    </div>

                    <!-- BODY -->
                    <div style="padding:25px; color:#333;">
                        <p style="font-size:16px;">
                            Hola <b>{nombre_completo}</b>,
                        </p>

                        <p>Te informamos que tienes una evaluación disponible en la plataforma <b>ACAASO Assessment</b>.</p>

                        <p><b>Detalles del quiz:</b></p>
                        <ul>
                            <li><b>Título:</b> {titulo_quiz}</li>
                            <div style="
                                background:#f1f5fb;
                                border:1px solid #d6e4ff;
                                border-radius:8px;
                                padding:18px;
                                margin:20px 0;
                                text-align:center;
                            ">

                                <div style="font-size:14px; color:#666;">
                                    Código de acceso
                                </div>

                                <div style="
                                    font-size:34px;
                                    font-weight:bold;
                                    color:#0d47a1;
                                    letter-spacing:4px;
                                    margin-top:8px;
                                ">
                                    {codigo_quiz}
                                </div>

                            </div>
                            <div style="text-align:center; margin:25px 0;">

                                <a href="{link_quiz}"
                                style="
                                        display:inline-block;
                                        background:#0d47a1;
                                        color:white;
                                        text-decoration:none;
                                        padding:14px 30px;
                                        border-radius:8px;
                                        font-size:16px;
                                        font-weight:bold;
                                ">
                                    Ingresar al Quiz
                                </a>

                            </div>
                            <p style="font-size:13px; color:#777; text-align:center; margin-top:15px;">
                                Si el botón no funciona, copie y pegue el siguiente enlace en su navegador:
                            </p>

                            <p style="text-align:center; word-break:break-all;">
                                <a href="{link_quiz}">{link_quiz}</a>
                            </p>
                        </ul>
                        <div style="text-align:center; margin:30px 0;">
                            <img src="cid:qr_quiz"
                                alt="Código QR"
                                style="width:220px; height:220px;">

                            <p style="margin-top:15px; font-size:15px; color:#555;">
                                Escanee este código QR para acceder directamente a la evaluación.
                            </p>
                        </div>

                        <p><b>Instrucciones:</b></p>
                        <ol>
                            <li>Ingresa a la plataforma.</li>
                            <li>Introduce el código.</li>
                            <li>Resuelve el quiz.</li>
                        </ol>

                        <p>Atentamente,<br>ACAASO Assessment</p>
                    </div>

                    <!-- FOOTER -->
                    <div style="background:#f4f6f8; padding:15px; text-align:center; font-size:12px; color:#777;">
                        © ACAASO Assessment
                    </div>

                </div>

            </div>
            """,
            "attachments": [
                {
                    "filename": "banner_enviar_quiz.png",
                    "content": banner_base64,
                    "content_type": "image/png",
                    "content_id": "banner_quiz"
                },
                {
                    "filename": "qr_quiz.png",
                    "content": base64.b64encode(qr_png).decode("utf-8"),
                    "content_type": "image/png",
                    "content_id": "qr_quiz"
                }
            ]
        })
         

    except Exception as e:
        print("❌ ERROR RESEND:", str(e))
        raise
 
@app.route('/enviar_acceso_salon', methods=['POST'])
def enviar_acceso_salon():

    data = request.json

    quiz_id = data.get('quiz_id')
    salon_quiz_id = data.get('salon_quiz_id')
    alumnos = data.get('alumnos')

    if not quiz_id or not salon_quiz_id or not alumnos:
        return jsonify({
        "error": "Datos incompletos"
    }), 400

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔹 Obtener título del quiz y código específico del salón
    cur.execute("""
        SELECT
            q.titulo,
            sq.codigo
        FROM salon_quiz sq
        JOIN quiz q
            ON q.id = sq.quiz_id
        WHERE sq.id = %s
        AND q.id = %s
    """, (
        salon_quiz_id,
        quiz_id
    ))

    quiz = cur.fetchone()

    if not quiz:
        cur.close()
        conn.close()

        return jsonify({
            "error": "Quiz o asignación al salón no encontrada"
        }), 404

    titulo_quiz, codigo_quiz = quiz

    # 🔹 Obtener todos los alumnos seleccionados
    cur.execute("""
        SELECT id, nombre, apellido, correo
        FROM alumnos
        WHERE id = ANY(%s)
        ORDER BY apellido, nombre
    """, (alumnos,))

    alumnos_data = cur.fetchall()

    enviados = 0

    # 🔹 Insertar un correo por cada alumno seleccionado
    for alumno in alumnos_data:
    
        _, nombre, apellido, correo = alumno

        nombre_completo = f"{nombre} {apellido}"

        if not correo:
            continue

        cur.execute("""
            INSERT INTO cola_email
            (
                destinatario,
                nombre,
                titulo_quiz,
                codigo_quiz,
                tipo_acceso
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                'salon'
            )
        """, (
            correo,
            nombre_completo,
            titulo_quiz,
            codigo_quiz
        ))

        enviados += 1

    conn.commit()

    cur.close()
    conn.close()

    return jsonify({
        "ok": True,
        "enviados": enviados,
        "mensaje": f"{enviados} correos agregados a la cola de envío."
    })   
        
def generar_qr(texto):
    
    qr = qrcode.QRCode(
        version=1,
        box_size=8,
        border=2
    )
    
    qr.add_data(texto)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")  
    
    buffer.seek(0)

    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8") 
    
    return buffer.getvalue()
    
    
    
    
#=======================================================
#
# PROFESOR
#
#=======================================================
@app.route('/crear_profesor', methods=['POST'])
def crear_profesor():

    if not require_admin():
        return redirect(url_for('login'))

    usuario = request.form['usuario']
    password = request.form['password']
    dni = request.form['dni']
    nombre = request.form['nombre']
    apellido = request.form['apellido']
    correo = request.form['correo']
    cempre = session['cempre']
    
    if not usuario or not password or not dni or not nombre or not apellido:
        return "Campos obligatorios vacíos", 400

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔍 validar usuario único
    cur.execute("""
        SELECT id FROM usuarios WHERE usuario = %s
    """, (usuario,))
    if cur.fetchone():
        return "Usuario ya existe", 400

    # 🔍 validar dni único
    cur.execute("""
        SELECT id FROM usuarios WHERE dni = %s
    """, (dni,))
    if cur.fetchone():
        return "DNI ya existe", 400

    # 🔐 insertar profesor
    cur.execute("""
        INSERT INTO usuarios 
        (usuario, password, rol, dni, nombre, apellido, correo, cempre)
        VALUES (%s, %s, 'profesor', %s, %s, %s, %s, %s)
    """, (usuario, password, dni, nombre, apellido, correo, cempre))

    conn.commit()
    cur.close()
    conn.close()

    return "OK", 200

@app.route('/editar_profesor', methods=['POST'])
def editar_profesor():

    if not require_admin():
        return redirect(url_for('login'))

    profesor_id = request.form['id']
    nombre = request.form['nombre']
    apellido = request.form['apellido']
    correo = request.form['correo']
    dni = request.form['dni']
    cempre = session['cempre']

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔒 validar pertenencia a empresa
    cur.execute("""
        SELECT id FROM usuarios 
        WHERE id = %s AND cempre = %s AND rol = 'profesor'
    """, (profesor_id, cempre))

    if not cur.fetchone():
        return "No autorizado", 403

    # 🔍 validar dni único (excluyendo actual)
    cur.execute("""
        SELECT id FROM usuarios 
        WHERE dni = %s AND id != %s
    """, (dni, profesor_id))

    if cur.fetchone():
        return "DNI duplicado", 400

    cur.execute("""
        UPDATE usuarios
        SET nombre = %s,
            apellido = %s,
            correo = %s,
            dni = %s
        WHERE id = %s AND cempre = %s
    """, (nombre, apellido, correo, dni, profesor_id, cempre))

    conn.commit()
    cur.close()
    conn.close()

    return "OK", 200

@app.route('/eliminar_profesor', methods=['POST'])
def eliminar_profesor():

    if not require_admin():
        return redirect(url_for('login'))

    profesor_id = request.form['id']
    cempre = session['cempre']

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔍 verificar quizzes asociados
    cur.execute("""
        SELECT id FROM quiz 
        WHERE usuario_id = %s AND cempre = %s
    """, (profesor_id, cempre))

    if cur.fetchone():
        return "Profesor tiene quizzes asociados", 400

    # 🔒 eliminar solo si pertenece a empresa
    cur.execute("""
        DELETE FROM usuarios
        WHERE id = %s AND cempre = %s AND rol = 'profesor'
    """, (profesor_id, cempre))

    conn.commit()
    cur.close()
    conn.close()

    return "OK", 200

@app.route('/importar_profesores', methods=['POST'])
def importar_profesores():

    if not require_admin():
        return redirect(url_for('login'))

    archivo = request.files['archivo']
    cempre = session['cempre']

    conn = get_db_connection()
    cur = conn.cursor()

    lineas = archivo.read().decode('utf-8').splitlines()

    insertados = 0
    ignorados = 0

    for linea in lineas:
        try:
            usuario, password, dni, nombre, apellido, correo = linea.split(',')

            # validar duplicados
            cur.execute("""
                SELECT id FROM usuarios 
                WHERE usuario = %s OR dni = %s
            """, (usuario, dni))

            if cur.fetchone():
                ignorados += 1
                continue

            cur.execute("""
                INSERT INTO usuarios
                (usuario, password, rol, dni, nombre, apellido, correo, cempre)
                VALUES (%s, %s, 'profesor', %s, %s, %s, %s, %s)
            """, (usuario, password, dni, nombre, apellido, correo, cempre))

            insertados += 1

        except:
            ignorados += 1

    conn.commit()
    cur.close()
    conn.close()

    return {
        "insertados": insertados,
        "ignorados": ignorados
    }
 
@app.route('/crear_plan', methods=['GET', 'POST'])
def crear_plan():

    if session.get('rol') != 'root':
        return "Acceso denegado", 403

    if request.method == 'POST':

        tipo = request.form['tipo']
        nombre = request.form['nombre']
        precio = request.form['precio']
        admins = request.form['admins']
        profesores = request.form['profesores']
        alumnos = request.form['alumnos']
        quizzes = request.form['quizzes']
        orden = request.form['orden']

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
           INSERT INTO planes (tipo, nombre, precio, admins, profesores, alumnos, quizzes, orden)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (tipo, nombre, precio, admins, profesores, alumnos, quizzes,orden))

        conn.commit()
        cur.close()
        conn.close()

        return redirect('/mantenimiento_planes')

    return render_template('crear_plan.html')   
    
@app.route('/mantenimiento_planes')
@requiere_licencia
def mantenimiento_planes():

    if session.get('rol') != 'root':
        return "Acceso denegado", 403

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, tipo, nombre, precio, admins, profesores, alumnos, quizzes,orden, activo
        FROM planes
        ORDER BY orden
    """)

    planes = cur.fetchall()

    cur.close()
    conn.close()

    return render_template('mantenimiento_planes.html', planes=planes)

@app.route('/actualizar_plan', methods=['POST'])
def actualizar_plan():

    if session.get('rol') != 'root':
        return {"ok": False}

    data = request.get_json()

    id = data['id']
    campo = data['campo']
    valor = data['valor']

    conn = get_db_connection()
    cur = conn.cursor()

    query = f"UPDATE planes SET {campo} = %s WHERE id = %s"
    cur.execute(query, (valor, id))

    conn.commit()
    cur.close()
    conn.close()

    return {"ok": True}

@app.route('/toggle_plan', methods=['POST'])
def toggle_plan():

    if session.get('rol') != 'root':
        return {"ok": False}

    data = request.get_json()

    id = data['id']
    activo = data['activo']

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("UPDATE planes SET activo = %s WHERE id = %s", (activo, id))

    conn.commit()
    cur.close()
    conn.close()

    return {"ok": True}

@app.route('/eliminar_plan', methods=['POST'])
def eliminar_plan():

    if session.get('rol') != 'root':
        return {"ok": False}

    data = request.get_json()
    id = data['id']

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM planes WHERE id = %s", (id,))

    conn.commit()
    cur.close()
    conn.close()

    return {"ok": True}

@app.route('/metricas_db')
@requiere_licencia
def metricas_db():

    if session.get('rol') != 'root':
        return "Acceso denegado", 403

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            c.relname AS tabla,
            s.n_live_tup AS filas,
            pg_total_relation_size(c.oid) / 1024 AS size_kb
        FROM pg_class c
        JOIN pg_stat_user_tables s ON c.oid = s.relid
        ORDER BY size_kb DESC;
    """)

    datos = cur.fetchall()

    cur.close()
    conn.close()

    # calcular total
    total = sum([fila[1] for fila in datos])
    # 🔹 tabla más pesada
    tabla_mas_pesada = max(datos, key=lambda x: x[2])

    # 🔹 tabla con más filas
    tabla_mas_filas = max(datos, key=lambda x: x[1])

    # 🔹 promedio tamaño por tabla
    promedio_kb = total / len(datos) if datos else 0
    
    limite_kb = 250000  # ejemplo: 500 MB (ajústalo a tu plan real)
    porcentaje = (total / limite_kb) * 100 if limite_kb > 0 else 0
    if porcentaje < 60:
        estado = "ok"
    elif porcentaje < 85:
        estado = "warning"
    else:
        estado = "danger"

    return render_template(
        'metricas_db.html',
        datos=datos,
        total=total,
        tabla_mas_pesada=tabla_mas_pesada,
        tabla_mas_filas=tabla_mas_filas,
        promedio_kb=promedio_kb,
        porcentaje=porcentaje,
        estado=estado,
        limite_kb=limite_kb
    )
    
def obtener_examen_alumno(alumno_id, quiz_id, intento_id):
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, intento_numero, quiz_id
        FROM intentos_quiz
        WHERE alumno_id = %s AND quiz_id = %s  AND activo=TRUE
    """, (alumno_id, quiz_id))

    
    # 🔥 convertir intento_numero → intento_id real
    
    cur.execute("""
        SELECT id
        FROM intentos_quiz
        WHERE alumno_id = %s 
        AND quiz_id = %s 
        AND intento_numero = %s
        AND activo=TRUE
    """, (alumno_id, quiz_id, intento_id))

    row = cur.fetchone()
    intento_id = row[0] if row else None

    
    # 🔹 obtener respuestas reales del alumno (CLAVE)
    cur.execute("""
        SELECT pregunta_id, opcion_id
        FROM respuestas_alumno
        WHERE alumno_id = %s
        AND intento_id = %s
    """, (alumno_id, intento_id))

    respuestas = cur.fetchall()

    # convertir a diccionario
    map_respuestas = {r[0]: r[1] for r in respuestas}

    # 🔹 preguntas + opciones + respuestas
    cur.execute("""
    SELECT 
        p.id,
        p.texto AS pregunta,
        o.texto AS opcion,
        o.es_correcta,
        o.id AS opcion_id,

        (
            SELECT ra.opcion_id
            FROM respuestas_alumno ra
            WHERE ra.pregunta_id = p.id
            AND ra.alumno_id = %s
            AND ra.intento_id = %s
            AND ra.quiz_id = %s
            ORDER BY ra.id DESC
            LIMIT 1
        ) AS respuesta_alumno

    FROM preguntas p
    JOIN opciones o ON o.pregunta_id = p.id
    WHERE p.quiz_id = %s
    ORDER BY p.id, o.id
    """, (alumno_id, intento_id, quiz_id, quiz_id))

    rows = cur.fetchall()

    detalle = []
    pregunta_actual = None
    correctas = 0
    total = 0

    for row in rows:
        pid, ptexto, otexto, ocorrecta, oid, respuesta_alumno = row

        if not pregunta_actual or pregunta_actual["pregunta"] != ptexto:
            pregunta_actual = {
                "pregunta": ptexto,
                "opciones": [],
                "puntaje": 0
            }
            detalle.append(pregunta_actual)
            total += 1

        marcada = (respuesta_alumno == oid)

        if ocorrecta and marcada:
            pregunta_actual["puntaje"] = 1
            correctas += 1

        pregunta_actual["opciones"].append({
            "texto": otexto,
            "correcta": ocorrecta,
            "marcada": marcada,
            "estado": (
                "correcta" if ocorrecta and marcada else
                "incorrecta" if not ocorrecta and marcada else
                "neutral"
            )
        })

    # 🔹 alumno
    cur.execute("""
        SELECT nombre, apellido, dni, correo
        FROM alumnos
        WHERE id = %s
    """, (alumno_id,))
    alumno = cur.fetchone()

    nombre = f"{alumno[0]} {alumno[1]}"
    dni = alumno[2]
    correo = alumno[3]

    # 🔹 quiz
    cur.execute("""
        SELECT titulo
        FROM quiz
        WHERE id = %s
    """, (quiz_id,))
    titulo = cur.fetchone()[0]

    nota = round((correctas / total) * 20, 2) if total > 0 else 0

    cur.close()
    conn.close()

    return {
        "detalle": detalle,
        "nota": nota,
        "correo": correo,
        "nombre": nombre,
        "dni": dni,
        "titulo": titulo
    }
    
from flask import send_file
from io import BytesIO

@app.route('/descargar_reporte/<int:quiz_id>/<int:alumno_id>')
def descargar_reporte(quiz_id, alumno_id):

    intento = request.args.get('intento', 1)

    # 🔹 obtener examen (YA FUNCIONA)
    examen = obtener_examen_alumno(alumno_id, quiz_id, intento)

    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

    # 🔥 reutilizar tu función pero SIN enviar correo
    buffer = generar_y_enviar_reporte(
        examen["detalle"],
        examen["nota"],
        examen["correo"],
        examen["nombre"],
        alumno_id,
        examen["titulo"],
        examen["dni"],
        fecha,
        False   # 🔴 clave: no enviar email
    )

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"reporte_{alumno_id}.pdf",
        mimetype='application/pdf'
    )

#=========================================================  
# MEJORAS
#=========================================================   
@app.route('/mejoras')
@requiere_licencia
def mejoras():
    
    if not session.get('usuario'):
        return redirect('/login')  # 🔥 redirige si no está logeado
    
    conn = get_db_connection()
    cur = conn.cursor()

    if session.get('rol') == 'root':
        # 🔥 ROOT solo ve enviados
        cur.execute("""
            SELECT descripcion, usuario, fecha, estado, version, id, tipo
            FROM mejoras
            WHERE estado in ('enviado','revisado','terminado')
            ORDER BY estado asc,fecha DESC
        """)
    else:
        # 🔥 usuarios normales ven todo
        cur.execute("""
            SELECT descripcion, usuario, fecha, estado, version, id, tipo
            FROM mejoras
            ORDER BY 
                CASE 
                    WHEN estado = 'nuevo' THEN 1
                    WHEN estado = 'enviado' THEN 2
                    WHEN estado = 'revisado' THEN 3
                    WHEN estado = 'terminado' THEN 4
                END,
                fecha DESC
        """)
    mejoras = cur.fetchall()

    cur.close()
    conn.close()
    
    try:
        with open("version.txt", "r") as f:
            version_actual = f.read().strip()
    except:
        version_actual = "v?"

    return render_template('mejoras.html', mejoras=mejoras, version_actual=version_actual)

@app.route('/mejoras/nueva')
@requiere_licencia
def nueva_mejora():
    return render_template('nueva_mejora.html')

from datetime import datetime

@app.route('/mejoras/guardar', methods=['POST'])
def guardar_mejora():
    
    if session.get('rol') not in ['admin', 'profesor', 'root']:
        return "No autorizado", 403
    
    descripcion = request.form['descripcion']
    usuario = session.get('usuario', 'anonimo')
    tipo = request.form.get('tipo', 'M')
    id_mejora = request.form.get('id')
    
    conn = get_db_connection()
    cur = conn.cursor()

    if id_mejora and id_mejora.strip() != "":

        # 🔍 obtener datos actuales
        cur.execute("SELECT descripcion, usuario, estado FROM mejoras WHERE id = %s", (id_mejora,))
        descripcion_actual, usuario_creador, estado_actual = cur.fetchone()
             
        # 🔒 bloqueo total si está terminado
        if estado_actual == 'terminado':
            cur.close()
            conn.close()
            return "No se puede editar un ticket terminado", 403
        
        # 🔥 si estaba enviado → volver a nuevo
        # 🔥 regla: si el usuario edita su ticket enviado → vuelve a nuevo
        if estado_actual == 'enviado' and usuario == usuario_creador:
           nuevo_estado = 'nuevo'
        else:
            nuevo_estado = estado_actual

        # 🔥 SI ES ROOT → agrega respuesta, no reemplaza
        if session.get('rol') == 'root':

            fecha = datetime.now().strftime("%d/%m/%Y %H:%M")

            nueva_descripcion = descripcion_actual + f"""
            <br><br>
            <b>Respuesta de Desarrollo - {fecha}</b><br>
            {descripcion}
            """
            cur.execute("""
                UPDATE mejoras
                SET descripcion = %s
                WHERE id = %s
            """, (nueva_descripcion, id_mejora))

        # 🔥 SI ES USUARIO NORMAL → comportamiento actual
        else:

            # regla enviado → nuevo
            if estado_actual == 'enviado' and usuario == usuario_creador:
                nuevo_estado = 'nuevo'
            else:
                nuevo_estado = estado_actual

            cur.execute("""
                UPDATE mejoras
                SET descripcion = %s,
                    tipo = %s,
                    estado = %s
                WHERE id = %s
            """, (descripcion, tipo, nuevo_estado, id_mejora))






        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"id": id_mejora})

    else:  # 🔥 NUEVO

        cur.execute("""
            INSERT INTO mejoras (descripcion, usuario, fecha, tipo)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (descripcion, usuario, datetime.now(), tipo))

        nuevo_id = cur.fetchone()[0]

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"id": nuevo_id})

     

@app.route('/mejoras/actualizar', methods=['POST'])
def actualizar_mejora():

    usuario = session.get('usuario')
    rol = session.get('rol')

    if not usuario:
        return "No autorizado", 403

    id_mejora = request.form.get('id')
    estado = request.form.get('estado')
    version = request.form.get('version')

    conn = get_db_connection()
    cur = conn.cursor()

    # Obtener datos actuales
    cur.execute("SELECT usuario, estado FROM mejoras WHERE id = %s", (id_mejora,))
    fila = cur.fetchone()
    
    estados_validos = ['nuevo', 'enviado', 'revisado', 'terminado']

    if estado not in estados_validos:
        return "Estado inválido", 400

    if not fila:
        cur.close()
        conn.close()
        return "No existe", 404

    usuario_creador, estado_actual = fila

    # 🔒 Reglas de seguridad
    puede_editar = False

    if rol == 'root':
        # 🔒 solo puede cambiar desde 'enviado'
        if estado_actual == 'enviado' and estado in ['revisado', 'terminado']:
            puede_editar = True
    elif estado_actual == 'nuevo' and usuario_creador == usuario:
        puede_editar = True

    if not puede_editar:
        cur.close()
        conn.close()
        return "No autorizado", 403

    # Actualizar
    cur.execute("""
        UPDATE mejoras
        SET estado = %s,
            version = %s
        WHERE id = %s
    """, (estado, version, id_mejora))

    conn.commit()
    cur.close()
    conn.close()

    return '', 204

@app.route('/mejoras/eliminar', methods=['POST'])
def eliminar_mejora():

    usuario = session.get('usuario')
    rol = session.get('rol')

    if not usuario:
        return "No autorizado", 403

    id_mejora = request.form.get('id')

    conn = get_db_connection()
    cur = conn.cursor()

    # 🔍 obtener datos
    cur.execute("SELECT usuario, estado FROM mejoras WHERE id = %s", (id_mejora,))
    fila = cur.fetchone()

    if not fila:
        cur.close()
        conn.close()
        return "No existe", 404

    usuario_creador, estado_actual = fila
    
    # 🔒 bloqueo total si está terminado
    if estado_actual == 'terminado':
        cur.close()
        conn.close()
        return "No se puede eliminar un ticket terminado", 403

    # 🔒 reglas (igual que editar)
    puede_eliminar = False

    if rol == 'root':
        puede_eliminar = True
    elif estado_actual in ['nuevo', 'enviado'] and usuario_creador == usuario:
        puede_eliminar = True

    if not puede_eliminar:
        cur.close()
        conn.close()
        return "No autorizado", 403

    # 🗑 eliminar
    cur.execute("DELETE FROM mejoras WHERE id = %s", (id_mejora,))

    conn.commit()
    cur.close()
    conn.close()

    return '', 204

@app.route('/ayuda')
def ayuda():

    usuario_logeado = session.get('usuario')

    plan = None
    licencia = None

    # Valores por defecto
    uso = {
        'admins': 0,
        'profesores': 0,
        'alumnos': 0,
        'quizzes': 0,

        'disp_admins': 0,
        'disp_profesores': 0,
        'disp_alumnos': 0,
        'disp_quizzes': 0
    }

    # =========================================================
    # SOLO SI ESTÁ LOGEADO
    # =========================================================
    if usuario_logeado:

        # Empresa del usuario logeado
        cempre = session.get('cempre')

        conn = get_db_connection()
        cur = conn.cursor()

        try:

            # =================================================
            # PLAN
            # =================================================
            cur.execute("""
                SELECT nombre, precio, profesores, alumnos, quizzes
                FROM planes
                ORDER BY orden
                LIMIT 1
            """)

            plan = cur.fetchone()


            # =================================================
            # LICENCIA MÁS RECIENTE DE LA EMPRESA
            # =================================================
            if cempre:

                cur.execute("""
                    SELECT
                        codigo,
                        fecha_emision,
                        fecha_inicio,
                        fecha_expiracion,
                        estado,
                        periodicidad,
                        cantidad_periodos,
                        max_admins,
                        max_profesores,
                        max_alumnos,
                        max_quizzes,
                        fecha_activacion,
                        origen
                    FROM licencias
                    WHERE cempre = %s
                    ORDER BY fecha_emision DESC
                    LIMIT 1
                """, (cempre,))

                licencia = cur.fetchone()
                
                print("DEBUG cempre:", cempre)
                print("DEBUG licencia:", licencia)


                # =================================================
                # SI EXISTE LICENCIA, CONTAMOS LOS RECURSOS USADOS
                # =================================================
                if licencia:

                    # ---------------------------------------------
                    # ADMINISTRADORES
                    # Root NO cuenta contra la licencia
                    # ---------------------------------------------
                    cur.execute("""
                        SELECT COUNT(*)
                        FROM usuarios
                        WHERE cempre = %s
                        AND LOWER(rol) = 'admin'
                    """, (cempre,))

                    uso['admins'] = cur.fetchone()[0]


                    # ---------------------------------------------
                    # PROFESORES
                    # ---------------------------------------------
                    cur.execute("""
                        SELECT COUNT(*)
                        FROM usuarios
                        WHERE cempre = %s
                        AND LOWER(rol) = 'profesor'
                    """, (cempre,))

                    uso['profesores'] = cur.fetchone()[0]


                    # ---------------------------------------------
                    # ALUMNOS ACTIVOS
                    # ---------------------------------------------
                    cur.execute("""
                        SELECT COUNT(*)
                        FROM alumnos
                        WHERE cempre = %s
                        AND estado = 'A'
                    """, (cempre,))

                    uso['alumnos'] = cur.fetchone()[0]


                    # ---------------------------------------------
                    # QUIZZES ACTIVOS
                    # ---------------------------------------------
                    cur.execute("""
                        SELECT COUNT(*)
                        FROM quiz
                        WHERE cempre = %s
                        AND estado = 'A'
                    """, (cempre,))

                    uso['quizzes'] = cur.fetchone()[0]


                    # =================================================
                    # DISPONIBLE = MÁXIMO DE LICENCIA - UTILIZADO
                    # =================================================

                    # licencia[7]  = max_admins
                    # licencia[8]  = max_profesores
                    # licencia[9]  = max_alumnos
                    # licencia[10] = max_quizzes

                    uso['disp_admins'] = max(
                        0,
                        licencia[7] - uso['admins']
                    )

                    uso['disp_profesores'] = max(
                        0,
                        licencia[8] - uso['profesores']
                    )

                    uso['disp_alumnos'] = max(
                        0,
                        licencia[9] - uso['alumnos']
                    )

                    uso['disp_quizzes'] = max(
                        0,
                        licencia[10] - uso['quizzes']
                    )

            print("DEBUG uso:", uso)
        finally:

            cur.close()
            conn.close()


    return render_template(
        'ayuda.html',
        plan=plan,
        licencia=licencia,
        uso=uso,
        usuario_logeado=usuario_logeado
    )

@app.route('/estado_cola/<int:cola_id>')
def estado_cola(cola_id):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT estado
        FROM cola_ia
        WHERE id = %s
    """, (cola_id,))

    fila = cur.fetchone()

    cur.close()
    conn.close()

    if not fila:
        return jsonify({"estado": "no_existe"})

    return jsonify({"estado": fila[0]})


@app.route("/eliminar_pregunta/<int:pregunta_id>", methods=["POST"])
@requiere_licencia
def eliminar_pregunta(pregunta_id):

    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT COUNT(*)
        FROM respuestas_alumno
        WHERE pregunta_id = %s
    """, (pregunta_id,))

    tiene_respuestas = cur.fetchone()[0] > 0
    
    if tiene_respuestas and not request.json.get("forzar", False):
        cur.close()
        conn.close()
        return {
            "ok": False,
            "requiere_confirmacion": True,
            "mensaje": "Esta pregunta ya tiene respuestas de alumnos. Si la elimina también se eliminarán esas respuestas. ¿Desea continuar?"
        }

    # Borrar respuestas de esa pregunta
    cur.execute("""
        DELETE FROM respuestas_alumno
        WHERE pregunta_id = %s
    """, (pregunta_id,))

    # Borrar opciones
    cur.execute("""
        DELETE FROM opciones
        WHERE pregunta_id = %s
    """, (pregunta_id,))

    # Borrar pregunta
    cur.execute("""
        DELETE FROM preguntas
        WHERE id = %s
    """, (pregunta_id,))

    conn.commit()

    cur.close()
    conn.close()
    
    return {"ok": True}

@app.route('/logros_alumno')
def logros_alumno():

    if "alumno_id" not in session:
        return redirect(url_for("login_alumno"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            q.id,
            q.titulo,
            i.nota_final,
            i.fecha_fin,
            i.intento_numero
        FROM intentos_quiz i
        JOIN quiz q
            ON q.id = i.quiz_id
        WHERE i.alumno_id = %s
            AND i.nota_final IS NOT NULL
            AND i.activo=TRUE
        ORDER BY i.fecha_fin DESC;
    """, (session["alumno_id"],))

    completados = cur.fetchall()
    
    cur.execute("""
        SELECT
            id,
            titulo
        FROM quiz
        WHERE publico = TRUE
        AND estado = 'A'
        ORDER BY titulo
    """)

    publicos = cur.fetchall()
    

    cur.close()
    conn.close()

    return render_template(
        "logros_alumno.html",
        nombre=session["nombre"],
        completados=completados,
        publicos=publicos
    )

@app.route('/login_alumno', methods=['GET', 'POST'])
def login_alumno():

    if request.method == 'POST':

        dni = request.form['dni'].strip()

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT id, dni, nombre, apellido, correo, cempre
            FROM alumnos
            WHERE dni=%s
        """, (dni,))

        alumno = cur.fetchone()

        cur.close()
        conn.close()

        if alumno:

            session["alumno_id"] = alumno[0]
            session["dni"] = alumno[1]
            session["nombre"] = alumno[2]
            session["apellido"] = alumno[3]
            session["correo"] = alumno[4]
            session["cempre"] = alumno[5]

            return redirect(url_for("logros_alumno"))

        return render_template(
            "login_alumno.html",
            error="DNI no registrado"
        )

    return render_template("login_alumno.html")


def obtener_config_default():
    return {

        # General
        "multiple_intentos": True,
        "enviar_solucionario": True,
        "privado": True,

        # Evaluación
        "tiempo_minutos": None,
        "modo": "normal",

        # Gamificación
        "comodin": False

    }
    
def generar_codigo_unico_salon(cur):
    
    while True:

        codigo = ''.join(
            random.choices(
                string.ascii_uppercase + string.digits,
                k=6
            )
        )

        cur.execute(
            "SELECT 1 FROM salon_quiz WHERE codigo=%s",
            (codigo,)
        )

        if not cur.fetchone():
            return codigo
        
@app.route('/configuracion')
def configuracion():

    cempre = session.get('cempre')

    logo_actual = "/static/img/logo.png"

    if cempre:

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT logo_header
            FROM configuracion_reportes
            WHERE cempre = %s
        """, (cempre,))

        resultado = cur.fetchone()

        cur.close()
        conn.close()

        if resultado and resultado[0]:
            logo_actual = resultado[0]

    return render_template(
        'configuracion.html',
        logo_actual=logo_actual
    )

@app.route('/configuracion/logo', methods=['POST'])
def guardar_logo_configuracion():

    if 'cempre' not in session:
        return jsonify({"ok": False, "mensaje": "Sesión no válida"}), 401

    archivo = request.files.get('logo')

    if not archivo:
        return jsonify({"ok": False, "mensaje": "No se recibió ningún logo"}), 400

    cempre = session['cempre']

    carpeta = os.path.join(
        app.root_path,
        'static',
        'uploads',
        'empresas',
        str(cempre)
    )

    os.makedirs(carpeta, exist_ok=True)

    extension = os.path.splitext(archivo.filename)[1].lower()

    nombre_archivo = 'logo_header' + extension

    ruta = os.path.join(carpeta, nombre_archivo)

    archivo.save(ruta)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO configuracion_reportes
            (cempre, logo_header, fecha_actualizacion)
        VALUES
            (%s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (cempre)
        DO UPDATE SET
            logo_header = EXCLUDED.logo_header,
            fecha_actualizacion = CURRENT_TIMESTAMP
    """, (cempre, f"/static/uploads/empresas/{cempre}/{nombre_archivo}"))

    conn.commit()

    cur.close()
    conn.close()

    return jsonify({
        "ok": True,
        "mensaje": "Logo guardado correctamente"
    })
    
    
@app.route('/procesar_pago', methods=['POST'])
def procesar_pago():

    # =========================================================
    # 1. VALIDAR SESIÓN
    # =========================================================

    if 'user_id' not in session or 'cempre' not in session:
        return jsonify({
            "ok": False,
            "mensaje": "Debe iniciar sesión"
        }), 401


    data = request.get_json()
    plan_id = data.get("plan_id")

    try:
        plan_id = int(plan_id)
    except (ValueError, TypeError):
        return jsonify({
            "ok": False,
            "mensaje": "Plan inválido"
        }), 400


    user_id = session['user_id']
    cempre = session['cempre']


    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =========================================================
        # 2. OBTENER PLAN REAL DESDE BD
        # =========================================================

        cur.execute("""
            SELECT
                id,
                nombre,
                precio,
                periodicidad,
                admins,
                profesores,
                alumnos,
                quizzes
            FROM planes
            WHERE id = %s
              AND activo = TRUE
        """, (plan_id,))

        plan = cur.fetchone()

        if not plan:
            conn.rollback()

            return jsonify({
                "ok": False,
                "mensaje": "El plan no existe o está inactivo"
            }), 400


        plan_id = plan[0]
        plan_nombre = plan[1]
        precio_final = float(plan[2])
        periodicidad = plan[3]
        max_admins = plan[4]
        max_profesores = plan[5]
        max_alumnos = plan[6]
        max_quizzes = plan[7]


        # Nunca procesar Starter aquí
        if precio_final <= 0:

            conn.rollback()

            return jsonify({
                "ok": False,
                "mensaje": "Este plan no requiere pago"
            }), 400


        # =========================================================
        # 3. OBTENER USUARIO Y EMPRESA
        # =========================================================

        cur.execute("""
            SELECT
                u.correo,
                e.ruc,
                nombre,
                apellido
            FROM usuarios u
            INNER JOIN empresa e
                ON e.cempre = u.cempre
            WHERE u.id = %s
              AND u.cempre = %s
        """, (user_id, cempre))

        cliente = cur.fetchone()

        if not cliente:
            raise Exception("No se encontró el usuario o empresa")


        correo_cliente = cliente[0]
        ruc = cliente[1]
        nombre = cliente[2]
        apellido = cliente[3]


        # =========================================================
        # 4. PAGO SIMULADO APROBADO
        # =========================================================

        referencia_pago = "SIM-" + secrets.token_hex(8).upper()

        pago_aprobado = True

        if not pago_aprobado:

            conn.rollback()

            return jsonify({
                "ok": False,
                "mensaje": "Pago rechazado"
            }), 400


        # =========================================================
        # 5. DETERMINAR FA O BV
        # =========================================================

        if ruc:
            tipo_documento = "FA"
        else:
            tipo_documento = "BV"


        # =========================================================
        # 6. OBTENER IGV
        # =========================================================

        cur.execute("""
            SELECT id, porcentaje
            FROM impuesto
            WHERE activo = TRUE
            ORDER BY id
            LIMIT 1
        """)

        impuesto = cur.fetchone()

        if not impuesto:
            raise Exception("No existe un impuesto activo")


        impuesto_id = impuesto[0]
        porcentaje_impuesto = float(impuesto[1])


        # Precio mostrado ya incluye IGV
        factor_igv = 1 + (porcentaje_impuesto / 100)

        importe_bruto = round(
            precio_final / factor_igv,
            2
        )

        importe_impuesto = round(
            precio_final - importe_bruto,
            2
        )

        importe_neto = precio_final


        # =========================================================
        # 7. SOLO AHORA GENERAR NUMERACIÓN
        # =========================================================

        cur.execute("""
            SELECT ultusado
            FROM numerador
            WHERE tipodedc = %s
            FOR UPDATE
        """, (tipo_documento,))

        numerador = cur.fetchone()

        if not numerador:
            raise Exception(
                f"No existe numerador para {tipo_documento}"
            )


        numero_documento = numerador[0] + 1


        cur.execute("""
            UPDATE numerador
            SET ultusado = %s
            WHERE tipodedc = %s
        """, (
            numero_documento,
            tipo_documento
        ))


        # =========================================================
        # 8. CREAR LICENCIA
        # =========================================================

        from datetime import datetime, timedelta

        fecha_inicio = datetime.now()


        # Por ahora manejamos las periodicidades actuales
        if periodicidad == "MENSUAL":
            fecha_expiracion = fecha_inicio + timedelta(days=30)

        elif periodicidad == "TRIMESTRAL":
            fecha_expiracion = fecha_inicio + timedelta(days=90)

        elif periodicidad == "SEMESTRAL":
            fecha_expiracion = fecha_inicio + timedelta(days=180)

        elif periodicidad == "ANUAL":
            fecha_expiracion = fecha_inicio + timedelta(days=365)

        elif periodicidad == "SEMANAL":
            fecha_expiracion = fecha_inicio + timedelta(days=7)

        else:
            raise Exception(
                f"Periodicidad no válida: {periodicidad}"
            )


        codigo = "TA-" + secrets.token_hex(10).upper()


        cur.execute("""
            INSERT INTO licencias (
                codigo,
                cempre,
                plan_id,
                fecha_inicio,
                fecha_expiracion,
                estado,
                periodicidad,
                cantidad_periodos,
                max_admins,
                max_profesores,
                max_alumnos,
                max_quizzes,
                origen,
                creado_por
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                'GENERADA',
                %s,
                1,
                %s,
                %s,
                %s,
                %s,
                'AUTOMATICA',
                %s
            )
            RETURNING id
        """, (
            codigo,
            cempre,
            plan_id,
            fecha_inicio,
            fecha_expiracion,
            periodicidad,
            max_admins,
            max_profesores,
            max_alumnos,
            max_quizzes,
            user_id
        ))

        licencia_id = cur.fetchone()[0]


        # =========================================================
        # 9. CREAR VENTA PAGADA
        # =========================================================

        cur.execute("""
            INSERT INTO ventas (
                cempre,
                plan_id,
                licencia_id,
                tipo_documento,
                numero_documento,
                periodicidad,
                cantidad_periodos,
                importe_bruto,
                impuesto_id,
                porcentaje_impuesto,
                importe_impuesto,
                importe_neto,
                moneda,
                estado,
                metodo_pago,
                referencia_pago,
                correo_cliente,
                creado_por
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                1,
                %s,
                %s,
                %s,
                %s,
                %s,
                'PEN',
                'PAGADA',
                'SIMULADO',
                %s,
                %s,
                %s
            )
            RETURNING id
        """, (
            cempre,
            plan_id,
            licencia_id,
            tipo_documento,
            numero_documento,
            periodicidad,
            importe_bruto,
            impuesto_id,
            porcentaje_impuesto,
            importe_impuesto,
            importe_neto,
            referencia_pago,
            correo_cliente,
            user_id
        ))

        venta_id = cur.fetchone()[0]


        # =========================================================
        # 10. CONFIRMAR TODA LA OPERACIÓN
        # =========================================================

        conn.commit()

        # =========================================================
        # ENVIAR CORREO CON LA LICENCIA
        # =========================================================

        nombre_completo = f"{nombre} {apellido}".strip()

        correo_enviado = enviar_correo_licencia(
            correo=correo_cliente,
            nombre_completo=nombre_completo,
            plan_nombre=plan_nombre,
            codigo_licencia=codigo,
            fecha_inicio=fecha_inicio.strftime("%d/%m/%Y"),
            fecha_expiracion=fecha_expiracion.strftime("%d/%m/%Y")
        )
        print("CORREO LICENCIA ENVIADO:", correo_enviado)

        # Guardamos el resultado para pago_exito
        session['ultimo_pago'] = {
            "venta_id": venta_id,
            
            # Pago
            "plan_nombre": plan_nombre,
            "importe": precio_final,
            "tipo_documento": tipo_documento,
            "numero_documento": numero_documento,

            # Licencia
            "licencia_id": licencia_id,
            "codigo_licencia": codigo,
            "fecha_inicio": fecha_inicio.strftime("%d/%m/%Y"),
            "fecha_expiracion": fecha_expiracion.strftime("%d/%m/%Y"),
            "periodicidad": periodicidad,
            "correo": correo_cliente
        }


        return jsonify({
            "ok": True,
            "mensaje": "Pago aprobado",
            "venta_id": venta_id
        })


    except Exception as e:

        conn.rollback()

        print("========================================")
        print("ERROR PROCESANDO PAGO:")
        print(str(e))
        print("========================================")

        return jsonify({
            "ok": False,
            "mensaje": "ERROR REAL: " + str(e)
        }), 500


    finally:

        cur.close()
        conn.close()
        
@app.route('/pago_exito')
def pago_exito():

    ultimo_pago = session.get('ultimo_pago')

    if not ultimo_pago:
        return redirect(url_for('login'))

    return render_template(
        'pago_exito.html',
        pago=ultimo_pago
    )
    
    
@app.route("/activar_licencia", methods=["POST"])
def activar_licencia():

    # ==========================================
    # VALIDAR SESIÓN
    # ==========================================
    if "user_id" not in session or "cempre" not in session:
        return jsonify({
            "ok": False,
            "mensaje": "Sesión no válida"
        }), 401

    datos = request.get_json()

    if not datos or not datos.get("codigo"):
        return jsonify({
            "ok": False,
            "mensaje": "Debe ingresar una licencia"
        }), 400

    codigo = datos["codigo"].strip().upper()

    cempre = session["cempre"]
    usuario_id = session["user_id"]

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # ==========================================
        # BUSCAR LICENCIA
        # ==========================================
        cur.execute("""
            SELECT
                l.id,
                l.codigo,
                l.cempre,
                l.plan_id,
                l.fecha_inicio,
                l.fecha_expiracion,
                l.estado,
                l.periodicidad,
                l.cantidad_periodos,
                l.max_admins,
                l.max_profesores,
                l.max_alumnos,
                l.max_quizzes,
                p.nombre
            FROM licencias l
            LEFT JOIN planes p
                ON p.id = l.plan_id
            WHERE UPPER(l.codigo) = %s
        """, (codigo,))

        licencia = cur.fetchone()
 

        # ==========================================
        # NO EXISTE
        # ==========================================
        if not licencia:

            return jsonify({
                "ok": False,
                "mensaje": "El código de licencia no existe"
            }), 404

        # ==========================================
        # NO PERTENECE A ESTA EMPRESA
        # ==========================================
        if licencia[2] != cempre:

            return jsonify({
                "ok": False,
                "mensaje": "Esta licencia no pertenece a esta empresa"
            }), 403

        licencia_id = licencia[0]
        estado = licencia[6]

        # ==========================================
        # SI YA ESTÁ ACTIVA
        # ==========================================
        if estado == "ACTIVA":

            return jsonify({
                "ok": False,
                "mensaje": "Esta licencia ya se encuentra activa"
            }), 400
            
        licencia_id = licencia[0]
        estado = licencia[6]
        fecha_expiracion = licencia[5]

        # ==========================================
        # SI YA ESTÁ ACTIVA
        # ==========================================
        if estado == "ACTIVA":

            return jsonify({
                "ok": False,
                "mensaje": "Esta licencia ya se encuentra activa"
            }), 400

        # ==========================================
        # LICENCIA VENCIDA
        # ==========================================
        if estado == "VENCIDA":

            return jsonify({
                "ok": False,
                "mensaje": "Esta licencia se encuentra vencida y no puede activarse"
            }), 400

        # ==========================================
        # VALIDAR FECHA DE EXPIRACIÓN
        # ==========================================
        if fecha_expiracion < datetime.now():

            return jsonify({
                "ok": False,
                "mensaje": "Esta licencia ha expirado y no puede activarse"
            }), 400

        # ==========================================
        # SOLO SE PUEDE ACTIVAR UNA LICENCIA GENERADA
        # ==========================================
        if estado != "GENERADA":

            return jsonify({
                "ok": False,
                "mensaje": "Esta licencia no está disponible para activación"
            }), 400

        # ==========================================
        # VENCER LICENCIA ACTIVA ANTERIOR
        # ==========================================
        cur.execute("""
            UPDATE licencias
            SET
                estado = 'VENCIDA',
                fecha_modificacion = CURRENT_TIMESTAMP,
                usuario_modificador = %s
            WHERE cempre = %s
              AND estado = 'ACTIVA'
        """, (
            usuario_id,
            cempre
        ))

        # ==========================================
        # ACTIVAR NUEVA LICENCIA
        # ==========================================
        cur.execute("""
            UPDATE licencias
            SET
                estado = 'ACTIVA',
                fecha_activacion = CURRENT_TIMESTAMP,
                fecha_modificacion = CURRENT_TIMESTAMP,
                usuario_modificador = %s
            WHERE id = %s
        """, (
            usuario_id,
            licencia_id
        ))

        conn.commit()

        return jsonify({
            "ok": True,
            "mensaje": "Licencia activada correctamente",

            "plan": licencia[13],
            "codigo": licencia[1],
            "estado": "ACTIVA",

            "fecha_inicio": licencia[4].strftime("%d/%m/%Y"),
            "fecha_expiracion": licencia[5].strftime("%d/%m/%Y"),

            "periodicidad": licencia[7],
            "cantidad_periodos": licencia[8],

            "max_admins": licencia[9],
            "max_profesores": licencia[10],
            "max_alumnos": licencia[11],
            "max_quizzes": licencia[12]
        })
        
    except Exception as e:

        conn.rollback()

        print("❌ ERROR ACTIVANDO LICENCIA:", str(e))

        return jsonify({
            "ok": False,
            "mensaje": str(e)
        }), 500

    finally:

        cur.close()
        conn.close()
        
@app.route('/validar_licencia', methods=['POST'])
def validar_licencia():

    # =========================================================
    # VALIDAR SESIÓN
    # =========================================================
    if 'user_id' not in session or 'cempre' not in session:
        return jsonify({
            'ok': False,
            'mensaje': 'Sesión no válida.'
        }), 401

    data = request.get_json()

    if not data:
        return jsonify({
            'ok': False,
            'mensaje': 'No se recibió información.'
        }), 400

    codigo = data.get('codigo', '').strip().upper()

    if not codigo:
        return jsonify({
            'ok': False,
            'mensaje': 'Ingrese el código de licencia.'
        }), 400

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =====================================================
        # BUSCAR LA LICENCIA SOLO POR EL CÓDIGO INGRESADO
        # =====================================================
        cur.execute("""
            SELECT
                id,
                codigo,
                cempre,
                fecha_inicio,
                fecha_expiracion,
                estado,
                max_admins,
                max_profesores,
                max_alumnos,
                max_quizzes
            FROM licencias
            WHERE UPPER(codigo) = %s
        """, (codigo,))

        licencia = cur.fetchone()

        if not licencia:
            return jsonify({
                'ok': False,
                'mensaje': 'El código de licencia no existe.'
            }), 404

        # =====================================================
        # POR AHORA SOLO VALIDAMOS
        # =====================================================
        return jsonify({
            'ok': True,
            'mensaje': 'Licencia encontrada correctamente.',
            'codigo': licencia[1],
            'estado': licencia[5]
        })

    finally:
        cur.close()
        conn.close()
        
@app.route('/empresas')
def empresas():

    if 'user_id' not in session:
        return redirect(url_for('login'))

    if session.get('rol') != 'root':
        return redirect(url_for('home'))

    # =========================================================
    # FILTRO DE EMPRESAS
    # =========================================================

    mostrar_inactivos = request.args.get('inactivos') == '1'

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        if mostrar_inactivos:

            cur.execute("""
                SELECT
                    cempre,
                    dempre,
                    ruc,
                    licencia,
                    estado,
                    fcreacion,
                    fecha_modificacion,
                    nombre_comercial,
                    direccion_domicilio
                FROM empresa
                WHERE estado = FALSE
                ORDER BY dempre
            """)

        else:

            cur.execute("""
                SELECT
                    cempre,
                    dempre,
                    ruc,
                    licencia,
                    estado,
                    fcreacion,
                    fecha_modificacion,
                    nombre_comercial,
                    direccion_domicilio
                FROM empresa
                WHERE estado = TRUE
                ORDER BY dempre
            """)

        empresas = cur.fetchall()

    finally:

        cur.close()
        conn.close()

    return render_template(
        'empresas.html',
        empresas=empresas,
        mostrar_inactivos=mostrar_inactivos
    )
    
@app.route('/guardar_empresa', methods=['POST'])
def guardar_empresa():

    # =========================================================
    # SOLO ROOT
    # =========================================================
    if 'user_id' not in session:
        return jsonify({
            'ok': False,
            'mensaje': 'Sesión no válida.'
        }), 401

    if session.get('rol') != 'root':
        return jsonify({
            'ok': False,
            'mensaje': 'No tiene autorización para administrar empresas.'
        }), 403

    data = request.get_json()

    if not data:
        return jsonify({
            'ok': False,
            'mensaje': 'No se recibieron datos.'
        }), 400

    # =========================================================
    # DATOS BÁSICOS
    # =========================================================

    empresa_id = data.get('empresa_id')

    try:
        empresa_id = int(empresa_id) if empresa_id else None
    except (ValueError, TypeError):

        return jsonify({
            'ok': False,
            'mensaje': 'Empresa inválida.'
        }), 400

    estado_recibido = data.get('estado', True)

    if isinstance(estado_recibido, bool):
        estado = estado_recibido
    else:
        estado = str(estado_recibido).strip().lower() == 'true'

    # =========================================================
    # SI HAY EMPRESA, PRIMERO DETERMINAMOS
    # SI LA OPERACIÓN ES DESACTIVAR
    # =========================================================

    if empresa_id:

        conn = get_db_connection()
        cur = conn.cursor()

        try:

            cur.execute("""
                SELECT
                    estado,
                    dempre
                FROM empresa
                WHERE cempre = %s
            """, (empresa_id,))

            empresa_actual = cur.fetchone()

            if not empresa_actual:

                return jsonify({
                    'ok': False,
                    'mensaje': 'La empresa no existe.'
                }), 404

            estado_actual = bool(empresa_actual[0])
            nombre_empresa = empresa_actual[1]

            estado_nuevo = estado

            # =================================================
            # DESACTIVAR
            # AQUÍ SALIMOS DIRECTAMENTE
            # NO VALIDAMOS RUC
            # NO VALIDAMOS NOMBRE COMERCIAL
            # NO VALIDAMOS DIRECCIÓN
            # =================================================

            if estado_actual and not estado_nuevo:

                cur.execute("""
                    UPDATE empresa
                    SET
                        estado = FALSE,
                        fecha_modificacion = CURRENT_TIMESTAMP
                    WHERE cempre = %s
                """, (empresa_id,))

                conn.commit()

                return jsonify({
                    'ok': True,
                    'mensaje': f'La empresa "{nombre_empresa}" fue desactivada correctamente.',
                    'cempre': empresa_id
                })

        except Exception as e:

            conn.rollback()

            print("❌ ERROR DESACTIVANDO EMPRESA:", str(e))

            return jsonify({
                'ok': False,
                'mensaje': 'Error al desactivar la empresa.'
            }), 500

        finally:

            cur.close()
            conn.close()

    # =========================================================
    # A PARTIR DE AQUÍ:
    # CREAR / EDITAR / REACTIVAR
    #
    # Estas operaciones SÍ VALIDAN LOS DATOS
    # =========================================================

    razon_social = data.get('razon_social', '').strip()
    ruc = data.get('ruc', '').strip()
    nombre_comercial = data.get('nombre_comercial', '').strip()
    direccion = data.get('direccion', '').strip()

    # =========================================================
    # VALIDAR CAMPOS OBLIGATORIOS
    # =========================================================

    if not razon_social:
        return jsonify({
            'ok': False,
            'campo': 'razon_social',
            'mensaje': 'Ingrese el nombre o razón social.'
        }), 400

    if not ruc:
        return jsonify({
            'ok': False,
            'campo': 'ruc',
            'mensaje': 'Ingrese el RUC.'
        }), 400

    if not ruc.isdigit() or len(ruc) != 11:
        return jsonify({
            'ok': False,
            'campo': 'ruc',
            'mensaje': 'El RUC debe contener exactamente 11 dígitos.'
        }), 400

    if not nombre_comercial:
        return jsonify({
            'ok': False,
            'campo': 'nombre_comercial',
            'mensaje': 'Ingrese el nombre comercial.'
        }), 400

    if not direccion:
        return jsonify({
            'ok': False,
            'campo': 'direccion',
            'mensaje': 'Ingrese la dirección del domicilio.'
        }), 400

    # =========================================================
    # CONEXIÓN
    # =========================================================

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =====================================================
        # VALIDAR RUC
        # =====================================================

        if empresa_id:

            cur.execute("""
                SELECT cempre
                FROM empresa
                WHERE ruc = %s
                  AND cempre <> %s
            """, (ruc, empresa_id))

        else:

            cur.execute("""
                SELECT cempre
                FROM empresa
                WHERE ruc = %s
            """, (ruc,))

        if cur.fetchone():

            return jsonify({
                'ok': False,
                'campo': 'ruc',
                'mensaje': 'El RUC ya está registrado en otra empresa.'
            }), 409

        # =====================================================
        # EDITAR / REACTIVAR
        # =====================================================

        if empresa_id:

            # -----------------------------------------------
            # Verificar estado actual
            # -----------------------------------------------
            cur.execute("""
                SELECT
                    estado,
                    dempre
                FROM empresa
                WHERE cempre = %s
            """, (empresa_id,))

            empresa_actual = cur.fetchone()

            if not empresa_actual:

                return jsonify({
                    'ok': False,
                    'mensaje': 'La empresa no existe.'
                }), 404

            estado_actual = bool(empresa_actual[0])
            estado_nuevo = bool(estado)

            cur.execute("""
                UPDATE empresa
                SET
                    dempre = %s,
                    ruc = %s,
                    nombre_comercial = %s,
                    direccion_domicilio = %s,
                    estado = %s,
                    fecha_modificacion = CURRENT_TIMESTAMP
                WHERE cempre = %s
            """, (
                razon_social,
                ruc,
                nombre_comercial,
                direccion,
                estado_nuevo,
                empresa_id
            ))

            if not estado_actual and estado_nuevo:

                mensaje = 'Empresa reactivada correctamente.'

            else:

                mensaje = 'Empresa actualizada correctamente.'

            cempre = empresa_id

        # =====================================================
        # CREAR
        # =====================================================

        else:

            cur.execute("""
                INSERT INTO empresa (
                    dempre,
                    ruc,
                    nombre_comercial,
                    direccion_domicilio,
                    licencia,
                    estado,
                    fcreacion
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    FALSE,
                    TRUE,
                    CURRENT_TIMESTAMP
                )
                RETURNING cempre
            """, (
                razon_social,
                ruc,
                nombre_comercial,
                direccion
            ))

            cempre = cur.fetchone()[0]

            mensaje = 'Empresa creada correctamente.'

        # =====================================================
        # CONFIRMAR
        # =====================================================

        conn.commit()

        return jsonify({
            'ok': True,
            'mensaje': mensaje,
            'cempre': cempre
        })

    except Exception as e:

        conn.rollback()

        print("❌ ERROR GUARDANDO EMPRESA:", str(e))

        return jsonify({
            'ok': False,
            'mensaje': 'Error al guardar la empresa.'
        }), 500

    finally:

        cur.close()
        conn.close()
        
@app.route('/desactivar_empresa/<int:empresa_id>', methods=['POST'])
def desactivar_empresa(empresa_id):

    # =========================================================
    # SOLO ROOT
    # =========================================================
    if 'user_id' not in session:
        return jsonify({
            'ok': False,
            'mensaje': 'Sesión no válida.'
        }), 401

    if session.get('rol') != 'root':
        return jsonify({
            'ok': False,
            'mensaje': 'No tiene autorización para desactivar empresas.'
        }), 403

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =====================================================
        # VERIFICAR EMPRESA
        # =====================================================
        cur.execute("""
            SELECT
                cempre,
                dempre,
                estado
            FROM empresa
            WHERE cempre = %s
        """, (empresa_id,))

        empresa = cur.fetchone()

        if not empresa:
            return jsonify({
                'ok': False,
                'mensaje': 'La empresa no existe.'
            }), 404

        if not empresa[2]:
            return jsonify({
                'ok': False,
                'mensaje': 'La empresa ya está inactiva.'
            }), 400

        # =====================================================
        # DESACTIVAR
        # =====================================================
        cur.execute("""
            UPDATE empresa
            SET
                estado = FALSE,
                fecha_modificacion = CURRENT_TIMESTAMP
            WHERE cempre = %s
        """, (empresa_id,))

        conn.commit()

        return jsonify({
            'ok': True,
            'mensaje': f'La empresa "{empresa[1]}" fue desactivada correctamente.'
        })

    except Exception as e:

        conn.rollback()

        print("❌ ERROR DESACTIVANDO EMPRESA:", str(e))

        return jsonify({
            'ok': False,
            'mensaje': 'Error al desactivar la empresa.'
        }), 500

    finally:

        cur.close()
        conn.close()
               
@app.route('/obtener_empresa/<int:empresa_id>')
def obtener_empresa(empresa_id):

    # =========================================================
    # SOLO ROOT
    # =========================================================
    if 'user_id' not in session:
        return jsonify({
            'ok': False,
            'mensaje': 'Sesión no válida.'
        }), 401

    if session.get('rol') != 'root':
        return jsonify({
            'ok': False,
            'mensaje': 'No autorizado.'
        }), 403

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                cempre,
                dempre,
                ruc,
                nombre_comercial,
                direccion_domicilio,
                estado
            FROM empresa
            WHERE cempre = %s
        """, (empresa_id,))

        empresa = cur.fetchone()

        if not empresa:
            return jsonify({
                'ok': False,
                'mensaje': 'La empresa no existe.'
            }), 404

        return jsonify({
            'ok': True,
            'empresa': {
                'id': empresa[0],
                'razon_social': empresa[1],
                'ruc': empresa[2] or '',
                'nombre_comercial': empresa[3] or '',
                'direccion': empresa[4] or '',
                'estado': empresa[5]
            }
        })

    except Exception as e:

        print("❌ ERROR OBTENIENDO EMPRESA:", str(e))

        return jsonify({
            'ok': False,
            'mensaje': 'Error al consultar la empresa.'
        }), 500

    finally:

        cur.close()
        conn.close()
        

    
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)